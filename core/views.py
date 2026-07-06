from decimal import Decimal
from datetime import datetime, timedelta
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from .forms import ProfileImageUpdateForm
from .decorators import get_user_role, role_required
from django.core.files.base import ContentFile
from .models import *

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Prefetch, Q
from django.shortcuts import get_object_or_404, redirect, render

from .decorators import get_user_role, role_required
from .forms import *
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum,Count
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

import base64
from io import BytesIO
from urllib.parse import quote

import qrcode

from decimal import Decimal
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.utils import timezone

from .forms import DealerInvoicePaymentForm

from django.http import JsonResponse
from decimal import Decimal, InvalidOperation
import json
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Sum
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from .decorators import role_required
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import get_object_or_404, redirect, render
from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import redirect, render
from django.utils import timezone
from urllib.parse import quote
from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone

import math
from decimal import Decimal
from django.db.models import Sum
from django.http import JsonResponse
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
import calendar as py_calendar
from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Sum
from django.utils import timezone

from .models import EmployeeAttendance
import base64
import re
from io import BytesIO

import qrcode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.safestring import mark_safe

from .models import Product, ProductStickerSetting
import base64
import uuid

from django.core.files.base import ContentFile
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Q
from decimal import Decimal
from django.db import transaction, IntegrityError
from django.db.models import Sum
from django.utils import timezone
from django.core.exceptions import ValidationError
from io import BytesIO
from datetime import datetime
from xml.sax.saxutils import escape

from django.http import HttpResponse
from django.db.models import Q, Sum
from django.utils import timezone
from django.utils.dateparse import parse_date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import FarmerData, Dealer
from PIL import Image, ImageDraw, ImageFont, ImageOps
from django.core.files.base import ContentFile
from io import BytesIO
from decimal import Decimal



User = get_user_model()

ROLE_LEVELS = {
    "DEALER": 0,
    "DEVELOPMENT_OFFICER": 1,
    "SALES_OFFICER_JUNIOR": 2,
    "SALES_OFFICER_SENIOR": 3,
    "WAREHOUSE_MANAGER": 3,
    "PRODUCTION_INCHARGE": 3,
    "ASM": 4,
    "REGIONAL_MANAGER": 5,
    "STATE_HEAD": 6,
    "ACCOUNTANT": 6,
    "HR": 7,
    "ADMIN": 8,
}


def get_role_level(role):
    return ROLE_LEVELS.get(role, 0)

def can_edit_employee_profile(editor, employee_profile):
    if editor.is_superuser:
        return True

    editor_profile = getattr(editor, "profile", None)

    if not editor_profile:
        return False

    # Nobody can edit their own full employee profile from employee edit page
    if employee_profile.user == editor:
        return False

    # Dealer profile should not be edited from employee section
    if employee_profile.role == "DEALER":
        return False

    # Only Admin and Accountant can edit employee profiles
    return editor_profile.role in ["ADMIN","HR","ACCOUNTANT"]


def find_assigned_asm_for_sales_officer(sales_officer):
    """
    Dealer will automatically go to the ASM assigned as Sales Officer's manager.
    If direct manager is not ASM, it checks one level above also.
    """

    profile = getattr(sales_officer, "profile", None)

    if not profile or not profile.manager:
        return None

    manager = profile.manager
    manager_profile = getattr(manager, "profile", None)

    if manager_profile and manager_profile.role == "ASM":
        return manager

    if manager_profile and manager_profile.manager:
        second_manager = manager_profile.manager
        second_manager_profile = getattr(second_manager, "profile", None)

        if second_manager_profile and second_manager_profile.role == "ASM":
            return second_manager

    return None


def safe_reverse(url_name):
    try:
        return reverse(url_name)
    except NoReverseMatch:
        return "#"

def clean_phone_for_link(phone):
    if not phone:
        return ""

    phone = str(phone)
    digits = "".join(ch for ch in phone if ch.isdigit())

    if len(digits) == 10:
        digits = "91" + digits

    return digits


def get_dealer_contact_context(dealer):
    sales_officer = dealer.created_by_sales_officer
    asm = dealer.concerned_asm

    sales_profile = getattr(sales_officer, "profile", None) if sales_officer else None
    asm_profile = getattr(asm, "profile", None) if asm else None

    sales_phone = clean_phone_for_link(sales_profile.phone if sales_profile else "")
    asm_phone = clean_phone_for_link(asm_profile.phone if asm_profile else "")

    return {
        "sales_officer": sales_officer,
        "sales_profile": sales_profile,
        "sales_phone": sales_phone,
        "asm": asm,
        "asm_profile": asm_profile,
        "asm_phone": asm_phone,
    }


def get_or_create_profile(user):
    profile = getattr(user, "profile", None)

    if profile:
        return profile

    return UserProfile.objects.create(
        user=user,
        role="ADMIN" if user.is_superuser else "DEALER"
    )


def get_dashboard_links(user):
    role = get_user_role(user)

    common = [
        {
            "title": "My Profile",
            "subtitle": "View your full profile",
            "url": safe_reverse("my_profile"),
            "icon": "profile",
        },
    ]

    role_links = {
        "ADMIN": [
            {
                "title": "Create Dealer",
                "subtitle": "Add new dealer",
                "url": safe_reverse("dealer_create"),
                "icon": "dealer",
            },
            {
                "title": "Dealer Approvals",
                "subtitle": "View approval requests",
                "url": safe_reverse("dealer_approval_list"),
                "icon": "approval",
            },
            {
                "title": "Warehouses",
                "subtitle": "Manage warehouses",
                "url": safe_reverse("warehouse_list"),
                "icon": "warehouse",
            },
            {
                "title": "Create Warehouse",
                "subtitle": "Add new warehouse",
                "url": safe_reverse("warehouse_create"),
                "icon": "warehouse",
            },
        ],

        "SALES_OFFICER_SENIOR": [
            {
                "title": "Create Dealer",
                "subtitle": "Create dealer and send to ASM",
                "url": safe_reverse("dealer_create"),
                "icon": "dealer",
            },
            {
                "title": "Farmer Meet Approvals",
                "subtitle": "Approve farmer meeting requests",
                "url": safe_reverse("farmer_meet_list"),
                "icon": "farmers",
            },
        ],

        "SALES_OFFICER_JUNIOR": [
            {
                "title": "Create Dealer",
                "subtitle": "Create dealer and send to ASM",
                "url": safe_reverse("dealer_create"),
                "icon": "dealer",
            },
            {
                "title": "Farmer Meet Approvals",
                "subtitle": "Approve farmer meeting requests",
                "url": safe_reverse("farmer_meet_list"),
                "icon": "farmers",
            },
        ],

        "ASM": [
            {
                "title": "Dealer Approvals",
                "subtitle": "Approve or forward dealers",
                "url": safe_reverse("dealer_approval_list"),
                "icon": "approval",
            },
            {
                "title": "Farmer Meet Approvals",
                "subtitle": "Approve farmer meeting requests",
                "url": safe_reverse("farmer_meet_list"),
                "icon": "farmers",
            },
        ],

        "REGIONAL_MANAGER": [
            {
                "title": "Dealer Approvals",
                "subtitle": "Approve or forward dealers",
                "url": safe_reverse("dealer_approval_list"),
                "icon": "approval",
            },
            {
                "title": "Farmer Meet Approvals",
                "subtitle": "Approve farmer meeting requests",
                "url": safe_reverse("farmer_meet_list"),
                "icon": "farmers",
            },
        ],

        "STATE_HEAD": [
            {
                "title": "Dealer Approvals",
                "subtitle": "Final approval for dealers",
                "url": safe_reverse("dealer_approval_list"),
                "icon": "approval",
            },
            {
                "title": "Farmer Meet Approvals",
                "subtitle": "Final approval requests",
                "url": safe_reverse("farmer_meet_list"),
                "icon": "farmers",
            },
        ],

        "WAREHOUSE_MANAGER": [
            {
                "title": "My Warehouse",
                "subtitle": "View warehouse details",
                "url": safe_reverse("warehouse_list"),
                "icon": "warehouse",
            },
        ],

        "DEVELOPMENT_OFFICER": [
            {
                "title": "Create Farmer Meet",
                "subtitle": "Request farmer meeting approval",
                "url": safe_reverse("farmer_meet_create"),
                "icon": "farmers",
            },
            {
                "title": "My Farmer Meets",
                "subtitle": "Track farmer meeting status",
                "url": safe_reverse("farmer_meet_list"),
                "icon": "reports",
            },
        ],

        "DEALER": [
            {
                "title": "My Dealer Profile",
                "subtitle": "View dealer details",
                "url": safe_reverse("my_profile"),
                "icon": "dealer",
            },
        ],
    }

    return common + role_links.get(role, [])


User = get_user_model()

@login_required
def employee_list(request):
    user = request.user
    profile = getattr(user, "profile", None)
    role = profile.role if profile else ""

    search = request.GET.get("q", "").strip()
    role_filter = request.GET.get("role", "").strip()
    current_year = timezone.localdate().year

    def is_admin_or_accountant():
        return user.is_superuser or role in ["ADMIN","HR", "ACCOUNTANT"]

    def get_user_name(u):
        if not u:
            return "-"
        return u.get_full_name() or u.email or u.username

    def get_all_subordinate_user_ids(manager_user):
        """
        Gets all lower employees under a manager.
        Example:
        ASM can see Sales Officers under him.
        RSM can see ASM and Sales Officers under those ASMs.
        """
        collected = set()
        current_manager_ids = {manager_user.id}

        for _ in range(6):
            direct_ids = set(
                UserProfile.objects.filter(
                    manager_id__in=current_manager_ids
                ).exclude(
                    role="DEALER"
                ).values_list("user_id", flat=True)
            )

            new_ids = direct_ids - collected

            if not new_ids:
                break

            collected.update(new_ids)
            current_manager_ids = new_ids

        return collected

    employee_profiles = UserProfile.objects.select_related(
        "user",
        "manager",
        "manager__profile",
        "state",
        "district",
    ).exclude(
        role="DEALER"
    ).order_by(
        "role",
        "user__first_name",
        "user__username"
    )

    # -----------------------------
    # ROLE BASED VISIBILITY
    # -----------------------------
    if user.is_superuser or role in ["ADMIN", "ACCOUNTANT", "HR"]:
        # Admin / Accountant can see all employees.
        pass

    elif role == "STATE_HEAD":
        subordinate_ids = get_all_subordinate_user_ids(user)
        visible_ids = subordinate_ids | {user.id}

        if profile and profile.state_id:
            employee_profiles = employee_profiles.filter(
                Q(user_id__in=visible_ids) |
                Q(state_id=profile.state_id)
            )
        else:
            employee_profiles = employee_profiles.filter(user_id__in=visible_ids)

    elif role == "REGIONAL_MANAGER":
        subordinate_ids = get_all_subordinate_user_ids(user)
        visible_ids = subordinate_ids | {user.id}
        employee_profiles = employee_profiles.filter(user_id__in=visible_ids)

    elif role == "ASM":
        # ASM can see himself and all employees whose manager is ASM.
        subordinate_ids = get_all_subordinate_user_ids(user)
        visible_ids = subordinate_ids | {user.id}
        employee_profiles = employee_profiles.filter(user_id__in=visible_ids)

    elif role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR", "DEVELOPMENT_OFFICER"]:
        # Lower employee can see own profile and manager profile.
        visible_ids = {user.id}

        if profile and profile.manager_id:
            visible_ids.add(profile.manager_id)

        employee_profiles = employee_profiles.filter(user_id__in=visible_ids)

    else:
        employee_profiles = employee_profiles.filter(user=user)

    # -----------------------------
    # SEARCH / FILTER
    # -----------------------------
    if search:
        employee_profiles = employee_profiles.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search) |
            Q(phone__icontains=search) |
            Q(manager__first_name__icontains=search) |
            Q(manager__last_name__icontains=search) |
            Q(manager__username__icontains=search) |
            Q(manager__email__icontains=search)
        )

    if role_filter:
        employee_profiles = employee_profiles.filter(role=role_filter)

    employee_rows = []

    for emp_profile in employee_profiles:
        target = EmployeeYearlyTarget.objects.filter(
            employee=emp_profile.user,
            year=current_year
        ).first()

        if target:
            percentage = target.achieved_percentage or 0
            target_assigned = True

            if percentage >= 100:
                target_status = "Achieved"
                target_status_class = "achieved"
            else:
                target_status = "Not Achieved"
                target_status_class = "pending"
        else:
            percentage = 0
            target_assigned = False
            target_status = "No Target Assigned"
            target_status_class = "no-target"

        manager_user = emp_profile.manager

        manager_profile = None
        if manager_user:
            manager_profile = getattr(manager_user, "profile", None)

        employee_rows.append({
            "profile": emp_profile,
            "manager_name": get_user_name(manager_user),
            "manager_role": manager_profile.get_role_display() if manager_profile else "-",
            "manager_phone": manager_profile.phone if manager_profile and manager_profile.phone else "-",
            "target_assigned": target_assigned,
            "percentage": percentage,
            "target_status": target_status,
            "target_status_class": target_status_class,
            "can_edit": is_admin_or_accountant(),
        })

    return render(
        request,
        "core/employee_list.html",
        {
            "employee_rows": employee_rows,
            "role_choices": UserProfile.ROLE_CHOICES,
            "search": search,
            "role_filter": role_filter,
            "current_year": current_year,
        }
    )



User = get_user_model()


TARGET_LEAF_ROLES = [
    "SALES_OFFICER_SENIOR",
    "SALES_OFFICER_JUNIOR",
]

AUTO_TARGET_MANAGER_ROLES = [
    "ASM",
    "REGIONAL_MANAGER",
    "STATE_HEAD",
]

TARGET_COUNTED_ROLES = [
    "SALES_OFFICER_SENIOR",
    "SALES_OFFICER_JUNIOR",
    "ASM",
    "REGIONAL_MANAGER",
    "STATE_HEAD",
]


def decimal_value(value):
    if value in [None, ""]:
        return Decimal("0.00")
    return Decimal(str(value))


def get_employee_year_target(employee_user, year):
    return EmployeeYearlyTarget.objects.filter(
        employee=employee_user,
        year=year
    ).first()


def recalculate_single_manager_target(manager_user, year):
    if not manager_user:
        return

    manager_profile = getattr(manager_user, "profile", None)

    if not manager_profile:
        return

    if manager_profile.role not in AUTO_TARGET_MANAGER_ROLES:
        return

    child_profiles = UserProfile.objects.filter(
        manager=manager_user,
        role__in=TARGET_COUNTED_ROLES,
        user__is_active=True,
    ).select_related("user")

    base_team_target = Decimal("0.00")
    base_team_achieved = Decimal("0.00")
    has_child_target = False

    for child_profile in child_profiles:
        child_target = get_employee_year_target(child_profile.user, year)

        if child_target:
            has_child_target = True
            base_team_target += child_target.target_amount or Decimal("0.00")
            base_team_achieved += child_target.achieved_amount or Decimal("0.00")

    if not has_child_target or base_team_target <= 0:
        EmployeeYearlyTarget.objects.filter(
            employee=manager_user,
            year=year
        ).delete()
        return

    extra_percentage = manager_profile.manager_target_extra_percentage or Decimal("0.00")

    extra_amount = (base_team_target * extra_percentage) / Decimal("100.00")
    final_manager_target = base_team_target + extra_amount

    EmployeeYearlyTarget.objects.update_or_create(
        employee=manager_user,
        year=year,
        defaults={
            "target_amount": final_manager_target,
            "achieved_amount": base_team_achieved,
        }
    )


def recalculate_target_chain_from_manager(manager_user, year):
    current_manager = manager_user
    visited = set()

    while current_manager:
        if current_manager.id in visited:
            break

        visited.add(current_manager.id)

        recalculate_single_manager_target(current_manager, year)

        current_profile = getattr(current_manager, "profile", None)

        if not current_profile:
            break

        current_manager = current_profile.manager


def save_employee_target_and_recalculate(
    employee_user,
    year,
    target_amount,
    achieved_amount,
    previous_manager=None
):
    profile = employee_user.profile
    role = profile.role

    target_amount = decimal_value(target_amount)
    achieved_amount = decimal_value(achieved_amount)

    if role in TARGET_LEAF_ROLES:
        EmployeeYearlyTarget.objects.update_or_create(
            employee=employee_user,
            year=year,
            defaults={
                "target_amount": target_amount,
                "achieved_amount": achieved_amount,
            }
        )

    elif role in AUTO_TARGET_MANAGER_ROLES:
        recalculate_single_manager_target(employee_user, year)

    else:
        EmployeeYearlyTarget.objects.update_or_create(
            employee=employee_user,
            year=year,
            defaults={
                "target_amount": target_amount,
                "achieved_amount": achieved_amount,
            }
        )

    current_manager = profile.manager

    if current_manager:
        recalculate_target_chain_from_manager(current_manager, year)

    if previous_manager and previous_manager != current_manager:
        recalculate_target_chain_from_manager(previous_manager, year)


@login_required
@role_required(["ADMIN", "HR", "ACCOUNTANT"])
@transaction.atomic
def employee_create(request):
    if request.method == "POST":
        form = EmployeeForm(
            request.POST,
            request.FILES,
            request_user=request.user
        )

        if form.is_valid():
            email = form.cleaned_data["email"].lower()

            employee_user = User.objects.create_user(
                username=email,
                email=email,
                password=form.cleaned_data["password"],
                first_name=form.cleaned_data["first_name"],
                last_name=form.cleaned_data["last_name"],
                is_active=True,
            )

            profile = form.save(commit=False)
            profile.user = employee_user
            profile.save()

            target_year = form.cleaned_data.get("target_year") or timezone.now().year
            target_amount = form.cleaned_data.get("target_amount") or Decimal("0.00")
            achieved_amount = form.cleaned_data.get("achieved_amount") or Decimal("0.00")

            save_employee_target_and_recalculate(
                employee_user=employee_user,
                year=target_year,
                target_amount=target_amount,
                achieved_amount=achieved_amount,
            )

            messages.success(request, "Employee created successfully.")
            return redirect("employee_list")

    else:
        form = EmployeeForm(request_user=request.user)

    return render(
        request,
        "core/employee_form.html",
        {
            "form": form,
            "title": "Create Employee",
            "button_text": "Create Employee",
        }
    )


@login_required
@transaction.atomic
def employee_edit(request, pk):
    profile = get_object_or_404(UserProfile, pk=pk)

    if not can_edit_employee_profile(request.user, profile):
        messages.error(request, "You do not have permission to edit this employee profile.")
        return redirect("employee_list")

    employee_user = profile.user
    previous_manager = profile.manager

    if request.method == "POST":
        form = EmployeeForm(
            request.POST,
            request.FILES,
            instance=profile,
            request_user=request.user
        )

        if form.is_valid():
            email = form.cleaned_data["email"].lower()

            employee_user.username = email
            employee_user.email = email
            employee_user.first_name = form.cleaned_data["first_name"]
            employee_user.last_name = form.cleaned_data["last_name"]

            password = form.cleaned_data.get("password")
            if password:
                employee_user.set_password(password)

            employee_user.save()

            profile = form.save(commit=False)
            profile.user = employee_user
            profile.save()

            target_year = form.cleaned_data.get("target_year") or timezone.now().year
            target_amount = form.cleaned_data.get("target_amount") or Decimal("0.00")
            achieved_amount = form.cleaned_data.get("achieved_amount") or Decimal("0.00")

            save_employee_target_and_recalculate(
                employee_user=employee_user,
                year=target_year,
                target_amount=target_amount,
                achieved_amount=achieved_amount,
                previous_manager=previous_manager,
            )

            messages.success(request, "Employee updated successfully.")
            return redirect("employee_list")

    else:
        form = EmployeeForm(
            instance=profile,
            request_user=request.user
        )

    return render(
        request,
        "core/employee_form.html",
        {
            "form": form,
            "title": "Edit Employee",
            "button_text": "Update Employee",
        }
    )


def get_target_context(user):
    role = get_user_role(user)

    context = {
        "show": False,
        "title": "",
        "target_amount": Decimal("0.00"),
        "achieved_amount": Decimal("0.00"),
        "balance_amount": Decimal("0.00"),
        "percentage": 0,
    }

    if role == "DEALER":
        dealer = Dealer.objects.filter(user=user, approval_status="APPROVED", is_active=True).first()

        if dealer:
            context.update({
                "show": True,
                "title": "Dealer Yearly Target",
                "target_amount": dealer.yearly_target_amount,
                "achieved_amount": dealer.achieved_amount,
                "balance_amount": dealer.target_balance_amount,
                "percentage": dealer.target_achieved_percentage,
            })

        return context

    current_year = timezone.now().year

    target = EmployeeYearlyTarget.objects.filter(
        employee=user,
        year=current_year
    ).first()

    if target:
        context.update({
            "show": True,
            "title": f"Employee Target - {current_year}",
            "target_amount": target.target_amount,
            "achieved_amount": target.achieved_amount,
            "balance_amount": target.balance_amount,
            "percentage": target.achieved_percentage,
        })

    return context


@login_required
def dashboard(request):
    user = request.user
    profile = getattr(user, "profile", None)
    role = profile.role if profile else ""

    dealer = None
    if role == "DEALER":
        dealer = Dealer.objects.filter(user=user).first()

    def safe_reverse(name):
        try:
            return reverse(name)
        except Exception:
            return "#"

    def add_link(title, subtitle, url_name, icon="menu", count=0, urgent=False):
        return {
            "title": title,
            "subtitle": subtitle,
            "url": safe_reverse(url_name),
            "icon": icon,
            "count": count,
            "urgent": urgent,
        }

    def count_orders(status, user_field=None):
        qs = DealerOrder.objects.filter(status=status)

        if user_field:
            qs = qs.filter(**{user_field: user})

        return qs.count()

    def count_farmer_meets(status, user_field=None):
        qs = FarmerMeetRequest.objects.filter(approval_status=status)

        if user_field:
            qs = qs.filter(**{user_field: user})

        return qs.count()

    def get_subordinate_user_ids(manager_user):
        return list(
            UserProfile.objects.filter(
                manager=manager_user
            ).values_list("user_id", flat=True)
        )

    # =====================
    # ATTENDANCE / LEAVE COUNTS
    # =====================
    pending_manager_attendance_count = 0
    pending_manager_leave_count = 0
    pending_hr_attendance_count = 0
    pending_accountant_attendance_count = 0

    if role != "DEALER":
        subordinate_ids = get_subordinate_user_ids(user)

        if user.is_superuser or role == "ADMIN":
            pending_manager_attendance_count = EmployeeAttendance.objects.filter(
                status="PENDING_MANAGER"
            ).count()

            pending_manager_leave_count = EmployeeLeaveRequest.objects.filter(
                status="PENDING"
            ).count()

        elif subordinate_ids:
            pending_manager_attendance_count = EmployeeAttendance.objects.filter(
                employee_id__in=subordinate_ids,
                status="PENDING_MANAGER"
            ).count()

            pending_manager_leave_count = EmployeeLeaveRequest.objects.filter(
                employee_id__in=subordinate_ids,
                status="PENDING"
            ).count()

        if user.is_superuser or role in ["ADMIN", "HR"]:
            pending_hr_attendance_count = EmployeeAttendance.objects.filter(
                status="MANAGER_APPROVED"
            ).count()

        if user.is_superuser or role in ["ADMIN", "ACCOUNTANT"]:
            pending_accountant_attendance_count = EmployeeAttendance.objects.filter(
                status="HR_APPROVED"
            ).count()

    action_links = []
    dashboard_links = []
    dealer_invoice_summary = None
    dealer_invoice_chart = [0, 0]
    recent_pending_invoices = []

    # GOLD DEALER POINTS SUMMARY
    dealer_points_summary = None

    # =====================
    # DEALER PRODUCT DATA (for dealer dashboard)
    # =====================
    top_selling_products = []
    categories = []
    top_searched_products = []
    promoted_products = []
    products_count = 0
    categories_count = 0
    farmers_count = 0
    my_orders_count = 0

    if role == "DEALER" and dealer:
        from django.db.models import Sum, Count, Q, F
        from django.db.models.functions import Coalesce
        
        # Get all active products
        all_products = Product.objects.filter(is_active=True)
        products_count = all_products.count()
        
        # Get categories with product count
        categories_list = ProductCategory.objects.filter(is_active=True)
        categories = []
        for cat in categories_list:
            cat.product_count = Product.objects.filter(category=cat, is_active=True).count()
            categories.append(cat)
        categories_count = categories_list.count()
        
        # Farmers count for this dealer
        farmers_count = FarmerData.objects.filter(dealer=dealer).count()
        
        # Top Selling Products (based on order item quantities)
        # Get pack sizes with highest total quantity sold
        top_packs = ProductPackSize.objects.filter(
            product__is_active=True
        ).annotate(
            total_sold=Coalesce(Sum('order_items__quantity_boxes'), 0)
        ).filter(total_sold__gt=0).order_by('-total_sold')[:20]
        
        # Get unique products from these packs
        product_ids = set()
        top_selling_products_list = []
        for pack in top_packs:
            if pack.product.id not in product_ids and pack.product.is_active:
                product_ids.add(pack.product.id)
                top_selling_products_list.append(pack.product)
        
        if top_selling_products_list:
            top_selling_products = top_selling_products_list[:10]
        else:
            # Fallback: show featured products or random products
            top_selling_products = list(all_products.order_by('-id')[:10])
        
        # Top Searched Products - based on cart additions
        # Get pack sizes with most cart items
        top_cart_packs = ProductPackSize.objects.filter(
            product__is_active=True
        ).annotate(
            cart_count=Count('cart_items', distinct=True),
            order_count=Count('order_items', distinct=True)
        ).order_by('-cart_count', '-order_count')[:10]
        
        product_ids_searched = set()
        top_searched_list = []
        for pack in top_cart_packs:
            if pack.product.id not in product_ids_searched and pack.product.is_active:
                product_ids_searched.add(pack.product.id)
                top_searched_list.append(pack.product)
        
        if top_searched_list:
            top_searched_products = top_searched_list[:5]
        else:
            top_searched_products = list(all_products.order_by('-id')[:5])
        
        # Add search_count attribute
        for idx, product in enumerate(top_searched_products):
            product.search_count = (idx + 1) * 5 + 10  # Dummy data
        
        # Promoted Products (products with active schemes)
        promoted_by_scheme = Product.objects.filter(
            is_active=True,
            schemes__is_active=True
        ).distinct()[:6]
        
        if promoted_by_scheme.exists():
            promoted_products = list(promoted_by_scheme)
        else:
            # If no schemes, show products with pack sizes that have MRP > sale price
            # Get pack sizes with discount - using mrp_per_unit instead of mrp_price
            discounted_packs = ProductPackSize.objects.filter(
                product__is_active=True,
                mrp_per_unit__gt=F('sale_price_per_unit')
            ).select_related('product').distinct()[:10]
            
            product_ids_promo = set()
            promo_list = []
            for pack in discounted_packs:
                if pack.product.id not in product_ids_promo and pack.product.is_active:
                    product_ids_promo.add(pack.product.id)
                    promo_list.append(pack.product)
            
            if promo_list:
                promoted_products = promo_list[:6]
            else:
                promoted_products = list(all_products.order_by('-id')[:6])
        
        # Calculate discount percentages for promoted products
        for product in promoted_products:
            first_pack = product.pack_sizes.first()
            if first_pack:
                if first_pack.mrp_per_unit and first_pack.mrp_per_unit > 0:
                    product.discount_percentage = int(
                        ((first_pack.mrp_per_unit - first_pack.sale_price_per_unit) / first_pack.mrp_per_unit) * 100
                    )
                else:
                    product.discount_percentage = 0
                product.discounted_price = first_pack.sale_price_per_unit
                product.original_price = first_pack.mrp_per_unit
            else:
                product.discount_percentage = 0
                product.discounted_price = 0
                product.original_price = 0

    # =====================
    # COMMON LINKS
    # =====================
    dashboard_links.append(
        add_link(
            "My Profile",
            "View your profile, role and account information.",
            "my_profile",
            "profile",
        )
    )

    # =====================
    # ATTENDANCE / LEAVE COMMON LINKS
    # =====================
    if role != "DEALER":
        dashboard_links += [
            add_link(
                "My Attendance / Clock In",
                "Clock-in, clock-out, GPS tracking, TA and DA summary.",
                "employee_attendance_dashboard",
                "attendance",
            ),
            add_link(
                "My Leaves",
                "View your leave requests and approval status.",
                "employee_leave_list",
                "leave",
            ),
            add_link(
                "Apply Leave",
                "Apply paid, sick, unpaid or other leave.",
                "employee_leave_request_create",
                "leave",
            ),
        ]

        if user.is_superuser or role in ["ADMIN", "STATE_HEAD", "REGIONAL_MANAGER", "ASM", "HR"]:
            dashboard_links.append(
                add_link(
                    "TA/DA Manager Approvals",
                    "Review employee travel KM and send claims to HR.",
                    "manager_attendance_approvals",
                    "approval",
                    pending_manager_attendance_count,
                    pending_manager_attendance_count > 0,
                )
            )

            dashboard_links.append(
                add_link(
                    "Leave Approvals",
                    "Approve or reject employee leave requests.",
                    "manager_leave_approvals",
                    "approval",
                    pending_manager_leave_count,
                    pending_manager_leave_count > 0,
                )
            )

            if pending_manager_attendance_count:
                action_links.append(
                    add_link(
                        "TA/DA Approval Pending",
                        f"{pending_manager_attendance_count} TA/DA claim(s) need manager approval.",
                        "manager_attendance_approvals",
                        "approval",
                        pending_manager_attendance_count,
                        True,
                    )
                )

            if pending_manager_leave_count:
                action_links.append(
                    add_link(
                        "Leave Approval Pending",
                        f"{pending_manager_leave_count} leave request(s) need approval.",
                        "manager_leave_approvals",
                        "approval",
                        pending_manager_leave_count,
                        True,
                    )
                )

        if user.is_superuser or role in ["ADMIN", "HR"]:
            dashboard_links.append(
                add_link(
                    "HR TA/DA Claims",
                    "Verify manager-approved TA/DA claims.",
                    "hr_attendance_claims",
                    "employee",
                    pending_hr_attendance_count,
                    pending_hr_attendance_count > 0,
                )
            )

            if pending_hr_attendance_count:
                action_links.append(
                    add_link(
                        "HR TA/DA Claims Pending",
                        f"{pending_hr_attendance_count} claim(s) need HR approval.",
                        "hr_attendance_claims",
                        "employee",
                        pending_hr_attendance_count,
                        True,
                    )
                )

        if user.is_superuser or role in ["ADMIN", "ACCOUNTANT"]:
            dashboard_links.append(
                add_link(
                    "Release TA/DA Funds",
                    "Release HR-approved TA/DA funds.",
                    "accountant_attendance_claims",
                    "accounts",
                    pending_accountant_attendance_count,
                    pending_accountant_attendance_count > 0,
                )
            )

            if pending_accountant_attendance_count:
                action_links.append(
                    add_link(
                        "TA/DA Fund Release Pending",
                        f"{pending_accountant_attendance_count} claim(s) are waiting for fund release.",
                        "accountant_attendance_claims",
                        "accounts",
                        pending_accountant_attendance_count,
                        True,
                    )
                )

    # =====================
    # HR MANAGEMENT DASHBOARD LINK
    # =====================
    if user.is_superuser or role in ["ADMIN", "HR", "ACCOUNTANT"]:
        dashboard_links.append(
            add_link(
                "HR Management Dashboard",
                "Employee attendance, TA/DA, top performers and leave analytics.",
                "hr_management_dashboard",
                "employee",
            )
        )

    # =====================
    # DEALER DASHBOARD
    # =====================
    if role == "DEALER":
        my_orders_count = DealerOrder.objects.filter(dealer_user=user).exclude(
            status__in=["DELIVERED", "REJECTED"]
        ).count()

        my_pending_invoice_count = DealerInvoice.objects.filter(
            order__dealer_user=user,
            pending_amount__gt=0,
        ).count()

        # =====================
        # DEALER PAYMENT SUMMARY
        # =====================
        dealer_invoice_summary = {
            "total_invoice_amount": Decimal("0.00"),
            "total_paid_amount": Decimal("0.00"),
            "total_pending_amount": Decimal("0.00"),
            "total_discount_amount": Decimal("0.00"),
            "net_payable_amount": Decimal("0.00"),
            "paid_percentage": 0,
            "pending_percentage": 0,
            "paid_invoice_count": 0,
            "pending_invoice_count": 0,
            "total_invoice_count": 0,
        }

        dealer_invoice_chart = [0, 0]
        recent_pending_invoices = []

        dealer_invoice_qs = DealerInvoice.objects.filter(
            order__dealer_user=user
        ).select_related("order").order_by("-id")

        total_invoice_amount = Decimal("0.00")
        total_paid_amount = Decimal("0.00")
        total_pending_amount = Decimal("0.00")
        total_discount_amount = Decimal("0.00")

        paid_invoice_count = 0
        pending_invoice_count = 0

        for invoice in dealer_invoice_qs:
            invoice_total = getattr(invoice, "total_amount", Decimal("0.00")) or Decimal("0.00")
            pending_amount = getattr(invoice, "pending_amount", Decimal("0.00")) or Decimal("0.00")
            discount_amount = getattr(invoice, "discount_amount", Decimal("0.00")) or Decimal("0.00")

            paid_amount = getattr(invoice, "paid_amount", None)

            if paid_amount is None:
                paid_amount = invoice_total - discount_amount - pending_amount

            paid_amount = paid_amount or Decimal("0.00")

            if paid_amount < 0:
                paid_amount = Decimal("0.00")

            total_invoice_amount += invoice_total
            total_paid_amount += paid_amount
            total_pending_amount += pending_amount
            total_discount_amount += discount_amount

            if pending_amount > 0:
                pending_invoice_count += 1
            else:
                paid_invoice_count += 1

        net_payable_amount = total_paid_amount + total_pending_amount

        paid_percentage = 0
        pending_percentage = 0

        if net_payable_amount > 0:
            paid_percentage = round((total_paid_amount / net_payable_amount) * 100, 2)
            pending_percentage = round((total_pending_amount / net_payable_amount) * 100, 2)

        dealer_invoice_summary = {
            "total_invoice_amount": total_invoice_amount,
            "total_paid_amount": total_paid_amount,
            "total_pending_amount": total_pending_amount,
            "total_discount_amount": total_discount_amount,
            "net_payable_amount": net_payable_amount,
            "paid_percentage": paid_percentage,
            "pending_percentage": pending_percentage,
            "paid_invoice_count": paid_invoice_count,
            "pending_invoice_count": pending_invoice_count,
            "total_invoice_count": dealer_invoice_qs.count(),
        }

        dealer_invoice_chart = [
            float(total_paid_amount),
            float(total_pending_amount),
        ]

        recent_pending_invoices = dealer_invoice_qs.filter(
            pending_amount__gt=0
        ).order_by("-id")[:5]

        # =====================
        # DEALER GOLD POINTS SUMMARY
        # =====================
        if dealer:
            setting = DealerPointSetting.get_settings()

            total_earned_points = DealerPointLedger.objects.filter(
                dealer=dealer,
                transaction_type="CREDIT"
            ).aggregate(total=Sum("points"))["total"] or 0

            redeemed_points = DealerPointLedger.objects.filter(
                dealer=dealer,
                transaction_type="DEBIT"
            ).aggregate(total=Sum("points"))["total"] or 0

            balance_points = total_earned_points - redeemed_points
            point_value = Decimal(balance_points) * setting.rupees_per_point

            farmers_count_local = FarmerData.objects.filter(
                dealer=dealer
            ).count()

            pending_redemptions = DealerPointRedemptionRequest.objects.filter(
                dealer=dealer
            ).exclude(
                status__in=["APPROVED", "REJECTED"]
            ).count()

            minimum_money_points = setting.minimum_money_redemption_points or 1

            progress_percentage = int((balance_points / minimum_money_points) * 100)

            if progress_percentage > 100:
                progress_percentage = 100

            if progress_percentage < 0:
                progress_percentage = 0

            dealer_points_summary = {
                "balance_points": balance_points,
                "total_earned_points": total_earned_points,
                "redeemed_points": redeemed_points,
                "point_value": point_value,
                "farmers_count": farmers_count_local,
                "pending_redemptions": pending_redemptions,
                "farmer_points": setting.farmer_points,
                "rupees_per_point": setting.rupees_per_point,
                "minimum_money_points": setting.minimum_money_redemption_points,
                "minimum_product_points": setting.minimum_product_redemption_points,
                "progress_percentage": progress_percentage,
            }

        dashboard_links += [
            add_link(
                "Products",
                "Browse products and add items to cart.",
                "dealer_products",
                "products",
            ),
            add_link(
                "My Cart",
                "Check selected products before placing order.",
                "dealer_cart",
                "cart",
            ),
            add_link(
                "My Orders",
                "Track current order status.",
                "dealer_order_list",
                "orders",
                my_orders_count,
                my_orders_count > 0,
            ),
            add_link(
                "My Invoices",
                "View dealer copy and pending bill details.",
                "dealer_invoice_list",
                "invoice",
                my_pending_invoice_count,
                my_pending_invoice_count > 0,
            ),
            add_link(
                "Credit Notes",
                "View sales return credit notes shared by company.",
                "sales_return_credit_note_list",
                "return",
            ),
            add_link(
                "My Points & Redemption",
                "Check your points balance and redeem rewards.",
                "dealer_points_dashboard",
                "star",
            ),
            add_link(
                "Farmer Data Upload",
                "Upload farmer data to earn reward points.",
                "dealer_farmer_data",
                "farmers",
            ),
        ]

        if my_orders_count:
            action_links.append(
                add_link(
                    "Orders in Progress",
                    f"{my_orders_count} order(s) are still active.",
                    "dealer_order_list",
                    "orders",
                    my_orders_count,
                    True,
                )
            )

        if my_pending_invoice_count:
            action_links.append(
                add_link(
                    "Pending Invoice Balance",
                    f"₹{total_pending_amount} pending from {my_pending_invoice_count} invoice(s).",
                    "dealer_invoice_list",
                    "invoice",
                    my_pending_invoice_count,
                    True,
                )
            )

    # =====================
    # SALES OFFICER
    # =====================
    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        sales_order_count = count_orders(
            "SALES_APPROVAL",
            "concerned_sales_officer"
        )

        farmer_meet_count = count_farmer_meets(
            "PENDING_SALES_OFFICER",
            "sales_officer"
        )

        own_dealer_count = Dealer.objects.filter(
            created_by_sales_officer=user
        ).count()

        dashboard_links += [
            add_link(
                "Sales Officer Order Approvals",
                "Approve or reject orders assigned to you.",
                "dealer_order_list",
                "approval",
                sales_order_count,
                sales_order_count > 0,
            ),
            add_link(
                "My Dealers",
                "View dealers created or assigned to you.",
                "dealer_list",
                "dealer",
                own_dealer_count,
                False,
            ),
            add_link(
                "Create Dealer",
                "Create a new dealer and send to ASM approval.",
                "dealer_create",
                "dealer",
            ),
            add_link(
                "Farmer Meets",
                "Review farmer meet requests assigned to you.",
                "farmer_meet_list",
                "farmers",
                farmer_meet_count,
                farmer_meet_count > 0,
            ),
        ]

        if sales_order_count:
            action_links.append(
                add_link(
                    "Order Approval Required",
                    f"{sales_order_count} order(s) need your approval.",
                    "dealer_order_list",
                    "approval",
                    sales_order_count,
                    True,
                )
            )

        if farmer_meet_count:
            action_links.append(
                add_link(
                    "Farmer Meet Approval",
                    f"{farmer_meet_count} farmer meet request(s) need action.",
                    "farmer_meet_list",
                    "farmers",
                    farmer_meet_count,
                    True,
                )
            )

    # =====================
    # ASM
    # =====================
    if role == "ASM":
        asm_order_count = count_orders("ASM_APPROVAL", "concerned_asm")

        dealer_approval_count = Dealer.objects.filter(
            approval_status="PENDING_ASM",
            concerned_asm=user,
        ).count()

        farmer_meet_count = count_farmer_meets("PENDING_ASM", "asm")

        dashboard_links += [
            add_link(
                "ASM Order Approvals",
                "Approve pending dealer orders.",
                "dealer_order_list",
                "approval",
                asm_order_count,
                asm_order_count > 0,
            ),
            add_link(
                "Dealer Approvals",
                "Approve new dealers created by sales officers.",
                "dealer_approval_list",
                "dealer",
                dealer_approval_count,
                dealer_approval_count > 0,
            ),
            add_link(
                "Dealer List",
                "View dealers under your area.",
                "dealer_list",
                "dealer",
            ),
            add_link(
                "Farmer Meets",
                "Approve farmer meet requests.",
                "farmer_meet_list",
                "farmers",
                farmer_meet_count,
                farmer_meet_count > 0,
            ),
        ]

        if asm_order_count:
            action_links.append(
                add_link(
                    "ASM Order Approval",
                    f"{asm_order_count} order(s) are waiting for ASM approval.",
                    "dealer_order_list",
                    "approval",
                    asm_order_count,
                    True,
                )
            )

        if dealer_approval_count:
            action_links.append(
                add_link(
                    "Dealer Approval Pending",
                    f"{dealer_approval_count} dealer profile(s) need approval.",
                    "dealer_approval_list",
                    "dealer",
                    dealer_approval_count,
                    True,
                )
            )

        if farmer_meet_count:
            action_links.append(
                add_link(
                    "Farmer Meet Approval",
                    f"{farmer_meet_count} farmer meet request(s) need action.",
                    "farmer_meet_list",
                    "farmers",
                    farmer_meet_count,
                    True,
                )
            )

    # =====================
    # REGIONAL MANAGER
    # =====================
    if role == "REGIONAL_MANAGER":
        rsm_order_count = count_orders("RSM_APPROVAL", "concerned_rsm")

        dealer_approval_count = Dealer.objects.filter(
            approval_status="FORWARDED_RM",
            forwarded_regional_manager=user,
        ).count()

        farmer_meet_count = count_farmer_meets(
            "PENDING_REGIONAL_MANAGER",
            "regional_manager"
        )

        dashboard_links += [
            add_link(
                "RSM Order Approvals",
                "Approve orders forwarded to regional manager.",
                "dealer_order_list",
                "approval",
                rsm_order_count,
                rsm_order_count > 0,
            ),
            add_link(
                "Dealer Approvals",
                "Approve dealers forwarded by ASM.",
                "dealer_approval_list",
                "dealer",
                dealer_approval_count,
                dealer_approval_count > 0,
            ),
            add_link(
                "Dealer List",
                "View dealers in your region.",
                "dealer_list",
                "dealer",
            ),
            add_link(
                "Farmer Meets",
                "Review farmer meet requests.",
                "farmer_meet_list",
                "farmers",
                farmer_meet_count,
                farmer_meet_count > 0,
            ),
        ]

        if rsm_order_count:
            action_links.append(
                add_link(
                    "RSM Order Approval",
                    f"{rsm_order_count} order(s) are waiting for RSM approval.",
                    "dealer_order_list",
                    "approval",
                    rsm_order_count,
                    True,
                )
            )

        if dealer_approval_count:
            action_links.append(
                add_link(
                    "Dealer Approval Pending",
                    f"{dealer_approval_count} dealer profile(s) need regional approval.",
                    "dealer_approval_list",
                    "dealer",
                    dealer_approval_count,
                    True,
                )
            )

        if farmer_meet_count:
            action_links.append(
                add_link(
                    "Farmer Meet Approval",
                    f"{farmer_meet_count} farmer meet request(s) need action.",
                    "farmer_meet_list",
                    "farmers",
                    farmer_meet_count,
                    True,
                )
            )

    # =====================
    # STATE HEAD
    # =====================
    if role == "STATE_HEAD":
        dealer_approval_count = Dealer.objects.filter(
            approval_status="FORWARDED_STATE_HEAD",
            forwarded_state_head=user,
        ).count()

        farmer_meet_count = count_farmer_meets(
            "PENDING_STATE_HEAD",
            "state_head"
        )

        dashboard_links += [
            add_link(
                "Dealer Approvals",
                "Approve dealers forwarded to State Head.",
                "dealer_approval_list",
                "dealer",
                dealer_approval_count,
                dealer_approval_count > 0,
            ),
            add_link(
                "Dealer List",
                "View state dealers.",
                "dealer_list",
                "dealer",
            ),
            add_link(
                "Farmer Meets",
                "Approve high-value farmer meet requests.",
                "farmer_meet_list",
                "farmers",
                farmer_meet_count,
                farmer_meet_count > 0,
            ),
        ]

        if dealer_approval_count:
            action_links.append(
                add_link(
                    "Dealer Approval Pending",
                    f"{dealer_approval_count} dealer profile(s) need State Head approval.",
                    "dealer_approval_list",
                    "dealer",
                    dealer_approval_count,
                    True,
                )
            )

        if farmer_meet_count:
            action_links.append(
                add_link(
                    "Farmer Meet Approval",
                    f"{farmer_meet_count} farmer meet request(s) need action.",
                    "farmer_meet_list",
                    "farmers",
                    farmer_meet_count,
                    True,
                )
            )

    # =====================
    # ACCOUNTANT
    # =====================
    if role == "ACCOUNTANT":
        accountant_order_count = DealerOrder.objects.filter(
            status__in=[
                "ACCOUNTANT_ORDER_APPROVAL",
                "ACCOUNTANT_INVOICE_REVIEW",
            ]
        ).count()

        sales_conversion_count = DealerInvoice.objects.filter(
            is_converted_to_sales=False
        ).count()

        dashboard_links += [
            add_link(
                "Accountant Review",
                "Review orders and invoice details.",
                "dealer_order_list",
                "approval",
                accountant_order_count,
                accountant_order_count > 0,
            ),
            add_link(
                "Invoices",
                "View released invoices and balances.",
                "dealer_invoice_list",
                "invoice",
            ),
            add_link(
                "Convert to Sales",
                "Convert released invoices into sales/payment records.",
                "dealer_invoice_list",
                "accounts",
                sales_conversion_count,
                sales_conversion_count > 0,
            ),
            add_link(
                "Transport Copies",
                "View and share transporter copies.",
                "transport_copy_list",
                "transport",
            ),
            add_link(
                "Sales Return / Credit Note",
                "Create return invoices and credit notes.",
                "sales_return_credit_note_list",
                "return",
            ),
            add_link(
                "Accounts",
                "Payment in, payment out and bank accounts.",
                "accounts_dashboard",
                "accounts",
            ),
            add_link(
                "Paid Sales Report",
                "View paid sales graph and payment mode report.",
                "paid_sales_report",
                "report",
            ),
            add_link(
                "Order vs Payment Report",
                "Compare order value, received and pending amount.",
                "order_payment_difference_report",
                "report",
            ),
            add_link(
                "Dealer List",
                "View all dealer details.",
                "dealer_list",
                "dealer",
            ),
        ]

        if accountant_order_count:
            action_links.append(
                add_link(
                    "Accountant Action Required",
                    f"{accountant_order_count} order/invoice review(s) need action.",
                    "dealer_order_list",
                    "approval",
                    accountant_order_count,
                    True,
                )
            )

        if sales_conversion_count:
            action_links.append(
                add_link(
                    "Sales Conversion Pending",
                    f"{sales_conversion_count} invoice(s) are not converted to sales.",
                    "dealer_invoice_list",
                    "accounts",
                    sales_conversion_count,
                    True,
                )
            )

    # =====================
    # WAREHOUSE / PRODUCTION
    # =====================
    if role in ["WAREHOUSE_MANAGER", "PRODUCTION_INCHARGE"]:
        if role == "WAREHOUSE_MANAGER":
            warehouse_review_count = count_orders(
                "WAREHOUSE_REVIEW",
                "warehouse_manager"
            )
            dispatch_pending_count = count_orders(
                "DISPATCH_PENDING",
                "warehouse_manager"
            )
        else:
            warehouse_review_count = DealerOrder.objects.filter(
                status="WAREHOUSE_REVIEW"
            ).count()
            dispatch_pending_count = DealerOrder.objects.filter(
                status="DISPATCH_PENDING"
            ).count()

        dashboard_links += [
            add_link(
                "My Warehouse Products",
                "View all products assigned to your warehouse.",
                "warehouse_my_products",
                "products",
            ),
            add_link(
                "My Warehouse Orders",
                "View all orders assigned to your warehouse.",
                "warehouse_my_orders",
                "orders",
                warehouse_review_count + dispatch_pending_count,
                (warehouse_review_count + dispatch_pending_count) > 0,
            ),
            add_link(
                "Warehouse Review",
                "Check product HSN, batch, MFG and expiry details.",
                "dealer_order_list",
                "warehouse",
                warehouse_review_count,
                warehouse_review_count > 0,
            ),
            add_link(
                "Dispatch Pending",
                "Enter transport and driver details.",
                "dealer_order_list",
                "transport",
                dispatch_pending_count,
                dispatch_pending_count > 0,
            ),
            add_link(
                "Transport Copies",
                "Print transporter copies for driver/transport.",
                "transport_copy_list",
                "transport",
            ),
            add_link(
                "Warehouses",
                "View warehouse details.",
                "warehouse_list",
                "warehouse",
            ),
        ]

        if warehouse_review_count:
            action_links.append(
                add_link(
                    "Warehouse Review Pending",
                    f"{warehouse_review_count} order(s) need warehouse checking.",
                    "dealer_order_list",
                    "warehouse",
                    warehouse_review_count,
                    True,
                )
            )

        if dispatch_pending_count:
            action_links.append(
                add_link(
                    "Dispatch Details Required",
                    f"{dispatch_pending_count} order(s) need transport details.",
                    "dealer_order_list",
                    "transport",
                    dispatch_pending_count,
                    True,
                )
            )

    # =====================
    # ADMIN
    # =====================
    if user.is_superuser or role == "ADMIN":
        pending_order_count = DealerOrder.objects.filter(
            status__in=[
                "SALES_APPROVAL",
                "ASM_APPROVAL",
                "RSM_APPROVAL",
                "ACCOUNTANT_ORDER_APPROVAL",
                "WAREHOUSE_REVIEW",
                "ACCOUNTANT_INVOICE_REVIEW",
                "DISPATCH_PENDING",
            ]
        ).count()

        pending_dealer_count = Dealer.objects.exclude(
            approval_status__in=["APPROVED", "REJECTED"]
        ).count()

        pending_farmer_count = FarmerMeetRequest.objects.exclude(
            approval_status__in=["APPROVED", "REJECTED"]
        ).count()

        sales_conversion_count = DealerInvoice.objects.filter(
            is_converted_to_sales=False
        ).count()

        dashboard_links += [
            add_link(
                "Orders",
                "View all orders and pending stages.",
                "dealer_order_list",
                "orders",
                pending_order_count,
                pending_order_count > 0,
            ),
            add_link(
                "Dealers",
                "View, create and edit dealer profiles.",
                "dealer_list",
                "dealer",
                pending_dealer_count,
                pending_dealer_count > 0,
            ),
            add_link(
                "Dealer Approvals",
                "Approve or reject dealer profiles.",
                "dealer_approval_list",
                "approval",
                pending_dealer_count,
                pending_dealer_count > 0,
            ),
            add_link(
                "Products",
                "Manage products and pack sizes.",
                "product_list",
                "products",
            ),
            add_link(
                "Categories & Payment Rules",
                "Manage category payment discount rules.",
                "product_category_list",
                "products",
            ),
            add_link(
                "Employees",
                "Manage employee profiles and roles.",
                "employee_list",
                "employee",
            ),
            add_link(
                "Create Employee",
                "Create a new staff login.",
                "employee_create",
                "employee",
            ),
            add_link(
                "Accounts",
                "Manage bank accounts and payments.",
                "accounts_dashboard",
                "accounts",
            ),
            add_link(
                "Invoices",
                "View all invoices.",
                "dealer_invoice_list",
                "invoice",
            ),
            add_link(
                "Transport Copies",
                "Print/share all transporter copies.",
                "transport_copy_list",
                "transport",
            ),
            add_link(
                "Sales Return / Credit Note",
                "Create and manage credit notes.",
                "sales_return_credit_note_list",
                "return",
            ),
            add_link(
                "Paid Sales Report",
                "Sales graph and payment report.",
                "paid_sales_report",
                "report",
            ),
            add_link(
                "Order vs Payment Report",
                "Order amount vs received amount.",
                "order_payment_difference_report",
                "report",
            ),
            add_link(
                "Farmer Meets",
                "View farmer meet requests.",
                "farmer_meet_list",
                "farmers",
                pending_farmer_count,
                pending_farmer_count > 0,
            ),
        ]

        if pending_order_count:
            action_links.append(
                add_link(
                    "Orders Need Action",
                    f"{pending_order_count} order(s) are pending in workflow.",
                    "dealer_order_list",
                    "orders",
                    pending_order_count,
                    True,
                )
            )

        if pending_dealer_count:
            action_links.append(
                add_link(
                    "Dealer Approval Pending",
                    f"{pending_dealer_count} dealer profile(s) need approval.",
                    "dealer_approval_list",
                    "dealer",
                    pending_dealer_count,
                    True,
                )
            )

        if sales_conversion_count:
            action_links.append(
                add_link(
                    "Sales Conversion Pending",
                    f"{sales_conversion_count} invoice(s) are not converted to sales.",
                    "dealer_invoice_list",
                    "accounts",
                    sales_conversion_count,
                    True,
                )
            )

    # =====================
    # HR
    # =====================
    if role == "HR":
        dashboard_links += [
            add_link(
                "Employees",
                "View employee details.",
                "employee_list",
                "employee",
            ),
            add_link(
                "Create Employee",
                "Create new employee profile.",
                "employee_create",
                "employee",
            ),
        ]

    # =====================
    # DEVELOPMENT OFFICER
    # =====================
    if role == "DEVELOPMENT_OFFICER":
        dashboard_links += [
            add_link(
                "Create Farmer Meet",
                "Create a new farmer meet request.",
                "farmer_meet_create",
                "farmers",
            ),
            add_link(
                "Farmer Meets",
                "Track farmer meet approval status.",
                "farmer_meet_list",
                "farmers",
            ),
        ]

    # Remove duplicate URLs while keeping first item
    unique_links = []
    seen_urls = set()

    for link in dashboard_links:
        if link["url"] not in seen_urls:
            unique_links.append(link)
            seen_urls.add(link["url"])

    dashboard_links = unique_links

    pending_action_count = sum(link["count"] for link in action_links)

    # =====================
    # TARGET
    # =====================
    target = {
        "show": False,
        "title": "Target Progress",
        "target_amount": Decimal("0.00"),
        "achieved_amount": Decimal("0.00"),
        "balance_amount": Decimal("0.00"),
        "percentage": 0,
    }

    if role == "DEALER":
        if dealer:
            target = {
                "show": True,
                "title": "Dealer Target Progress",
                "target_amount": dealer.yearly_target_amount,
                "achieved_amount": dealer.achieved_amount,
                "balance_amount": dealer.target_balance_amount,
                "percentage": dealer.target_achieved_percentage,
            }

    elif role in [
        "SALES_OFFICER_SENIOR",
        "SALES_OFFICER_JUNIOR",
        "ASM",
        "REGIONAL_MANAGER",
        "STATE_HEAD",
    ]:
        current_year = timezone.localdate().year

        employee_target = EmployeeYearlyTarget.objects.filter(
            employee=user,
            year=current_year,
        ).first()

        if employee_target:
            target = {
                "show": True,
                "title": "Employee Target Progress",
                "target_amount": employee_target.target_amount,
                "achieved_amount": employee_target.achieved_amount,
                "balance_amount": employee_target.balance_amount,
                "percentage": employee_target.achieved_percentage,
            }

    return render(
        request,
        "core/dashboard.html",
        {
            "profile": profile,
            "dealer": dealer,
            "dashboard_links": dashboard_links,
            "action_links": action_links,
            "pending_action_count": pending_action_count,
            "target": target,
            "dealer_invoice_summary": dealer_invoice_summary,
            "dealer_invoice_chart": dealer_invoice_chart,
            "recent_pending_invoices": recent_pending_invoices,
            "dealer_points_summary": dealer_points_summary,

            # New data for dealer product sections
            "top_selling_products": top_selling_products,
            "categories": categories,
            "top_searched_products": top_searched_products,
            "promoted_products": promoted_products,
            "products_count": products_count,
            "categories_count": categories_count,
            "farmers_count": farmers_count,
            "my_orders_count": my_orders_count,

            "pending_manager_attendance_count": pending_manager_attendance_count,
            "pending_manager_leave_count": pending_manager_leave_count,
            "pending_hr_attendance_count": pending_hr_attendance_count,
            "pending_accountant_attendance_count": pending_accountant_attendance_count,
        }
    )


   
@login_required
def my_profile(request):
    profile = get_or_create_profile(request.user)
    dealer = Dealer.objects.filter(user=request.user).first()
    dealer_contacts = get_dealer_contact_context(dealer) if dealer else None
    current_year = timezone.now().year

    employee_target = EmployeeYearlyTarget.objects.filter(
        employee=request.user,
        year=current_year
    ).first()

    if request.method == "POST":
        profile_image_form = ProfileImageUpdateForm(
            request.POST,
            request.FILES,
            instance=profile
        )

        if profile_image_form.is_valid():
            profile_image_form.save()
            messages.success(request, "Profile picture updated successfully.")
            return redirect("my_profile")

    else:
        profile_image_form = ProfileImageUpdateForm(instance=profile)

    return render(
        request,
        "core/my_profile.html",
        {
            "profile": profile,
            "dealer": dealer,
            "dealer_contacts": dealer_contacts,
            "employee_target": employee_target,
            "target": get_target_context(request.user),
            "profile_image_form": profile_image_form,
        }
    )



@login_required
def dealer_pending(request):
    dealer = Dealer.objects.filter(user=request.user).first()

    return render(
        request,
        "core/dealer_pending.html",
        {
            "dealer": dealer,
        }
    )



@login_required
@role_required(["ADMIN", "SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"])
@transaction.atomic
def dealer_create(request):
    if request.method == "POST":
        form = DealerCreateForm(
            request.POST,
            request.FILES,
            request_user=request.user
        )

        if form.is_valid():
            selected_sales_officer = form.cleaned_data.get("created_by_sales_officer")

            if not selected_sales_officer:
                messages.error(request, "Please select dealer manager / sales officer.")
                return redirect("dealer_create")

            assigned_asm = find_assigned_asm_for_sales_officer(selected_sales_officer)

            if not assigned_asm:
                messages.error(
                    request,
                    "ASM is not assigned for the selected Sales Officer. Please assign ASM first."
                )
                return redirect("dealer_create")

            email = form.cleaned_data["email"].lower().strip()

            dealer_user = User.objects.create_user(
                username=email,
                email=email,
                password=form.cleaned_data["password"],
                is_active=False,
            )

            UserProfile.objects.update_or_create(
                user=dealer_user,
                defaults={
                    "role": "DEALER",
                    "phone": form.cleaned_data.get("phone"),
                    "address": form.cleaned_data.get("owner_address"),
                    "state": form.cleaned_data.get("state"),
                    "district": form.cleaned_data.get("district"),
                }
            )

            dealer = form.save(commit=False)
            dealer.user = dealer_user
            dealer.created_by_sales_officer = selected_sales_officer
            dealer.concerned_asm = assigned_asm
            dealer.approval_status = "PENDING_ASM"
            dealer.is_active = False
            dealer.deposit_mode = "CHEQUE"
            dealer.save()

            DealerApprovalHistory.objects.create(
                dealer=dealer,
                action="CREATED",
                performed_by=request.user,
                remarks=(
                    "Dealer created and automatically sent to assigned ASM approval. "
                    f"Dealer manager: {selected_sales_officer.get_full_name() or selected_sales_officer.email or selected_sales_officer.username}"
                ),
            )

            messages.success(
                request,
                "Dealer created and sent to assigned ASM approval."
            )
            return redirect("dealer_approval_list")

    else:
        form = DealerCreateForm(request_user=request.user)

    return render(
        request,
        "core/dealer_form.html",
        {
            "form": form,
            "title": "Create Dealer",
            "button_text": "Save Dealer",
        }
    )



def can_edit_dealer_profile(user):
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in ["ADMIN", "ACCOUNTANT"]


def get_dealers_for_user(user):
    if user.is_superuser:
        return Dealer.objects.all()

    profile = getattr(user, "profile", None)

    if not profile:
        return Dealer.objects.none()

    role = profile.role

    # Admin / Accountant can see all dealers
    if role in ["ADMIN", "ACCOUNTANT"]:
        return Dealer.objects.all()

    # State Head can see only his/her state dealers
    if role == "STATE_HEAD":
        if profile.state:
            return Dealer.objects.filter(state=profile.state)

        return Dealer.objects.none()

    # Sales officers can see only dealers created/assigned by them
    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        return Dealer.objects.filter(created_by_sales_officer=user)

    return Dealer.objects.none()


@login_required
def dealer_list(request):
    dealers = get_dealers_for_user(request.user)

    search = request.GET.get("search", "").strip()
    status = request.GET.get("status", "").strip()

    if search:
        search_query = (
            Q(firm_name__icontains=search) |
            Q(owner_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search) |
            Q(gst_number__icontains=search)
        )

        if any(field.name == "dealer_code" for field in Dealer._meta.get_fields()):
            search_query |= Q(dealer_code__icontains=search)

        dealers = dealers.filter(search_query)

    if status == "active":
        dealers = dealers.filter(user__is_active=True)

    elif status == "inactive":
        dealers = dealers.filter(user__is_active=False)

    dealers = dealers.select_related(
        "user",
        "state",
        "district",
        "created_by_sales_officer",
        "concerned_asm",
    ).order_by("-id")

    return render(
        request,
        "core/dealer_list.html",
        {
            "dealers": dealers,
            "search": search,
            "status": status,
            "can_edit_dealer": can_edit_dealer_profile(request.user),
        }
    )


@login_required
def dealer_edit(request, dealer_id):
    if not can_edit_dealer_profile(request.user):
        messages.error(request, "Only Admin or Accountant can edit dealer profiles.")
        return redirect("dealer_list")

    dealer = get_object_or_404(Dealer, id=dealer_id)

    if request.method == "POST":
        form = DealerEditForm(
            request.POST,
            request.FILES,
            instance=dealer,
            request_user=request.user,
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Dealer profile updated successfully.")
            return redirect("dealer_list")
    else:
        form = DealerEditForm(
            instance=dealer,
            request_user=request.user,
        )

    return render(
        request,
        "core/dealer_form.html",
        {
            "form": form,
            "title": "Edit Dealer",
            "button_text": "Update Dealer",
            "dealer": dealer,
        }
    )



@login_required
def dealer_approval_list(request):
    role = get_user_role(request.user)

    if request.user.is_superuser or role == "ADMIN":
        dealers = Dealer.objects.exclude(approval_status="APPROVED")

    elif role == "ASM":
        dealers = Dealer.objects.filter(
            concerned_asm=request.user,
            approval_status="PENDING_ASM"
        )

    elif role == "REGIONAL_MANAGER":
        dealers = Dealer.objects.filter(
            forwarded_regional_manager=request.user,
            approval_status="FORWARDED_RM"
        )

    elif role == "STATE_HEAD":
        dealers = Dealer.objects.filter(
            forwarded_state_head=request.user,
            approval_status="FORWARDED_STATE_HEAD"
        )

    else:
        messages.error(request, "You do not have permission.")
        return redirect("dashboard")

    return render(
        request,
        "core/dealer_approval_list.html",
        {
            "dealers": dealers,
        }
    )


@login_required
def dealer_approval_detail(request, dealer_id):
    dealer = get_object_or_404(Dealer, id=dealer_id)
    role = get_user_role(request.user)

    allowed = False

    if request.user.is_superuser or role == "ADMIN":
        allowed = True
    elif role == "ASM" and dealer.concerned_asm == request.user and dealer.approval_status == "PENDING_ASM":
        allowed = True
    elif role == "REGIONAL_MANAGER" and dealer.forwarded_regional_manager == request.user and dealer.approval_status == "FORWARDED_RM":
        allowed = True
    elif role == "STATE_HEAD" and dealer.forwarded_state_head == request.user and dealer.approval_status == "FORWARDED_STATE_HEAD":
        allowed = True

    if not allowed:
        messages.error(request, "You cannot access this dealer approval.")
        return redirect("dealer_approval_list")

    return render(
        request,
        "core/dealer_approval_detail.html",
        {
            "dealer": dealer,
            "role": role,
            "reject_form": DealerRejectForm(),
            "forward_rm_form": ForwardToRegionalManagerForm(),
            "forward_state_head_form": ForwardToStateHeadForm(),
        }
    )


@login_required
@transaction.atomic
def dealer_approve(request, dealer_id):
    dealer = get_object_or_404(Dealer, id=dealer_id)
    role = get_user_role(request.user)

    if request.method != "POST":
        return redirect("dealer_approval_detail", dealer_id=dealer.id)

    if request.user.is_superuser or role == "ADMIN":
        dealer.activate_dealer_login()

        DealerApprovalHistory.objects.create(
            dealer=dealer,
            action="ASM_APPROVED",
            performed_by=request.user,
            remarks="Dealer approved by Admin.",
        )

        messages.success(request, "Dealer approved. Dealer can login now.")
        return redirect("dealer_approval_list")

    if role == "ASM" and dealer.concerned_asm == request.user and dealer.approval_status == "PENDING_ASM":
        dealer.asm_approved_by = request.user
        dealer.asm_approved_at = timezone.now()
        dealer.activate_dealer_login()

        DealerApprovalHistory.objects.create(
            dealer=dealer,
            action="ASM_APPROVED",
            performed_by=request.user,
            remarks="Dealer approved by ASM.",
        )

        messages.success(request, "Dealer approved. Dealer can login now.")
        return redirect("dealer_approval_list")

    if role == "REGIONAL_MANAGER" and dealer.forwarded_regional_manager == request.user and dealer.approval_status == "FORWARDED_RM":
        dealer.regional_manager_approved_by = request.user
        dealer.regional_manager_approved_at = timezone.now()
        dealer.activate_dealer_login()

        DealerApprovalHistory.objects.create(
            dealer=dealer,
            action="RM_APPROVED",
            performed_by=request.user,
            remarks="Dealer approved by Regional Manager.",
        )

        messages.success(request, "Dealer approved. Dealer can login now.")
        return redirect("dealer_approval_list")

    if role == "STATE_HEAD" and dealer.forwarded_state_head == request.user and dealer.approval_status == "FORWARDED_STATE_HEAD":
        dealer.state_head_approved_by = request.user
        dealer.state_head_approved_at = timezone.now()
        dealer.activate_dealer_login()

        DealerApprovalHistory.objects.create(
            dealer=dealer,
            action="STATE_HEAD_APPROVED",
            performed_by=request.user,
            remarks="Dealer approved by State Head.",
        )

        messages.success(request, "Dealer approved. Dealer can login now.")
        return redirect("dealer_approval_list")

    messages.error(request, "You cannot approve this dealer.")
    return redirect("dealer_approval_list")


@login_required
@transaction.atomic
def dealer_forward_to_rm(request, dealer_id):
    dealer = get_object_or_404(Dealer, id=dealer_id)
    role = get_user_role(request.user)

    if role != "ASM" or dealer.concerned_asm != request.user or dealer.approval_status != "PENDING_ASM":
        messages.error(request, "Only concerned ASM can forward this dealer.")
        return redirect("dealer_approval_list")

    form = ForwardToRegionalManagerForm(request.POST)

    if form.is_valid():
        rm_user = form.cleaned_data["regional_manager"]

        dealer.asm_approved_by = request.user
        dealer.asm_approved_at = timezone.now()
        dealer.forwarded_regional_manager = rm_user
        dealer.approval_status = "FORWARDED_RM"
        dealer.save()

        DealerApprovalHistory.objects.create(
            dealer=dealer,
            action="FORWARDED_RM",
            performed_by=request.user,
            remarks=f"Forwarded to Regional Manager: {rm_user.username}",
        )

        messages.success(request, "Dealer forwarded to Regional Manager.")
        return redirect("dealer_approval_list")

    messages.error(request, "Invalid Regional Manager.")
    return redirect("dealer_approval_detail", dealer_id=dealer.id)


@login_required
@transaction.atomic
def dealer_forward_to_state_head(request, dealer_id):
    dealer = get_object_or_404(Dealer, id=dealer_id)
    role = get_user_role(request.user)

    if role != "REGIONAL_MANAGER" or dealer.forwarded_regional_manager != request.user or dealer.approval_status != "FORWARDED_RM":
        messages.error(request, "Only Regional Manager can forward this dealer.")
        return redirect("dealer_approval_list")

    form = ForwardToStateHeadForm(request.POST)

    if form.is_valid():
        state_head_user = form.cleaned_data["state_head"]

        dealer.regional_manager_approved_by = request.user
        dealer.regional_manager_approved_at = timezone.now()
        dealer.forwarded_state_head = state_head_user
        dealer.approval_status = "FORWARDED_STATE_HEAD"
        dealer.save()

        DealerApprovalHistory.objects.create(
            dealer=dealer,
            action="FORWARDED_STATE_HEAD",
            performed_by=request.user,
            remarks=f"Forwarded to State Head: {state_head_user.username}",
        )

        messages.success(request, "Dealer forwarded to State Head.")
        return redirect("dealer_approval_list")

    messages.error(request, "Invalid State Head.")
    return redirect("dealer_approval_detail", dealer_id=dealer.id)


@login_required
@transaction.atomic
def dealer_reject(request, dealer_id):
    dealer = get_object_or_404(Dealer, id=dealer_id)
    role = get_user_role(request.user)

    allowed = False

    if request.user.is_superuser or role == "ADMIN":
        allowed = True
    elif role == "ASM" and dealer.concerned_asm == request.user and dealer.approval_status == "PENDING_ASM":
        allowed = True
    elif role == "REGIONAL_MANAGER" and dealer.forwarded_regional_manager == request.user and dealer.approval_status == "FORWARDED_RM":
        allowed = True
    elif role == "STATE_HEAD" and dealer.forwarded_state_head == request.user and dealer.approval_status == "FORWARDED_STATE_HEAD":
        allowed = True

    if not allowed:
        messages.error(request, "You cannot reject this dealer.")
        return redirect("dealer_approval_list")

    form = DealerRejectForm(request.POST)

    if form.is_valid():
        reason = form.cleaned_data["rejection_reason"]

        dealer.reject_dealer(request.user, reason)

        DealerApprovalHistory.objects.create(
            dealer=dealer,
            action="REJECTED",
            performed_by=request.user,
            remarks=reason,
        )

        messages.success(request, "Dealer rejected successfully.")
        return redirect("dealer_approval_list")

    messages.error(request, "Rejection reason is required.")
    return redirect("dealer_approval_detail", dealer_id=dealer.id)


@login_required
@role_required(["ADMIN"])
def warehouse_create(request):
    if request.method == "POST":
        form = WarehouseForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Warehouse created successfully.")
            return redirect("warehouse_list")

    else:
        form = WarehouseForm()

    return render(
        request,
        "core/warehouse_form.html",
        {
            "form": form,
            "title": "Create Warehouse",
        }
    )


@login_required
def warehouse_list(request):
    role = get_user_role(request.user)

    if request.user.is_superuser or role == "ADMIN":
        warehouses = Warehouse.objects.all()

    elif role == "WAREHOUSE_MANAGER":
        warehouses = Warehouse.objects.filter(warehouse_manager=request.user)

    else:
        messages.error(request, "You do not have permission.")
        return redirect("dashboard")

    return render(
        request,
        "core/warehouse_list.html",
        {
            "warehouses": warehouses,
        }
    )


@login_required
@role_required(["DEVELOPMENT_OFFICER"])
def farmer_meet_create(request):
    if request.method == "POST":
        form = FarmerMeetCreateForm(request.POST)

        if form.is_valid():
            farmer_meet = form.save(commit=False)
            farmer_meet.created_by_mdo = request.user
            farmer_meet.approval_status = "PENDING_SALES_OFFICER"
            farmer_meet.save()

            FarmerMeetApprovalHistory.objects.create(
                farmer_meet=farmer_meet,
                action="Created and sent to Sales Officer",
                performed_by=request.user,
            )

            messages.success(request, "Farmer meeting request created successfully.")
            return redirect("farmer_meet_list")

    else:
        form = FarmerMeetCreateForm()

    return render(
        request,
        "core/farmer_meet_form.html",
        {
            "form": form,
            "title": "Create Farmer Meeting Request",
        }
    )

def farmer_meet_role_can_act(user, farmer_meet):
    role = get_user_role(user)

    if farmer_meet.approval_status in ["APPROVED", "REJECTED"]:
        return False

    if user.is_superuser or role == "ADMIN":
        return True

    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        return (
            farmer_meet.sales_officer_id == user.id and
            farmer_meet.approval_status == "PENDING_SALES_OFFICER"
        )

    if role == "ASM":
        return (
            farmer_meet.asm_id == user.id and
            farmer_meet.approval_status == "PENDING_ASM"
        )

    if role == "REGIONAL_MANAGER":
        return (
            farmer_meet.regional_manager_id == user.id and
            farmer_meet.approval_status == "PENDING_REGIONAL_MANAGER"
        )

    if role == "STATE_HEAD":
        return (
            farmer_meet.state_head_id == user.id and
            farmer_meet.approval_status == "PENDING_STATE_HEAD"
        )

    return False


def farmer_meet_user_can_view(user, farmer_meet):
    role = get_user_role(user)

    if user.is_superuser or role == "ADMIN":
        return True

    if role == "DEVELOPMENT_OFFICER":
        return farmer_meet.created_by_mdo_id == user.id

    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        return farmer_meet.sales_officer_id == user.id

    if role == "ASM":
        return farmer_meet.asm_id == user.id

    if role == "REGIONAL_MANAGER":
        return farmer_meet.regional_manager_id == user.id

    if role == "STATE_HEAD":
        return farmer_meet.state_head_id == user.id

    return False


def prepare_farmer_meet_ui(farmer_meet, user):
    farmer_meet.can_approve_now = farmer_meet_role_can_act(user, farmer_meet)

    if farmer_meet.approval_status == "APPROVED":
        farmer_meet.status_class = "approved"
        farmer_meet.status_text = "Approved / Visit Scheduled"

    elif farmer_meet.approval_status == "REJECTED":
        farmer_meet.status_class = "rejected"
        farmer_meet.status_text = "Rejected"

    elif farmer_meet.approval_status == "PENDING_SALES_OFFICER":
        farmer_meet.status_class = "pending"
        farmer_meet.status_text = "Pending Sales Officer"

    elif farmer_meet.approval_status == "PENDING_ASM":
        farmer_meet.status_class = "pending"
        farmer_meet.status_text = "Pending ASM"

    elif farmer_meet.approval_status == "PENDING_REGIONAL_MANAGER":
        farmer_meet.status_class = "forwarded"
        farmer_meet.status_text = "Pending Regional Manager"

    elif farmer_meet.approval_status == "PENDING_STATE_HEAD":
        farmer_meet.status_class = "forwarded"
        farmer_meet.status_text = "Pending State Head"

    else:
        farmer_meet.status_class = "pending"
        farmer_meet.status_text = farmer_meet.approval_status

    return farmer_meet



@login_required
def farmer_meet_list(request):
    role = get_user_role(request.user)
    status_filter = request.GET.get("status", "ALL").strip()

    farmer_meets = FarmerMeetRequest.objects.select_related(
        "created_by_mdo",
        "sales_officer",
        "asm",
        "regional_manager",
        "state_head",
        "rejected_by",
    )

    if role == "DEVELOPMENT_OFFICER":
        farmer_meets = farmer_meets.filter(created_by_mdo=request.user)

    elif role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        farmer_meets = farmer_meets.filter(sales_officer=request.user)

    elif role == "ASM":
        farmer_meets = farmer_meets.filter(asm=request.user)

    elif role == "REGIONAL_MANAGER":
        farmer_meets = farmer_meets.filter(regional_manager=request.user)

    elif role == "STATE_HEAD":
        farmer_meets = farmer_meets.filter(state_head=request.user)

    elif request.user.is_superuser or role == "ADMIN":
        farmer_meets = farmer_meets.all()

    else:
        messages.error(request, "You do not have permission.")
        return redirect("dashboard")

    if status_filter and status_filter != "ALL":
        farmer_meets = farmer_meets.filter(approval_status=status_filter)

    farmer_meets = farmer_meets.order_by("meeting_date", "-id")

    farmer_meet_list_data = []

    for meet in farmer_meets:
        farmer_meet_list_data.append(
            prepare_farmer_meet_ui(meet, request.user)
        )

    pending_count = farmer_meets.exclude(
        approval_status__in=["APPROVED", "REJECTED"]
    ).count()

    approved_count = farmer_meets.filter(
        approval_status="APPROVED"
    ).count()

    return render(
        request,
        "core/farmer_meet_list.html",
        {
            "farmer_meets": farmer_meet_list_data,
            "role": role,
            "status_filter": status_filter,
            "pending_count": pending_count,
            "approved_count": approved_count,
        }
    )

@login_required
def farmer_meet_detail(request, meet_id):
    farmer_meet = get_object_or_404(
        FarmerMeetRequest.objects.select_related(
            "created_by_mdo",
            "sales_officer",
            "asm",
            "regional_manager",
            "state_head",
            "rejected_by",
        ).prefetch_related("approval_history"),
        id=meet_id
    )

    if not farmer_meet_user_can_view(request.user, farmer_meet):
        messages.error(request, "You do not have permission to view this farmer meet.")
        return redirect("farmer_meet_list")

    farmer_meet = prepare_farmer_meet_ui(farmer_meet, request.user)

    return render(
        request,
        "core/farmer_meet_detail.html",
        {
            "farmer_meet": farmer_meet,
            "can_approve_now": farmer_meet.can_approve_now,
            "reject_form": FarmerMeetRejectForm(),
        }
    )


@login_required
@transaction.atomic
def farmer_meet_approve(request, meet_id):
    farmer_meet = get_object_or_404(FarmerMeetRequest, id=meet_id)

    if request.method != "POST":
        return redirect("farmer_meet_detail", meet_id=farmer_meet.id)

    if not farmer_meet_role_can_act(request.user, farmer_meet):
        messages.error(request, "You cannot approve this request now.")
        return redirect("farmer_meet_detail", meet_id=farmer_meet.id)

    old_status = farmer_meet.approval_status
    farmer_meet.approval_status = farmer_meet.get_next_status_after_approval()
    farmer_meet.save(update_fields=["approval_status"])

    FarmerMeetApprovalHistory.objects.create(
        farmer_meet=farmer_meet,
        action=f"Approved from {old_status} to {farmer_meet.approval_status}",
        performed_by=request.user,
    )

    messages.success(request, "Farmer meeting request approved.")
    return redirect("farmer_meet_list")


@login_required
@transaction.atomic
def farmer_meet_reject(request, meet_id):
    farmer_meet = get_object_or_404(FarmerMeetRequest, id=meet_id)

    if request.method != "POST":
        return redirect("farmer_meet_detail", meet_id=farmer_meet.id)

    if not farmer_meet_role_can_act(request.user, farmer_meet):
        messages.error(request, "You cannot reject this request now.")
        return redirect("farmer_meet_detail", meet_id=farmer_meet.id)

    form = FarmerMeetRejectForm(request.POST)

    if form.is_valid():
        reason = form.cleaned_data["rejection_reason"]

        farmer_meet.approval_status = "REJECTED"
        farmer_meet.rejected_by = request.user
        farmer_meet.rejected_at = timezone.now()
        farmer_meet.rejection_reason = reason
        farmer_meet.save(update_fields=[
            "approval_status",
            "rejected_by",
            "rejected_at",
            "rejection_reason",
        ])

        FarmerMeetApprovalHistory.objects.create(
            farmer_meet=farmer_meet,
            action="Rejected",
            performed_by=request.user,
            remarks=reason,
        )

        messages.success(request, "Farmer meeting request rejected.")
        return redirect("farmer_meet_list")

    messages.error(request, "Rejection reason is required.")
    return redirect("farmer_meet_detail", meet_id=farmer_meet.id)

    
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.safestring import mark_safe


import base64
import re
from io import BytesIO

import qrcode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.safestring import mark_safe

from .models import Product, ProductStickerSetting


def generate_qr_data_uri(data):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2,
    )
    qr.add_data(data)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    buffer = BytesIO()
    img.save(buffer, format="PNG")

    encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{encoded}"


def public_product_detail(request, product_id):
    product = get_object_or_404(
        Product.objects.select_related("category").prefetch_related(
            "pack_sizes",
            "pack_sizes__warehouse",
        ),
        id=product_id,
        is_active=True,
    )

    return render(
        request,
        "core/public_product_detail.html",
        {
            "product": product,
        }
    )


CODE39_PATTERNS = {
    "0": "nnnwwnwnn",
    "1": "wnnwnnnnw",
    "2": "nnwwnnnnw",
    "3": "wnwwnnnnn",
    "4": "nnnwwnnnw",
    "5": "wnnwwnnnn",
    "6": "nnwwwnnnn",
    "7": "nnnwnnwnw",
    "8": "wnnwnnwnn",
    "9": "nnwwnnwnn",
    "A": "wnnnnwnnw",
    "B": "nnwnnwnnw",
    "C": "wnwnnwnnn",
    "D": "nnnnwwnnw",
    "E": "wnnnwwnnn",
    "F": "nnwnwwnnn",
    "G": "nnnnnwwnw",
    "H": "wnnnnwwnn",
    "I": "nnwnnwwnn",
    "J": "nnnnwwwnn",
    "K": "wnnnnnnww",
    "L": "nnwnnnnww",
    "M": "wnwnnnnwn",
    "N": "nnnnwnnww",
    "O": "wnnnwnnwn",
    "P": "nnwnwnnwn",
    "Q": "nnnnnnwww",
    "R": "wnnnnnwwn",
    "S": "nnwnnnwwn",
    "T": "nnnnwnwwn",
    "U": "wwnnnnnnw",
    "V": "nwwnnnnnw",
    "W": "wwwnnnnnn",
    "X": "nwnnwnnnw",
    "Y": "wwnnwnnnn",
    "Z": "nwwnwnnnn",
    "-": "nwnnnnwnw",
    ".": "wwnnnnwnn",
    " ": "nwwnnnwnn",
    "$": "nwnwnwnnn",
    "/": "nwnwnnnwn",
    "+": "nwnnnwnwn",
    "%": "nnnwnwnwn",
    "*": "nwnnwnwnn",
}


def _clean_code39_value(value):
    value = str(value or "").upper().strip()
    value = re.sub(r"[^A-Z0-9\-\.\ \$\/\+\%]", "", value)

    if not value:
        value = "PRODUCT"

    return value[:28]


def _generate_code39_svg(value):
    value = _clean_code39_value(value)
    encoded = f"*{value}*"

    narrow = 1
    wide = 3
    gap = 1
    height = 42
    x = 0
    rects = []

    for char in encoded:
        pattern = CODE39_PATTERNS.get(char)

        if not pattern:
            continue

        for index, part in enumerate(pattern):
            width = wide if part == "w" else narrow

            if index % 2 == 0:
                rects.append(
                    f'<rect x="{x}" y="0" width="{width}" height="{height}" fill="#000"/>'
                )

            x += width

        x += gap

    svg = f"""
    <svg class="barcode-svg" viewBox="0 0 {x} {height}" preserveAspectRatio="none" xmlns="http://www.w3.org/2000/svg">
        {''.join(rects)}
    </svg>
    """

    return mark_safe(svg)


@login_required
def product_sticker_print(request, product_id):
    profile = getattr(request.user, "profile", None)

    allowed_print_roles = ["ADMIN", "INVENTORY_MANAGER"]

    if not request.user.is_superuser:
        if not profile or profile.role not in allowed_print_roles:
            messages.error(request, "Only Admin or Inventory Manager can print product stickers.")
            return redirect("dashboard")

    product = get_object_or_404(
        Product.objects.select_related("category").prefetch_related(
            "pack_sizes",
            "pack_sizes__warehouse",
        ),
        id=product_id,
    )

    pack_sizes = product.pack_sizes.select_related("warehouse").all().order_by(
        "warehouse__name",
        "pack_size",
    )

    setting, created = ProductStickerSetting.objects.get_or_create(
        product=product,
        defaults={
            "width_mm": 100,
            "height_mm": 75,
            "copies": 1,
        }
    )

    selected_pack = setting.pack_size or pack_sizes.first()

    if selected_pack:
        default_qty_per_case = f"{selected_pack.units_per_box} X {selected_pack.display_pack}"
        default_unit_mrp = selected_pack.mrp_per_unit
        default_case_mrp = selected_pack.box_sale_price
    else:
        default_qty_per_case = ""
        default_unit_mrp = ""
        default_case_mrp = ""

    if request.method == "POST":
        pack_id = request.POST.get("pack_id")

        if pack_id:
            selected_pack = pack_sizes.filter(id=pack_id).first()

        setting.pack_size = selected_pack
        setting.width_mm = request.POST.get("width_mm") or 100
        setting.height_mm = request.POST.get("height_mm") or 75
        setting.copies = request.POST.get("copies") or 1

        setting.company_name = request.POST.get("company_name") or setting.company_name
        setting.company_address = request.POST.get("company_address") or ""
        setting.company_email = request.POST.get("company_email") or ""
        setting.company_website = request.POST.get("company_website") or ""
        setting.customer_care = request.POST.get("customer_care", "").strip()
        setting.body_font_mm = request.POST.get("body_font_mm") or setting.body_font_mm or "2.25"

        setting.company_subtitle = request.POST.get("company_subtitle") or "PRODUCT LABEL / STICKER"

        setting.product_name = request.POST.get("product_name") or product.name
        setting.hsn_code = request.POST.get("hsn_code") or product.hsn_number or ""
        setting.qty_per_case = request.POST.get("qty_per_case") or default_qty_per_case

        setting.batch_no = request.POST.get("batch_no") or ""
        setting.mfg_date = request.POST.get("mfg_date") or ""
        setting.exp_date = request.POST.get("exp_date") or ""

        setting.usp_text = request.POST.get("usp_text") or ""
        setting.unit_mrp = request.POST.get("unit_mrp") or default_unit_mrp
        setting.case_mrp = request.POST.get("case_mrp") or default_case_mrp

        setting.barcode_value = _clean_code39_value(
            request.POST.get("barcode_value") or f"PRD{product.id:04d}"
        )

        setting.manufactured_by = request.POST.get("manufactured_by") or setting.company_name
        setting.marketed_by = request.POST.get("marketed_by") or setting.company_name

        setting.footer_text = request.POST.get("footer_text") or "Scan QR for product details."

        setting.show_barcode = True if request.POST.get("show_barcode") == "on" else False
        setting.show_qr = True if request.POST.get("show_qr") == "on" else False

        setting.save()

        messages.success(request, "Sticker settings saved successfully.")
        return redirect("product_sticker_print", product_id=product.id)

    product_name = setting.product_name or product.name
    hsn_code = setting.hsn_code or product.hsn_number or ""
    qty_per_case = setting.qty_per_case or default_qty_per_case
    unit_mrp = setting.unit_mrp or default_unit_mrp
    case_mrp = setting.case_mrp or default_case_mrp

    barcode_value = _clean_code39_value(setting.barcode_value or f"PRD{product.id:04d}")
    barcode_svg = _generate_code39_svg(barcode_value)

    product_detail_url = request.build_absolute_uri(
        reverse("public_product_detail", args=[product.id])
    )
    qr_code_data_uri = generate_qr_data_uri(product_detail_url)

    copies = int(setting.copies or 1)

    context = {
        "product": product,
        "pack_sizes": pack_sizes,
        "selected_pack": selected_pack,
        "setting": setting,

        "width_mm": setting.width_mm or 100,
        "height_mm": setting.height_mm or 75,
        "copies": copies,
        "copies_range": range(copies),

        "company_name": setting.company_name,
        "company_address": setting.company_address,
        "company_email": setting.company_email,
        "company_website": setting.company_website,
        "customer_care": setting.customer_care,
        "body_font_mm": setting.body_font_mm or "2.25",
        "company_subtitle": setting.company_subtitle,

        "product_name": product_name,
        "hsn_code": hsn_code,
        "qty_per_case": qty_per_case,
        "batch_no": setting.batch_no,
        "mfg_date": setting.mfg_date,
        "exp_date": setting.exp_date,
        "usp_text": setting.usp_text,
        "unit_mrp": unit_mrp,
        "case_mrp": case_mrp,

        "barcode_value": barcode_value,
        "barcode_svg": barcode_svg,

        "manufactured_by": setting.manufactured_by,
        "marketed_by": setting.marketed_by,
        "footer_text": setting.footer_text,

        "show_barcode": setting.show_barcode,
        "show_qr": setting.show_qr,

        "product_detail_url": product_detail_url,
        "qr_code_data_uri": qr_code_data_uri,
    }

    return render(request, "core/product_sticker_print.html", context)



@login_required
def product_category_list(request):
    role = user_role(request.user)

    if not request.user.is_superuser and role != "ADMIN":
        messages.error(request, "Only admin can view product categories.")
        return redirect("dashboard")

    categories = ProductCategory.objects.prefetch_related("payment_rules").order_by("name")

    def get_first_value(obj, names, default=None):
        for name in names:
            if hasattr(obj, name):
                value = getattr(obj, name)
                if value is not None and value != "":
                    return value
        return default

    def get_rule_type(rule):
        if hasattr(rule, "get_rule_type_display"):
            return rule.get_rule_type_display()

        if hasattr(rule, "get_payment_type_display"):
            return rule.get_payment_type_display()

        return get_first_value(
            rule,
            ["rule_type", "payment_type", "payment_option", "type"],
            "Payment Rule"
        )

    def get_percentage(rule):
        value = get_first_value(
            rule,
            [
                "percentage",
                "discount_percent",
                "discount_percentage",
                "discount",
                "discount_value",
            ],
            None
        )

        if value is None:
            return "-"

        return value

    def get_rule_name(rule):
        value = get_first_value(
            rule,
            ["rule_name", "name", "title"],
            None
        )

        if value:
            return value

        return get_rule_type(rule)

    def get_days(rule):
        from_day = get_first_value(
            rule,
            ["from_day", "start_day", "min_days", "days_from", "day_from"],
            None
        )

        to_day = get_first_value(
            rule,
            ["to_day", "end_day", "max_days", "days_to", "day_to"],
            None
        )

        rule_type_text = str(get_rule_type(rule)).lower()

        if from_day is None and to_day is None:
            if "instant" in rule_type_text:
                return "Same day"
            if "advance" in rule_type_text:
                return "Before / on order date"
            return "-"

        if from_day is None:
            return f"Up to {to_day} days"

        if to_day is None:
            return f"From {from_day} days"

        return f"{from_day} to {to_day} days"

    category_rows = []

    for category in categories:
        rule_rows = []

        for rule in category.payment_rules.all():
            rule_type = get_rule_type(rule)
            percentage = get_percentage(rule)

            rule_rows.append({
                "rule_name": get_rule_name(rule),
                "rule_type": rule_type,
                "percentage": percentage,
                "days": get_days(rule),
                "is_active": getattr(rule, "is_active", True),
                "is_interest": str(rule_type).upper() == "INTEREST",
            })

        category_rows.append({
            "category": category,
            "rules": rule_rows,
        })

    return render(
        request,
        "core/product_category_list.html",
        {
            "category_rows": category_rows,
        }
    )



@login_required
@role_required(["ADMIN"])
@transaction.atomic
def product_category_create(request):
    category = ProductCategory()

    if request.method == "POST":
        form = ProductCategoryForm(request.POST, instance=category)
        formset = CategoryPaymentRuleFormSet(request.POST, instance=category)

        if form.is_valid() and formset.is_valid():
            category = form.save()
            formset.instance = category
            formset.save()

            messages.success(request, "Category and payment rules saved successfully.")
            return redirect("product_category_create")
    else:
        form = ProductCategoryForm(instance=category)
        formset = CategoryPaymentRuleFormSet(instance=category)

    return render(
        request,
        "core/product_category_form.html",
        {
            "form": form,
            "formset": formset,
            "title": "Create Product Category",
        }
    )


@login_required
def product_category_edit(request, pk):
    role = user_role(request.user)

    if not request.user.is_superuser and role != "ADMIN":
        messages.error(request, "Only admin can edit product categories.")
        return redirect("dashboard")

    category = get_object_or_404(ProductCategory, pk=pk)

    if request.method == "POST":
        form = ProductCategoryForm(request.POST, instance=category)
        formset = CategoryPaymentRuleFormSet(request.POST, instance=category)

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()

            messages.success(request, "Category and payment rules updated successfully.")
            return redirect("product_category_list")
    else:
        form = ProductCategoryForm(instance=category)
        formset = CategoryPaymentRuleFormSet(instance=category)

    return render(
        request,
        "core/product_category_form.html",
        {
            "form": form,
            "formset": formset,
            "title": "Edit Product Category",
            "button_text": "Update Category",
            "category": category,
        }
    )


@login_required
@role_required(["ADMIN"])
@transaction.atomic
def product_create(request):
    product = Product()

    if request.method == "POST":
        form = ProductForm(
            request.POST,
            request.FILES,
            instance=product
        )

        pack_formset = ProductPackSizeFormSet(
            request.POST,
            instance=product,
            prefix="packs"
        )

        scheme_formset = ProductSchemeFormSet(
            request.POST,
            instance=product,
            prefix="schemes"
        )

        if form.is_valid() and pack_formset.is_valid() and scheme_formset.is_valid():
            product = form.save()

            pack_formset.instance = product
            pack_formset.save()

            scheme_formset.instance = product
            scheme_formset.save()

            messages.success(request, "Product created successfully.")
            return redirect("product_list")
        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = ProductForm(instance=product)

        pack_formset = ProductPackSizeFormSet(
            instance=product,
            prefix="packs"
        )

        scheme_formset = ProductSchemeFormSet(
            instance=product,
            prefix="schemes"
        )

    return render(
        request,
        "core/product_form.html",
        {
            "form": form,
            "pack_formset": pack_formset,
            "scheme_formset": scheme_formset,
            "title": "Create Product",
            "button_text": "Save Product",
        }
    )

@login_required
@role_required(["ADMIN", "INVENTORY_MANAGER"])
def product_list(request):
    search = request.GET.get("q", "").strip()

    profile = getattr(request.user, "profile", None)
    role = getattr(profile, "role", None)

    can_manage_products = request.user.is_superuser or role == "ADMIN"
    can_print_stickers = request.user.is_superuser or role in ["ADMIN", "INVENTORY_MANAGER"]

    products = Product.objects.select_related("category").prefetch_related(
        "pack_sizes",
        "pack_sizes__warehouse",
        "schemes"
    ).order_by("name")

    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(category__name__icontains=search) |
            Q(hsn_number__icontains=search)
        )

    return render(
        request,
        "core/product_list.html",
        {
            "products": products,
            "search": search,
            "can_manage_products": can_manage_products,
            "can_print_stickers": can_print_stickers,
        }
    )


@login_required
def dealer_product_detail(request, product_id):
    role = get_user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealers can access product catalogue.")
        return redirect("dashboard")

    dealer = Dealer.objects.filter(
        user=request.user,
        approval_status="APPROVED",
        is_active=True
    ).first()

    if not dealer:
        messages.error(request, "Your dealer profile is not approved yet.")
        return redirect("dashboard")

    product = get_object_or_404(
        Product.objects.filter(
            is_active=True,
            pack_sizes__is_active=True,
            pack_sizes__stock_boxes__gt=0
        ).select_related(
            "category"
        ).prefetch_related(
            Prefetch(
                "pack_sizes",
                queryset=ProductPackSize.objects.filter(
                    is_active=True,
                    stock_boxes__gt=0
                ).select_related("warehouse")
            ),
            Prefetch(
                "schemes",
                queryset=ProductScheme.objects.filter(is_active=True)
            ),
            "category__payment_rules"
        ).distinct(),
        id=product_id
    )

    return render(
        request,
        "core/dealer_product_detail.html",
        {
            "product": product,
            "dealer": dealer,
        }
    )


@login_required
@role_required(["ADMIN"])
@transaction.atomic
def product_edit(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == "POST":
        form = ProductForm(
            request.POST,
            request.FILES,
            instance=product
        )

        pack_formset = ProductPackSizeFormSet(
            request.POST,
            instance=product,
            prefix="packs"
        )

        scheme_formset = ProductSchemeFormSet(
            request.POST,
            instance=product,
            prefix="schemes"
        )

        if form.is_valid() and pack_formset.is_valid() and scheme_formset.is_valid():
            product = form.save()

            # Save edited/new pack sizes safely
            pack_instances = pack_formset.save(commit=False)

            for pack in pack_instances:
                pack.product = product
                pack.save()

            # Safe delete handling for old pack sizes
            for pack in pack_formset.deleted_objects:
                used_reasons = []

                if hasattr(pack, "production_recipes") and pack.production_recipes.exists():
                    used_reasons.append("Production Recipe")

                if hasattr(pack, "production_outputs") and pack.production_outputs.exists():
                    used_reasons.append("Production Batch")

                if hasattr(pack, "repacking_outputs") and pack.repacking_outputs.exists():
                    used_reasons.append("Repacking Batch")

                if used_reasons:
                    # Do not delete linked pack size. Just deactivate if field exists.
                    if hasattr(pack, "is_active"):
                        pack.is_active = False
                        pack.save(update_fields=["is_active"])

                        messages.warning(
                            request,
                            f"Pack size '{pack}' is already used in {', '.join(used_reasons)}. "
                            f"So it was not deleted, only deactivated."
                        )
                    else:
                        messages.warning(
                            request,
                            f"Pack size '{pack}' is already used in {', '.join(used_reasons)}. "
                            f"So it cannot be deleted."
                        )
                else:
                    pack.delete()

            if hasattr(pack_formset, "save_m2m"):
                pack_formset.save_m2m()

            scheme_formset.instance = product
            scheme_formset.save()

            messages.success(request, "Product updated successfully.")
            return redirect("product_list")

        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = ProductForm(instance=product)

        pack_formset = ProductPackSizeFormSet(
            instance=product,
            prefix="packs"
        )

        scheme_formset = ProductSchemeFormSet(
            instance=product,
            prefix="schemes"
        )

    return render(
        request,
        "core/product_form.html",
        {
            "form": form,
            "pack_formset": pack_formset,
            "scheme_formset": scheme_formset,
            "title": "Edit Product",
            "button_text": "Update Product",
            "product": product,
        }
    )

@login_required
@role_required(["ADMIN"])
@transaction.atomic
def product_deactivate(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method != "POST":
        return redirect("product_list")

    product.is_active = False
    product.save(update_fields=["is_active"])

    product.pack_sizes.update(is_active=False)

    messages.success(request, f"{product.name} deactivated successfully.")
    return redirect("product_list")


@login_required
@role_required(["ADMIN"])
def product_force_delete(request, product_id):
    product = get_object_or_404(
        Product.objects.prefetch_related("pack_sizes"),
        id=product_id
    )

    packs = list(product.pack_sizes.all())

    delete_summary = {
        "pack_count": len(packs),
        "cart_items": DealerCartItem.objects.filter(product_pack__in=packs).count(),
        "order_items": DealerOrderItem.objects.filter(product_pack__in=packs).count(),
        "sales_return_items": SalesReturnCreditNoteItem.objects.filter(product_pack__in=packs).count(),
        "sales_return_order_items": SalesReturnCreditNoteItem.objects.filter(order_item__product_pack__in=packs).count(),
        "old_production_recipes": ProductionRecipe.objects.filter(output_product_pack__in=packs).count(),
        "old_production_batches": ProductionBatch.objects.filter(output_product_pack__in=packs).count(),
        "repacking_batches": RepackingBatch.objects.filter(destination_product_pack__in=packs).count(),
        "smart_production_runs": ProductProductionRun.objects.filter(product=product).count(),
        "smart_production_outputs": ProductProductionRunOutputPack.objects.filter(product_pack__in=packs).count(),
        "smart_formulas": ProductProductionFormula.objects.filter(product=product).count(),
    }

    if request.method == "POST":
        confirm_text = (request.POST.get("confirm_text") or "").strip().upper()

        if confirm_text != "DELETE":
            messages.error(request, "Please type DELETE to confirm product delete.")
            return redirect("product_force_delete", product_id=product.id)

        try:
            with transaction.atomic():
                packs = list(product.pack_sizes.select_for_update().all())

                # 1. Remove cart items
                DealerCartItem.objects.filter(product_pack__in=packs).delete()

                # 2. Remove sales return item links before order items
                SalesReturnCreditNoteItem.objects.filter(product_pack__in=packs).delete()
                SalesReturnCreditNoteItem.objects.filter(order_item__product_pack__in=packs).delete()

                # 3. Remove dealer order items using this product pack
                DealerOrderItem.objects.filter(product_pack__in=packs).delete()

                # 4. Remove smart production output rows
                ProductProductionRunOutputPack.objects.filter(product_pack__in=packs).delete()

                # 5. Remove smart production runs/formulas for this product
                ProductProductionRun.objects.filter(product=product).delete()
                ProductProductionFormula.objects.filter(product=product).delete()

                # 6. Remove old production/repacking references
                ProductionBatch.objects.filter(output_product_pack__in=packs).delete()
                RepackingBatch.objects.filter(destination_product_pack__in=packs).delete()
                ProductionRecipe.objects.filter(output_product_pack__in=packs).delete()

                # 7. Delete pack sizes
                product.pack_sizes.all().delete()

                product_name = product.name

                # 8. Delete product
                product.delete()

            messages.success(request, f"{product_name} deleted successfully.")
            return redirect("product_list")

        except ProtectedError as error:
            messages.error(request, f"Product could not be deleted because it is still protected: {error}")
            return redirect("product_force_delete", product_id=product.id)

        except Exception as error:
            messages.error(request, f"Product could not be deleted: {error}")
            return redirect("product_force_delete", product_id=product.id)

    return render(request, "core/product_force_delete_confirm.html", {
        "product": product,
        "delete_summary": delete_summary,
    })
    

@login_required
def dealer_products(request):
    role = get_user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealers can access product catalogue.")
        return redirect("dashboard")

    dealer = Dealer.objects.filter(
        user=request.user,
        approval_status="APPROVED",
        is_active=True
    ).first()

    if not dealer:
        messages.error(request, "Your dealer profile is not approved yet.")
        return redirect("dashboard")

    selected_category = request.GET.get("category", "").strip()
    search = request.GET.get("q", "").strip()

    categories = ProductCategory.objects.filter(
        is_active=True,
        products__is_active=True
    ).distinct().order_by("name")

    products = Product.objects.filter(
        is_active=True,
        pack_sizes__is_active=True,
        pack_sizes__stock_boxes__gt=0
    ).select_related(
        "category"
    ).prefetch_related(
        Prefetch(
            "pack_sizes",
            queryset=ProductPackSize.objects.filter(
                is_active=True,
                stock_boxes__gt=0
            ).select_related("warehouse")
        ),
        Prefetch(
            "schemes",
            queryset=ProductScheme.objects.filter(is_active=True)
        ),
        "category__payment_rules"
    ).distinct()

    if selected_category:
        products = products.filter(category_id=selected_category)

    if search:
        products = products.filter(name__icontains=search)

    return render(
        request,
        "core/dealer_products.html",
        {
            "products": products,
            "dealer": dealer,
            "categories": categories,
            "selected_category": selected_category,
            "search": search,
        }
    )

@login_required
def dealer_add_to_cart(request, pack_id):
    role = get_user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealers can add products to cart.")
        return redirect("dashboard")

    dealer = Dealer.objects.filter(
        user=request.user,
        approval_status="APPROVED",
        is_active=True
    ).first()

    if not dealer:
        messages.error(request, "Your dealer profile is not approved yet.")
        return redirect("dashboard")

    pack = get_object_or_404(
        ProductPackSize,
        id=pack_id,
        is_active=True,
        stock_boxes__gt=0
    )

    quantity_boxes = int(request.POST.get("quantity_boxes", 1))

    if quantity_boxes <= 0:
        quantity_boxes = 1

    if quantity_boxes > pack.stock_boxes:
        messages.error(request, "Selected quantity is more than available stock.")
        return redirect("dealer_products")

    cart, created = DealerCart.objects.get_or_create(dealer=request.user)

    item, created = DealerCartItem.objects.get_or_create(
        cart=cart,
        product_pack=pack,
        defaults={"quantity_boxes": quantity_boxes}
    )

    if not created:
        item.quantity_boxes += quantity_boxes

        if item.quantity_boxes > pack.stock_boxes:
            item.quantity_boxes = pack.stock_boxes

        item.save()

    messages.success(request, "Product added to cart.")
    return redirect("dealer_products")


@login_required
def dealer_cart(request):
    role = get_user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealers can access cart.")
        return redirect("dashboard")

    cart, created = DealerCart.objects.get_or_create(dealer=request.user)

    cart_items = cart.items.select_related(
        "product_pack",
        "product_pack__product",
        "product_pack__warehouse"
    )

    return render(
        request,
        "core/dealer_cart.html",
        {
            "cart": cart,
            "cart_items": cart_items,
        }
    )

@login_required
def dealer_increase_cart_item(request, item_id):
    item = get_object_or_404(
        DealerCartItem,
        id=item_id,
        cart__dealer=request.user
    )

    if request.method == "POST":
        item.quantity_boxes += 1
        item.save()

    return redirect("dealer_cart")


@login_required
def dealer_decrease_cart_item(request, item_id):
    item = get_object_or_404(
        DealerCartItem,
        id=item_id,
        cart__dealer=request.user
    )

    if request.method == "POST":
        if item.quantity_boxes > 1:
            item.quantity_boxes -= 1
            item.save()
        else:
            item.delete()

    return redirect("dealer_cart")

@login_required
def dealer_remove_cart_item(request, item_id):
    item = get_object_or_404(
        DealerCartItem,
        id=item_id,
        cart__dealer=request.user
    )

    item.delete()

    messages.success(request, "Cart item removed.")
    return redirect("dealer_cart")



def user_role(user):
    if user.is_superuser:
        return "ADMIN"

    profile = getattr(user, "profile", None)

    if profile:
        return profile.role

    return None


def log_order_action(order, user, action, note=""):
    DealerOrderApprovalLog.objects.create(
        order=order,
        user=user,
        role=user_role(user),
        action=action,
        note=note,
    )


def get_rsm_from_asm(asm_user):
    if not asm_user:
        return None

    asm_profile = getattr(asm_user, "profile", None)

    if asm_profile and asm_profile.manager:
        manager_profile = getattr(asm_profile.manager, "profile", None)

        if manager_profile and manager_profile.role == "REGIONAL_MANAGER":
            return asm_profile.manager

    return None


def get_warehouse_manager_from_order(order):
    first_item = order.items.select_related(
        "product_pack",
        "product_pack__warehouse"
    ).first()

    if not first_item:
        return None

    warehouse = first_item.product_pack.warehouse

    if hasattr(warehouse, "manager") and warehouse.manager:
        return warehouse.manager

    if hasattr(warehouse, "warehouse_manager") and warehouse.warehouse_manager:
        return warehouse.warehouse_manager

    return None


def recalculate_order_totals(order):
    subtotal = Decimal("0.00")
    gst_total = Decimal("0.00")
    grand_total = Decimal("0.00")

    for item in order.items.select_related("product_pack").all():
        item.calculate()
        item.save()

        subtotal += item.taxable_amount
        gst_total += item.gst_amount
        grand_total += item.total_amount

    order.subtotal_amount = subtotal
    order.gst_amount = gst_total
    order.total_amount = grand_total
    order.save(update_fields=["subtotal_amount", "gst_amount", "total_amount", "updated_at"])


def get_order_progress(order):
    if order.status == "REJECTED":
        return [
            {"title": "Order Placed", "done": True},
            {"title": "Rejected", "done": True, "rejected": True},
        ]

    return [
        {
            "title": "Order Placed",
            "done": order.status in [
                "PLACED",
                "SALES_APPROVAL",
                "ASM_APPROVAL",
                "RSM_APPROVAL",
                "ACCOUNTANT_ORDER_APPROVAL",
                "WAREHOUSE_REVIEW",
                "ACCOUNTANT_INVOICE_REVIEW",
                "INVOICE_RELEASED",
                "DISPATCH_PENDING",
                "DISPATCHED",
                "DELIVERED",
            ],
        },
        {
            "title": "Processing",
            "done": order.status in [
                "WAREHOUSE_REVIEW",
                "ACCOUNTANT_INVOICE_REVIEW",
                "INVOICE_RELEASED",
                "DISPATCH_PENDING",
                "DISPATCHED",
                "DELIVERED",
            ],
        },
        {
            "title": "Invoice Released",
            "done": order.status in [
                "INVOICE_RELEASED",
                "DISPATCH_PENDING",
                "DISPATCHED",
                "DELIVERED",
            ],
        },
        {
            "title": "Dispatched",
            "done": order.status in ["DISPATCHED", "DELIVERED"],
        },
        {
            "title": "Delivered",
            "done": order.status == "DELIVERED",
        },
    ]


def get_order_permissions(order, user):
    role = user_role(user)

    permissions = {
        "can_approve": False,
        "can_reject": False,
        "can_warehouse_review": False,
        "can_accountant_invoice_review": False,
        "can_dispatch": False,
        "can_accept_delivery": False,
        "can_view_dealer_copy": False,
        "can_view_company_copy": False,
        "can_view_transporter_copy": False,
    }

    if role == "ADMIN" or user.is_superuser:
        for key in permissions:
            permissions[key] = True

        if order.status != "DISPATCHED":
            permissions["can_accept_delivery"] = False

        return permissions

    if role == "DEALER":
        permissions["can_view_dealer_copy"] = True

        if order.dealer_user == user and order.status == "DISPATCHED":
            permissions["can_accept_delivery"] = True

        return permissions

    if order.status == "SALES_APPROVAL" and role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        if order.concerned_sales_officer == user:
            permissions["can_approve"] = True
            permissions["can_reject"] = True

    if order.status == "ASM_APPROVAL" and role == "ASM":
        if order.concerned_asm == user:
            permissions["can_approve"] = True
            permissions["can_reject"] = True

    if order.status == "RSM_APPROVAL" and role == "REGIONAL_MANAGER":
        if order.concerned_rsm == user:
            permissions["can_approve"] = True
            permissions["can_reject"] = True

    if role == "ACCOUNTANT":
        permissions["can_view_dealer_copy"] = True
        permissions["can_view_company_copy"] = True
        permissions["can_view_transporter_copy"] = True

        if order.status == "ACCOUNTANT_ORDER_APPROVAL":
            permissions["can_approve"] = True
            permissions["can_reject"] = True

        if order.status == "ACCOUNTANT_INVOICE_REVIEW":
            permissions["can_accountant_invoice_review"] = True

    if role == "WAREHOUSE_MANAGER":
        permissions["can_view_company_copy"] = True
        permissions["can_view_transporter_copy"] = True

        if order.status == "WAREHOUSE_REVIEW":
            if not order.warehouse_manager or order.warehouse_manager == user:
                permissions["can_warehouse_review"] = True

        if order.status == "DISPATCH_PENDING":
            if not order.warehouse_manager or order.warehouse_manager == user:
                permissions["can_dispatch"] = True

    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        if order.concerned_sales_officer == user:
            permissions["can_view_dealer_copy"] = True

    if role == "ASM":
        if order.concerned_asm == user:
            permissions["can_view_dealer_copy"] = True

    if role == "REGIONAL_MANAGER":
        if order.concerned_rsm == user:
            permissions["can_view_dealer_copy"] = True

    return permissions

@login_required
@transaction.atomic
def dealer_place_order(request):
    role = user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealer can place order.")
        return redirect("dashboard")

    dealer = Dealer.objects.filter(
        user=request.user,
        approval_status="APPROVED",
        is_active=True
    ).first()

    if not dealer:
        messages.error(request, "Dealer profile is not approved.")
        return redirect("dashboard")

    cart = DealerCart.objects.filter(dealer=request.user).first()

    if not cart or not cart.items.exists():
        messages.error(request, "Cart is empty.")
        return redirect("dealer_cart")

    dealer_flag = getattr(dealer, "flag", "GREEN")

    concerned_sales_officer = getattr(dealer, "created_by_sales_officer", None)
    concerned_asm = getattr(dealer, "concerned_asm", None)
    concerned_rsm = get_rsm_from_asm(concerned_asm)

    if dealer_flag == "GREEN":
        first_status = "ASM_APPROVAL"
    else:
        first_status = "SALES_APPROVAL"

    order = DealerOrder.objects.create(
        dealer=dealer,
        dealer_user=request.user,
        dealer_flag=dealer_flag,
        status=first_status,
        concerned_sales_officer=concerned_sales_officer,
        concerned_asm=concerned_asm,
        concerned_rsm=concerned_rsm,
    )

    for cart_item in cart.items.select_related(
        "product_pack",
        "product_pack__product"
    ).all():
        pack = cart_item.product_pack

        order_item = DealerOrderItem.objects.create(
            order=order,
            product_pack=pack,
            product_name=pack.product.name,
            pack_label=pack.display_pack,
            quantity_boxes=cart_item.quantity_boxes,
            units_per_box=pack.units_per_box,
            unit_name=pack.get_unit_display(),
            hsn_code=pack.product.hsn_number,
            gst_percent=Decimal("5.00"),
        )

        order_item.calculate()
        order_item.save()

    recalculate_order_totals(order)

    cart.items.all().delete()

    log_order_action(order, request.user, "CREATED", "Dealer placed order.")

    messages.success(request, "Order placed successfully.")
    return redirect("dealer_order_detail", order_id=order.id)


@login_required
def dealer_order_list(request):
    role = user_role(request.user)

    orders = DealerOrder.objects.select_related(
        "dealer",
        "dealer_user",
        "concerned_sales_officer",
        "concerned_asm",
        "concerned_rsm",
    ).prefetch_related("items")

    if role == "DEALER":
        orders = orders.filter(dealer_user=request.user)

    elif role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        orders = orders.filter(concerned_sales_officer=request.user)

    elif role == "ASM":
        orders = orders.filter(concerned_asm=request.user)

    elif role == "REGIONAL_MANAGER":
        orders = orders.filter(concerned_rsm=request.user)

    elif role == "WAREHOUSE_MANAGER":
        orders = orders.filter(
            Q(status="WAREHOUSE_REVIEW") |
            Q(status="DISPATCH_PENDING") |
            Q(status="DISPATCHED")
        )

    elif role == "ACCOUNTANT":
        orders = orders.filter(
            Q(status="ACCOUNTANT_ORDER_APPROVAL") |
            Q(status="ACCOUNTANT_INVOICE_REVIEW") |
            Q(status="INVOICE_RELEASED") |
            Q(status="DISPATCH_PENDING")
        )

    elif role in ["ADMIN", "STATE_HEAD"]:
        pass

    else:
        orders = DealerOrder.objects.none()

    return render(
        request,
        "core/dealer_order_list.html",
        {
            "orders": orders,
            "role": role,
        }
    )


@login_required
def dealer_order_detail(request, order_id):
    order = get_object_or_404(
        DealerOrder.objects.select_related(
            "dealer",
            "dealer_user",
            "concerned_sales_officer",
            "concerned_asm",
            "concerned_rsm",
            "warehouse_manager",
        ).prefetch_related(
            "items",
            "approval_logs"
        ),
        id=order_id
    )

    role = user_role(request.user)

    if role == "DEALER" and order.dealer_user != request.user:
        messages.error(request, "You cannot access this order.")
        return redirect("dealer_order_list")

    progress = get_order_progress(order)
    permissions = get_order_permissions(order, request.user)

    return render(
        request,
        "core/dealer_order_detail.html",
        {
            "order": order,
            "progress": progress,
            "role": role,
            "permissions": permissions,
        }
    )

@login_required
@transaction.atomic
def dealer_order_approve(request, order_id):
    order = get_object_or_404(DealerOrder, id=order_id)
    role = user_role(request.user)

    allowed = False

    if order.status == "SALES_APPROVAL" and role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        allowed = order.concerned_sales_officer == request.user

    elif order.status == "ASM_APPROVAL" and role == "ASM":
        allowed = order.concerned_asm == request.user

    elif order.status == "RSM_APPROVAL" and role == "REGIONAL_MANAGER":
        allowed = order.concerned_rsm == request.user

    elif order.status == "ACCOUNTANT_ORDER_APPROVAL" and role == "ACCOUNTANT":
        allowed = True

    elif role == "ADMIN":
        allowed = True

    if not allowed:
        messages.error(request, "You are not allowed to approve this order.")
        return redirect("dealer_order_detail", order_id=order.id)

    if order.status == "SALES_APPROVAL":
        order.status = "ASM_APPROVAL"
        order.sales_approved_at = timezone.now()

    elif order.status == "ASM_APPROVAL":
        order.asm_approved_at = timezone.now()

        if order.dealer_flag == "GREEN":
            order.status = "WAREHOUSE_REVIEW"
        else:
            order.status = "RSM_APPROVAL"

    elif order.status == "RSM_APPROVAL":
        order.rsm_approved_at = timezone.now()

        if order.dealer_flag == "ORANGE":
            order.status = "WAREHOUSE_REVIEW"
        else:
            order.status = "ACCOUNTANT_ORDER_APPROVAL"

    elif order.status == "ACCOUNTANT_ORDER_APPROVAL":
        order.accountant_order_approved_at = timezone.now()
        order.status = "WAREHOUSE_REVIEW"

    order.save()

    if order.status == "WAREHOUSE_REVIEW":
        order.warehouse_manager = get_warehouse_manager_from_order(order)
        order.save(update_fields=["warehouse_manager", "updated_at"])

    log_order_action(order, request.user, "APPROVED", f"{role} approved order.")

    messages.success(request, "Order approved successfully.")
    return redirect("dealer_order_detail", order_id=order.id)


@login_required
@transaction.atomic
def dealer_order_reject(request, order_id):
    order = get_object_or_404(DealerOrder, id=order_id)

    if request.method == "POST":
        form = OrderRejectForm(request.POST)

        if form.is_valid():
            order.reject_order(
                request.user,
                form.cleaned_data["reason"]
            )

            log_order_action(
                order,
                request.user,
                "REJECTED",
                form.cleaned_data["reason"]
            )

            messages.success(request, "Order rejected.")
            return redirect("dealer_order_detail", order_id=order.id)

    else:
        form = OrderRejectForm()

    return render(
        request,
        "core/dealer_order_reject.html",
        {
            "form": form,
            "order": order,
        }
    )

@login_required
@transaction.atomic
def dealer_accept_delivery(request, order_id):
    order = get_object_or_404(DealerOrder, id=order_id)
    role = user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealer can accept delivery.")
        return redirect("dealer_order_detail", order_id=order.id)

    if order.dealer_user != request.user:
        messages.error(request, "You cannot accept this delivery.")
        return redirect("dealer_order_list")

    if order.status != "DISPATCHED":
        messages.error(request, "Delivery can be accepted only after dispatch.")
        return redirect("dealer_order_detail", order_id=order.id)

    if request.method == "POST":
        order.status = "DELIVERED"
        order.delivered_at = timezone.now()
        order.save(update_fields=["status", "delivered_at", "updated_at"])

        log_order_action(
            order,
            request.user,
            "DELIVERED",
            "Dealer accepted delivery."
        )

        messages.success(request, "Delivery accepted successfully.")
        return redirect("dealer_order_detail", order_id=order.id)

    return render(
        request,
        "core/dealer_accept_delivery.html",
        {
            "order": order,
        }
    )


@login_required
@transaction.atomic
def warehouse_order_review(request, order_id):
    order = get_object_or_404(DealerOrder, id=order_id)

    role = user_role(request.user)

    if role not in ["WAREHOUSE_MANAGER", "ADMIN"]:
        messages.error(request, "Only warehouse manager can review order.")
        return redirect("dealer_order_detail", order_id=order.id)

    if order.status != "WAREHOUSE_REVIEW":
        messages.error(request, "Order is not in warehouse review stage.")
        return redirect("dealer_order_detail", order_id=order.id)

    if request.method == "POST":
        formset = WarehouseOrderItemFormSet(request.POST, instance=order)

        if formset.is_valid():
            items = formset.save(commit=False)

            for item in items:
                item.warehouse_checked = True
                item.save()

            order.status = "ACCOUNTANT_INVOICE_REVIEW"
            order.warehouse_reviewed_at = timezone.now()
            order.save()

            log_order_action(order, request.user, "FORWARDED", "Warehouse reviewed and sent to accountant.")

            messages.success(request, "Order sent to accountant for invoice review.")
            return redirect("dealer_order_detail", order_id=order.id)

    else:
        formset = WarehouseOrderItemFormSet(instance=order)

    return render(
        request,
        "core/warehouse_order_review.html",
        {
            "order": order,
            "formset": formset,
        }
    )


@login_required
@transaction.atomic
def accountant_invoice_review(request, order_id):
    order = get_object_or_404(
        DealerOrder.objects.prefetch_related("items"),
        id=order_id
    )

    role = user_role(request.user)

    if role not in ["ACCOUNTANT", "ADMIN"]:
        messages.error(request, "Only accountant can release invoice.")
        return redirect("dealer_order_detail", order_id=order.id)

    if order.status != "ACCOUNTANT_INVOICE_REVIEW":
        messages.error(request, "Order is not ready for invoice review.")
        return redirect("dealer_order_detail", order_id=order.id)

    if request.method == "POST":
        formset = AccountantOrderItemFormSet(request.POST, instance=order)
        release_form = AccountantInvoiceReleaseForm(request.POST)

        if formset.is_valid() and release_form.is_valid():
            items = formset.save(commit=False)

            for item in items:
                item.accountant_checked = True
                item.calculate()
                item.save()

            recalculate_order_totals(order)

            fy = release_form.cleaned_data["financial_year"]

            sequence, created = InvoiceNumberSequence.objects.get_or_create(
                financial_year=fy
            )

            invoice_number = sequence.next_invoice_number()

            gstr1_code = release_form.cleaned_data.get("gstr1_code") or ""
            gstr1_code = str(gstr1_code).strip()

            gstr1_code_key = gstr1_code.replace(" ", "").upper()

            add_to_gstr1 = gstr1_code_key == "ARVIS"

            invoice = DealerInvoice.objects.create(
                order=order,
                financial_year=fy,
                invoice_number=invoice_number,
                invoice_date=release_form.cleaned_data["invoice_date"],
                place_of_supply=release_form.cleaned_data["place_of_supply"],
                subtotal_amount=order.subtotal_amount,
                gst_amount=order.gst_amount,
                total_amount=order.total_amount,
                pending_amount=order.total_amount,
                released_by=request.user,

                # GSTR-1 logic
                gstr1_code=gstr1_code,
                add_to_gstr1=add_to_gstr1,
                gstr1_added_at=timezone.now() if add_to_gstr1 else None,
            )

            order.status = "DISPATCH_PENDING"
            order.invoice_released_at = timezone.now()
            order.save()

            if add_to_gstr1:
                gstr_note = "Added to GSTR-1 report."
            else:
                gstr_note = "Not added to GSTR-1 report."

            log_order_action(
                order,
                request.user,
                "INVOICE_RELEASED",
                f"Invoice {invoice.invoice_number} released. {gstr_note}"
            )

            messages.success(request, f"Invoice released: {invoice.invoice_number}")
            return redirect("dealer_invoice_print", invoice_id=invoice.id, copy_type="COMPANY")

    else:
        formset = AccountantOrderItemFormSet(instance=order)
        release_form = AccountantInvoiceReleaseForm()

    return render(
        request,
        "core/accountant_invoice_review.html",
        {
            "order": order,
            "formset": formset,
            "release_form": release_form,
        }
    )



@login_required
@transaction.atomic
def warehouse_dispatch_order(request, order_id):
    order = get_object_or_404(DealerOrder, id=order_id)
    role = user_role(request.user)

    if role not in ["WAREHOUSE_MANAGER", "ADMIN"]:
        messages.error(request, "Only warehouse manager can dispatch order.")
        return redirect("dealer_order_detail", order_id=order.id)

    if order.status != "DISPATCH_PENDING":
        messages.error(request, "Order is not ready for dispatch.")
        return redirect("dealer_order_detail", order_id=order.id)

    dispatch = getattr(order, "dispatch", None)

    if request.method == "POST":
        form = DealerDispatchForm(request.POST, instance=dispatch)

        if form.is_valid():
            dispatch = form.save(commit=False)
            dispatch.order = order
            dispatch.created_by = request.user
            dispatch.save()

            order.status = "DISPATCHED"
            order.dispatched_at = timezone.now()
            order.save()

            log_order_action(order, request.user, "DISPATCHED", "Dispatch details entered.")

            messages.success(request, "Order dispatched successfully.")
            return redirect("dealer_order_detail", order_id=order.id)

    else:
        form = DealerDispatchForm(instance=dispatch)

    return render(
        request,
        "core/warehouse_dispatch_form.html",
        {
            "order": order,
            "form": form,
        }
    )



@login_required
def warehouse_my_products(request):
    profile = getattr(request.user, "profile", None)

    if not request.user.is_superuser:
        if not profile or profile.role not in ["WAREHOUSE_MANAGER", "ADMIN"]:
            messages.error(request, "Only Warehouse Manager can view warehouse products.")
            return redirect("dashboard")

    if request.user.is_superuser or profile.role == "ADMIN":
        warehouses = Warehouse.objects.filter(is_active=True)
    else:
        warehouses = Warehouse.objects.filter(
            warehouse_manager=request.user,
            is_active=True
        )

    pack_sizes = ProductPackSize.objects.filter(
        warehouse__in=warehouses,
        product__is_active=True
    ).select_related(
        "product",
        "product__category",
        "warehouse"
    ).order_by(
        "warehouse__name",
        "product__name",
        "pack_size"
    )

    products = Product.objects.filter(
        pack_sizes__warehouse__in=warehouses,
        is_active=True
    ).select_related(
        "category"
    ).distinct().order_by("name")

    total_stock_boxes = pack_sizes.aggregate(
        total=Sum("stock_boxes")
    )["total"] or 0

    context = {
        "warehouses": warehouses,
        "products": products,
        "pack_sizes": pack_sizes,
        "total_products": products.count(),
        "total_pack_sizes": pack_sizes.count(),
        "total_stock_boxes": total_stock_boxes,
    }

    return render(request, "core/warehouse_my_products.html", context)

@login_required
def warehouse_my_orders(request):
    profile = getattr(request.user, "profile", None)

    if not request.user.is_superuser:
        if not profile or profile.role not in ["WAREHOUSE_MANAGER", "PRODUCTION_INCHARGE", "ADMIN"]:
            messages.error(request, "Only Warehouse Manager can view warehouse orders.")
            return redirect("dashboard")

    status_filter = request.GET.get("status", "all")

    if request.user.is_superuser or profile.role in ["ADMIN", "PRODUCTION_INCHARGE"]:
        orders = DealerOrder.objects.all()
    else:
        orders = DealerOrder.objects.filter(
            warehouse_manager=request.user
        )

    if status_filter != "all":
        orders = orders.filter(status=status_filter)

    orders = orders.select_related(
        "dealer",
        "dealer_user",
        "concerned_sales_officer",
        "concerned_asm",
        "concerned_rsm",
        "warehouse_manager"
    ).prefetch_related(
        "items"
    ).order_by("-placed_at")

    context = {
        "orders": orders,
        "status_filter": status_filter,
        "total_orders": orders.count(),
        "warehouse_review_count": orders.filter(status="WAREHOUSE_REVIEW").count(),
        "dispatch_pending_count": orders.filter(status="DISPATCH_PENDING").count(),
        "dispatched_count": orders.filter(status="DISPATCHED").count(),
        "delivered_count": orders.filter(status="DELIVERED").count(),
    }

    return render(request, "core/warehouse_my_orders.html", context)


@login_required
def dealer_invoice_list(request):
    role = user_role(request.user)

    invoices = DealerInvoice.objects.select_related(
        "order",
        "order__dealer",
        "order__dealer_user",
        "order__concerned_sales_officer",
        "order__warehouse_manager",
    ).order_by("-invoice_date", "-id")

    if role == "DEALER":
        invoices = invoices.filter(order__dealer_user=request.user)

    elif role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        invoices = invoices.filter(order__concerned_sales_officer=request.user)

    elif role == "WAREHOUSE_MANAGER":
        invoices = invoices.filter(order__warehouse_manager=request.user)

    elif role in ["ACCOUNTANT", "ADMIN"]:
        pass

    else:
        invoices = DealerInvoice.objects.none()

    return render(
        request,
        "core/dealer_invoice_list.html",
        {
            "invoices": invoices,
            "role": role,
        }
    )


def number_to_words_indian(number):
    number = int(round(float(number)))

    if number == 0:
        return "Zero"

    ones = [
        "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
        "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
        "Sixteen", "Seventeen", "Eighteen", "Nineteen"
    ]

    tens = [
        "", "", "Twenty", "Thirty", "Forty", "Fifty",
        "Sixty", "Seventy", "Eighty", "Ninety"
    ]

    def two_digit_words(n):
        if n < 20:
            return ones[n]
        return tens[n // 10] + ("" if n % 10 == 0 else " " + ones[n % 10])

    def three_digit_words(n):
        word = ""
        if n >= 100:
            word += ones[n // 100] + " Hundred"
            n = n % 100
            if n:
                word += " "
        if n:
            word += two_digit_words(n)
        return word

    parts = []

    crore = number // 10000000
    number %= 10000000

    lakh = number // 100000
    number %= 100000

    thousand = number // 1000
    number %= 1000

    hundred = number

    if crore:
        parts.append(three_digit_words(crore) + " Crore")

    if lakh:
        parts.append(three_digit_words(lakh) + " Lakh")

    if thousand:
        parts.append(three_digit_words(thousand) + " Thousand")

    if hundred:
        parts.append(three_digit_words(hundred))

    return " ".join(parts)


def generate_upi_qr_data_uri(invoice):
    upi_id = "mab.037347040880032@axisbank"
    payee_name = "Arvis Fertilizers And Chemicals Pvt Ltd"
    display_amount = invoice.balance_amount or invoice.final_payable_amount or invoice.total_amount
    amount = f"{display_amount:.2f}"

    upi_url = (
        f"upi://pay?"
        f"pa={quote(upi_id)}"
        f"&pn={quote(payee_name)}"
        f"&am={quote(amount)}"
        f"&cu=INR"
        f"&tn={quote(invoice.invoice_number)}"
    )

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=4,
        border=2,
    )

    qr.add_data(upi_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    buffer = BytesIO()
    img.save(buffer, format="PNG")

    encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")

    return f"data:image/png;base64,{encoded}"



@login_required
def dealer_invoice_print(request, invoice_id, copy_type):
    invoice = get_object_or_404(
        DealerInvoice.objects.select_related(
            "order",
            "order__dealer",
            "order__dealer_user",
            "order__concerned_sales_officer",
            "order__concerned_asm",
            "order__concerned_rsm",
            "order__warehouse_manager",
        ).prefetch_related(
            "order__items"
        ),
        id=invoice_id
    )

    copy_type = copy_type.upper()

    if copy_type not in ["COMPANY", "DEALER", "TRANSPORTER"]:
        copy_type = "DEALER"

    order = invoice.order
    role = user_role(request.user)

    allowed = False

    if role == "ADMIN" or request.user.is_superuser:
        allowed = True

    elif role == "DEALER":
        allowed = order.dealer_user == request.user and copy_type == "DEALER"

    elif role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        allowed = order.concerned_sales_officer == request.user and copy_type == "DEALER"

    elif role == "ASM":
        allowed = order.concerned_asm == request.user and copy_type == "DEALER"

    elif role == "REGIONAL_MANAGER":
        allowed = order.concerned_rsm == request.user and copy_type == "DEALER"

    elif role == "WAREHOUSE_MANAGER":
        allowed = order.warehouse_manager == request.user and copy_type in ["COMPANY", "TRANSPORTER"]

    elif role == "ACCOUNTANT":
        allowed = copy_type in ["COMPANY", "DEALER", "TRANSPORTER"]

    if not allowed:
        messages.error(request, "You are not allowed to access this invoice copy.")
        return redirect("dealer_invoice_list")

    items = order.items.all()

    tax_summary = {}

    for item in items:
        key = item.hsn_code or "-"

        if key not in tax_summary:
            tax_summary[key] = {
                "taxable": Decimal("0.00"),
                "gst_rate": item.gst_percent,
                "gst": Decimal("0.00"),
                "total_tax": Decimal("0.00"),
            }

        tax_summary[key]["taxable"] += item.taxable_amount
        tax_summary[key]["gst"] += item.gst_amount
        tax_summary[key]["total_tax"] += item.gst_amount

    display_invoice_amount = invoice.final_payable_amount or invoice.total_amount

    if display_invoice_amount <= 0:
        display_invoice_amount = invoice.total_amount

    return render(
        request,
        "core/dealer_invoice_print.html",
        {
            "invoice": invoice,
            "order": order,
            "dealer": order.dealer,
            "items": items,
            "tax_summary": tax_summary,
            "copy_type": copy_type,
            "invoice_display_amount": display_invoice_amount,
            "invoice_amount_words": number_to_words_indian(display_invoice_amount),
            "upi_qr_data_uri": generate_upi_qr_data_uri(invoice),
            "upi_id": "mab.037347040880032@axisbank",
        }
    )



def can_accountant_or_admin(user):
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in ["ADMIN", "ACCOUNTANT"]


def get_invoice_snapshot(invoice):
    return {
        "invoice_number": invoice.invoice_number,
        "total_amount": str(invoice.total_amount),
        "paid_amount": str(invoice.paid_amount),
        "pending_amount": str(invoice.pending_amount),
        "discount_amount": str(getattr(invoice, "discount_amount", 0)),
        "final_payable_amount": str(getattr(invoice, "final_payable_amount", 0)),
        "is_converted_to_sales": getattr(invoice, "is_converted_to_sales", False),
    }



def get_category_discount_for_invoice(invoice, payment_date):
    """
    Finds best category payment rule discount from products in this invoice.

    Supports common field names:
    - payment_type / rule_type / payment_rule_type / payment_option / type
    - discount_percent / discount_percentage / discount / discount_value
    - from_day / start_day / min_days / days_from / day_from
    - to_day / end_day / max_days / days_to / day_to
    """

    from decimal import Decimal

    order = invoice.order
    ordered_date = order.placed_at.date()
    days = (payment_date - ordered_date).days

    best_discount = Decimal("0.00")
    best_note = "No category rule matched"

    def get_first_value(obj, names, default=None):
        for name in names:
            if hasattr(obj, name):
                value = getattr(obj, name)

                if value is not None:
                    return value

        return default

    def normalize(value):
        if value is None:
            return ""

        return str(value).upper().replace(" ", "_").replace("-", "_")

    def get_discount(rule):
        value = get_first_value(
            rule,
            [
                "discount_percent",
                "discount_percentage",
                "discount",
                "discount_value",
                "percentage",
            ],
            Decimal("0.00")
        )

        try:
            return Decimal(str(value))
        except Exception:
            return Decimal("0.00")

    def get_day_value(rule, names, default=0):
        value = get_first_value(rule, names, default)

        try:
            return int(value)
        except Exception:
            return default

    def is_rule_active(rule):
        if hasattr(rule, "is_active"):
            return bool(rule.is_active)

        if hasattr(rule, "active"):
            return bool(rule.active)

        return True

    def get_category_rules(category):
        possible_related_names = [
            "payment_rules",
            "category_payment_rules",
            "paymentrule_set",
            "categorypaymentrule_set",
            "productpaymentrule_set",
        ]

        for related_name in possible_related_names:
            manager = getattr(category, related_name, None)

            if manager is not None and hasattr(manager, "all"):
                return manager.all()

        return []

    for item in order.items.select_related(
        "product_pack",
        "product_pack__product",
        "product_pack__product__category"
    ):
        category = item.product_pack.product.category
        rules = get_category_rules(category)

        for rule in rules:
            if not is_rule_active(rule):
                continue

            rule_discount = get_discount(rule)

            if rule_discount <= 0:
                continue

            payment_type = normalize(
                get_first_value(
                    rule,
                    [
                        "payment_type",
                        "rule_type",
                        "payment_rule_type",
                        "payment_option",
                        "type",
                        "name",
                    ],
                    ""
                )
            )

            from_day = get_day_value(
                rule,
                ["from_day", "start_day", "min_days", "days_from", "day_from"],
                0
            )

            to_day = get_day_value(
                rule,
                ["to_day", "end_day", "max_days", "days_to", "day_to"],
                0
            )

            matched = False
            note = ""

            if "INSTANT" in payment_type:
                if days == 0:
                    matched = True
                    note = f"{category.name} instant payment discount {rule_discount}%"

            elif "ADVANCE" in payment_type:
                if payment_date <= ordered_date:
                    matched = True
                    note = f"{category.name} advance payment discount {rule_discount}%"

            elif "DAY" in payment_type or from_day or to_day:
                if from_day <= days <= to_day:
                    matched = True
                    note = f"{category.name} day-wise discount {rule_discount}% for {days} days"

            else:
                # fallback: if rule has no recognizable type, still apply it
                matched = True
                note = f"{category.name} category payment discount {rule_discount}%"

            if matched and rule_discount > best_discount:
                best_discount = rule_discount
                best_note = note

    return best_discount, best_note



def update_dealer_and_employee_targets(invoice, amount):
    """
    Updates dealer yearly achieved amount and concerned sales officer yearly target.
    This is defensive: it updates only fields that exist in your models.
    """

    order = invoice.order
    dealer = order.dealer
    year = invoice.invoice_date.year

    if hasattr(dealer, "achieved_amount"):
        dealer.achieved_amount = (dealer.achieved_amount or Decimal("0.00")) + amount
        dealer.save(update_fields=["achieved_amount"])

    elif hasattr(dealer, "target_achieved_amount"):
        dealer.target_achieved_amount = (dealer.target_achieved_amount or Decimal("0.00")) + amount
        dealer.save(update_fields=["target_achieved_amount"])

    sales_user = order.concerned_sales_officer

    if sales_user:
        target, created = EmployeeYearlyTarget.objects.get_or_create(
            employee=sales_user,
            year=year,
            defaults={
                "target_amount": Decimal("0.00"),
                "achieved_amount": Decimal("0.00"),
            }
        )

        target.achieved_amount = (target.achieved_amount or Decimal("0.00")) + amount
        target.save(update_fields=["achieved_amount"])


@login_required
def invoice_payment_preview(request, invoice_id):
    invoice = get_object_or_404(
        DealerInvoice.objects.select_related(
            "order",
            "order__dealer",
            "order__concerned_sales_officer",
        ).prefetch_related("order__items"),
        id=invoice_id
    )

    if not can_accountant_or_admin(request.user):
        return JsonResponse({
            "success": False,
            "message": "Not allowed."
        }, status=403)

    try:
        payment_status = request.GET.get("payment_status", "FULL")
        payment_date_value = request.GET.get("payment_date")
        payment_amount = Decimal(request.GET.get("payment_amount") or "0")
        manual_discount_percent = Decimal(request.GET.get("manual_discount_percent") or "0")
        selected_rule_key = request.GET.get("category_rule_choice") or ""

        if payment_date_value:
            payment_date = timezone.datetime.strptime(
                payment_date_value,
                "%Y-%m-%d"
            ).date()
        else:
            payment_date = timezone.localdate()

    except (InvalidOperation, ValueError):
        return JsonResponse({
            "success": False,
            "message": "Invalid preview values."
        })

    if payment_amount <= 0:
        payment_amount = invoice.balance_amount

    discount_percent = Decimal("0.00")
    discount_note = "No discount applied"
    selected_rule_name = ""
    order_date_display = invoice.order.placed_at.date().strftime("%d-%m-%Y")
    due_date_display = "-"
    rule_eligible = False

    if payment_status == "FULL" and selected_rule_key:
        rule_result = get_selected_invoice_category_rule(
            invoice,
            selected_rule_key,
            payment_date
        )

        discount_percent = rule_result["discount_percent"]
        discount_note = rule_result["discount_note"]
        selected_rule_name = rule_result["rule_name"]
        rule_eligible = rule_result["eligible"]

        if rule_result["order_date"]:
            order_date_display = rule_result["order_date"].strftime("%d-%m-%Y")

        if rule_result["due_date"]:
            due_date_display = rule_result["due_date"].strftime("%d-%m-%Y")

    elif payment_status == "PARTIAL":
        discount_percent = manual_discount_percent

        if discount_percent > 0:
            discount_note = f"Manual discount {discount_percent}%"
        else:
            discount_note = "No manual discount applied"

    else:
        discount_note = "Select a category payment rule to apply discount"

    discount_amount = (payment_amount * discount_percent) / Decimal("100.00")
    final_amount_after_discount = payment_amount - discount_amount

    current_paid = invoice.paid_amount or Decimal("0.00")
    current_discount = invoice.discount_amount or Decimal("0.00")

    total_discount_after_this = current_discount + discount_amount
    final_invoice_payable = invoice.total_amount - total_discount_after_this

    paid_after_this = current_paid + final_amount_after_discount
    balance_after_payment = final_invoice_payable - paid_after_this

    if balance_after_payment < 0:
        balance_after_payment = Decimal("0.00")

    return JsonResponse({
        "success": True,
        "invoice_total": f"{invoice.total_amount:.2f}",
        "current_paid": f"{current_paid:.2f}",
        "current_discount": f"{current_discount:.2f}",
        "payment_amount": f"{payment_amount:.2f}",
        "discount_percent": f"{discount_percent:.2f}",
        "discount_amount": f"{discount_amount:.2f}",
        "final_amount_after_discount": f"{final_amount_after_discount:.2f}",
        "final_invoice_payable": f"{final_invoice_payable:.2f}",
        "balance_after_payment": f"{balance_after_payment:.2f}",
        "discount_note": discount_note,
        "selected_rule_name": selected_rule_name,
        "order_date": order_date_display,
        "due_date": due_date_display,
        "rule_eligible": rule_eligible,
    })



def get_first_rule_value(obj, names, default=None):
    for name in names:
        if hasattr(obj, name):
            value = getattr(obj, name)

            if value is not None:
                return value

    return default


def normalize_rule_text(value):
    if value is None:
        return ""

    return str(value).upper().replace(" ", "_").replace("-", "_")


def get_rule_discount(rule):
    value = get_first_rule_value(
        rule,
        [
            "discount_percent",
            "discount_percentage",
            "discount",
            "discount_value",
            "percentage",
        ],
        Decimal("0.00")
    )

    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0.00")


def get_rule_day_value(rule, names, default=0):
    value = get_first_rule_value(rule, names, default)

    try:
        return int(value)
    except Exception:
        return default


def get_rule_payment_type_display(rule):
    field_names = [
        "payment_type",
        "rule_type",
        "payment_rule_type",
        "payment_option",
        "type",
        "name",
    ]

    for field_name in field_names:
        if hasattr(rule, field_name):
            raw_value = getattr(rule, field_name)

            display_method = getattr(rule, f"get_{field_name}_display", None)

            if display_method:
                try:
                    return str(display_method()), raw_value
                except Exception:
                    pass

            return str(raw_value), raw_value

    return "Category Payment Rule", ""


def is_category_rule_active(rule):
    if hasattr(rule, "is_active"):
        return bool(rule.is_active)

    if hasattr(rule, "active"):
        return bool(rule.active)

    return True


def get_rules_from_category(category):
    possible_related_names = [
        "payment_rules",
        "category_payment_rules",
        "paymentrule_set",
        "categorypaymentrule_set",
        "productpaymentrule_set",
    ]

    for related_name in possible_related_names:
        manager = getattr(category, related_name, None)

        if manager is not None and hasattr(manager, "all"):
            return manager.all()

    return []


def get_rule_dates(rule, ordered_date):
    display_name, raw_type = get_rule_payment_type_display(rule)
    normalized_type = normalize_rule_text(raw_type or display_name)

    from_day = get_rule_day_value(
        rule,
        ["from_day", "start_day", "min_days", "days_from", "day_from"],
        0
    )

    to_day = get_rule_day_value(
        rule,
        ["to_day", "end_day", "max_days", "days_to", "day_to"],
        0
    )

    if "INSTANT" in normalized_type:
        start_date = ordered_date
        due_date = ordered_date

    elif "ADVANCE" in normalized_type:
        start_date = ordered_date
        due_date = ordered_date

    elif "DAY" in normalized_type or from_day or to_day:
        start_date = ordered_date + timedelta(days=from_day)
        due_date = ordered_date + timedelta(days=to_day)

    else:
        start_date = ordered_date
        due_date = ordered_date + timedelta(days=to_day) if to_day else ordered_date

    return {
        "display_name": display_name,
        "raw_type": raw_type,
        "normalized_type": normalized_type,
        "from_day": from_day,
        "to_day": to_day,
        "start_date": start_date,
        "due_date": due_date,
    }


def get_invoice_category_rule_options(invoice):
    """
    Shows all category payment rules available for products in this invoice.
    Accountant can manually select one rule.
    """

    ordered_date = invoice.order.placed_at.date()

    choices = [
        ("", "No category payment rule / No discount")
    ]

    seen = set()

    for item in invoice.order.items.select_related(
        "product_pack",
        "product_pack__product",
        "product_pack__product__category"
    ):
        category = item.product_pack.product.category
        rules = get_rules_from_category(category)

        for rule in rules:
            if not is_category_rule_active(rule):
                continue

            discount = get_rule_discount(rule)

            if discount <= 0:
                continue

            rule_key = f"{category.id}:{rule.id}"

            if rule_key in seen:
                continue

            seen.add(rule_key)

            rule_dates = get_rule_dates(rule, ordered_date)

            label = (
                f"{category.name} | "
                f"{rule_dates['display_name']} | "
                f"{discount}% | "
                f"Order: {ordered_date.strftime('%d-%m-%Y')} | "
                f"Due: {rule_dates['due_date'].strftime('%d-%m-%Y')}"
            )

            choices.append((rule_key, label))

    return choices


def get_selected_invoice_category_rule(invoice, selected_rule_key, payment_date):
    """
    Returns selected rule discount details and checks eligibility based on payment date.
    """

    if not selected_rule_key:
        return {
            "matched": False,
            "eligible": False,
            "discount_percent": Decimal("0.00"),
            "discount_note": "No category rule selected",
            "rule_name": "",
            "order_date": invoice.order.placed_at.date(),
            "due_date": None,
        }

    ordered_date = invoice.order.placed_at.date()
    days = (payment_date - ordered_date).days

    for item in invoice.order.items.select_related(
        "product_pack",
        "product_pack__product",
        "product_pack__product__category"
    ):
        category = item.product_pack.product.category
        rules = get_rules_from_category(category)

        for rule in rules:
            rule_key = f"{category.id}:{rule.id}"

            if rule_key != selected_rule_key:
                continue

            if not is_category_rule_active(rule):
                break

            discount = get_rule_discount(rule)
            rule_dates = get_rule_dates(rule, ordered_date)
            normalized_type = rule_dates["normalized_type"]

            eligible = False

            if "INSTANT" in normalized_type:
                eligible = days == 0

            elif "ADVANCE" in normalized_type:
                eligible = payment_date <= ordered_date

            elif "DAY" in normalized_type or rule_dates["from_day"] or rule_dates["to_day"]:
                eligible = rule_dates["from_day"] <= days <= rule_dates["to_day"]

            else:
                eligible = payment_date <= rule_dates["due_date"]

            if eligible:
                discount_note = (
                    f"{category.name} - {rule_dates['display_name']} selected. "
                    f"Order date: {ordered_date.strftime('%d-%m-%Y')}, "
                    f"Due date: {rule_dates['due_date'].strftime('%d-%m-%Y')}"
                )
                final_discount = discount
            else:
                discount_note = (
                    f"{category.name} - {rule_dates['display_name']} selected, "
                    f"but payment date is not eligible. "
                    f"Order date: {ordered_date.strftime('%d-%m-%Y')}, "
                    f"Due date: {rule_dates['due_date'].strftime('%d-%m-%Y')}"
                )
                final_discount = Decimal("0.00")

            return {
                "matched": True,
                "eligible": eligible,
                "discount_percent": final_discount,
                "actual_rule_discount": discount,
                "discount_note": discount_note,
                "rule_name": f"{category.name} - {rule_dates['display_name']}",
                "order_date": ordered_date,
                "due_date": rule_dates["due_date"],
            }

    return {
        "matched": False,
        "eligible": False,
        "discount_percent": Decimal("0.00"),
        "discount_note": "Selected category rule not found",
        "rule_name": "",
        "order_date": ordered_date,
        "due_date": None,
    }


@login_required
@transaction.atomic
def convert_invoice_to_sales(request, invoice_id):
    invoice = get_object_or_404(
        DealerInvoice.objects.select_related(
            "order",
            "order__dealer",
            "order__concerned_sales_officer",
        ).prefetch_related(
            "order__items",
            "payments",
            "edit_history",
        ),
        id=invoice_id
    )

    if not can_accountant_or_admin(request.user):
        messages.error(request, "Only Accountant or Admin can convert invoice to sales.")
        return redirect("dealer_invoice_list")

    if request.method == "POST":
        category_rule_choices = get_invoice_category_rule_options(invoice)

        form = DealerInvoicePaymentForm(
            request.POST,
            request.FILES,
            invoice=invoice,
            category_rule_choices=category_rule_choices
        )

        if form.is_valid():
            old_values = get_invoice_snapshot(invoice)

            payment = form.save(commit=False)
            payment.invoice = invoice
            payment.created_by = request.user
            payment.gross_invoice_amount = invoice.total_amount

            payment_status = form.cleaned_data["payment_status"]
            payment_date = form.cleaned_data["payment_date"]
            payment_amount = form.cleaned_data["payment_amount"]
            manual_discount_percent = form.cleaned_data.get("manual_discount_percent") or Decimal("0.00")
            selected_rule_key = form.cleaned_data.get("category_rule_choice") or ""

            discount_percent = Decimal("0.00")
            discount_type = "NONE"
            category_rule_note = ""
            selected_rule_name = ""
            selected_rule_order_date = None
            selected_rule_due_date = None

            if payment_status == "FULL" and selected_rule_key:
                rule_result = get_selected_invoice_category_rule(
                    invoice,
                    selected_rule_key,
                    payment_date
                )

                discount_percent = rule_result["discount_percent"]
                category_rule_note = rule_result["discount_note"]
                selected_rule_name = rule_result["rule_name"]
                selected_rule_order_date = rule_result["order_date"]
                selected_rule_due_date = rule_result["due_date"]

                if discount_percent > 0:
                    discount_type = "CATEGORY_RULE"

            elif payment_status == "PARTIAL" and manual_discount_percent > 0:
                discount_percent = manual_discount_percent
                discount_type = "MANUAL"
                category_rule_note = f"Manual discount {manual_discount_percent}%"
            else:
                category_rule_note = "No discount applied"

            discount_amount = (payment_amount * discount_percent) / Decimal("100.00")
            net_received_amount = payment_amount - discount_amount

            payment.payment_amount = payment_amount
            payment.discount_type = discount_type
            payment.discount_percent = discount_percent
            payment.discount_amount = discount_amount
            payment.net_received_amount = net_received_amount
            payment.category_rule_note = category_rule_note
            payment.selected_category_rule_key = selected_rule_key
            payment.selected_category_rule_name = selected_rule_name
            payment.selected_category_rule_order_date = selected_rule_order_date
            payment.selected_category_rule_due_date = selected_rule_due_date

            new_paid_amount = (invoice.paid_amount or Decimal("0.00")) + net_received_amount
            total_discount = (getattr(invoice, "discount_amount", Decimal("0.00")) or Decimal("0.00")) + discount_amount

            final_payable_amount = invoice.total_amount - total_discount
            balance_after_payment = final_payable_amount - new_paid_amount

            if balance_after_payment < 0:
                balance_after_payment = Decimal("0.00")

            payment.balance_after_payment = balance_after_payment
            payment.save()

            invoice.paid_amount = new_paid_amount
            invoice.discount_amount = total_discount
            invoice.final_payable_amount = final_payable_amount
            invoice.pending_amount = balance_after_payment
            invoice.last_payment_date = payment_date
            invoice.is_converted_to_sales = True
            invoice.save()

            if not invoice.sales_converted_at:
                invoice.sales_converted_at = timezone.now()
                invoice.sales_converted_by = request.user

            invoice.save()

            update_dealer_and_employee_targets(invoice, net_received_amount)

            new_values = get_invoice_snapshot(invoice)

            create_invoice_history(
                invoice=invoice,
                user=request.user,
                action="Sales conversion / payment added",
                old_values=old_values,
                new_values=new_values,
                note=(
                    f"Payment: ₹{payment_amount}, "
                    f"Discount: ₹{discount_amount}, "
                    f"Net Received: ₹{net_received_amount}, "
                    f"Mode: {payment.payment_mode}"
                )
            )

            messages.success(request, "Invoice converted to sales and payment saved successfully.")
            return redirect("convert_invoice_to_sales", invoice_id=invoice.id)

    else:
        category_rule_choices = get_invoice_category_rule_options(invoice)

        form = DealerInvoicePaymentForm(
            invoice=invoice,
            category_rule_choices=category_rule_choices
        )

    return render(
        request,
        "core/convert_invoice_to_sales.html",
        {
            "invoice": invoice,
            "order": invoice.order,
            "dealer": invoice.order.dealer,
            "form": form,
            "payments": invoice.payments.all(),
            "history": invoice.edit_history.all(),
        }
    )


def make_history_value_safe(value):
    """
    Converts Decimal/date/datetime and nested values into safe values
    for JSONField or TextField history storage.
    """
    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, (date, datetime)):
        return value.isoformat()

    if isinstance(value, dict):
        return {
            key: make_history_value_safe(val)
            for key, val in value.items()
        }

    if isinstance(value, list):
        return [make_history_value_safe(item) for item in value]

    return value


def clean_history_dict(value):
    """
    Handles both JSONField dict and old TextField string dict like:
    {'paid_amount': '0.00', 'is_converted_to_sales': False}
    """
    if not value:
        return {}

    if isinstance(value, dict):
        return value

    if isinstance(value, str):
        # Try JSON first
        try:
            parsed = json.loads(value)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            pass

        # Try Python dict string
        try:
            parsed = ast.literal_eval(value)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            pass

    return {}


def get_history_label(key):
    labels = {
        "invoice_number": "Invoice Number",
        "total_amount": "Invoice Amount",
        "paid_amount": "Paid Amount",
        "pending_amount": "Pending Amount",
        "discount_amount": "Discount Amount",
        "final_payable_amount": "Final Payable Amount",
        "balance_amount": "Balance Amount",
        "is_converted_to_sales": "Converted to Sales",
        "payment_mode": "Payment Mode",
        "payment_reference": "Payment Reference",
        "reference_number": "Reference Number",
        "transaction_number": "Transaction Number",
        "cheque_number": "Cheque Number",
        "last_payment_date": "Last Payment Date",
        "return_credit_amount": "Return Credit Amount",
        "net_payable_after_return": "Net Payable After Return",
        "discount_rule_name": "Discount Rule",
        "discount_percentage": "Discount Percentage",
        "payment_type": "Payment Type",
    }

    return labels.get(key, key.replace("_", " ").title())


def format_history_value(key, value):
    if value is None or value == "":
        return "-"

    money_fields = [
        "total_amount",
        "paid_amount",
        "pending_amount",
        "discount_amount",
        "final_payable_amount",
        "balance_amount",
        "return_credit_amount",
        "net_payable_after_return",
    ]

    percentage_fields = [
        "discount_percentage",
    ]

    if key in money_fields:
        try:
            amount = Decimal(str(value or "0"))
            return f"₹{amount:,.2f}"
        except Exception:
            return value

    if key in percentage_fields:
        try:
            percent = Decimal(str(value or "0"))
            return f"{percent}%"
        except Exception:
            return value

    if key == "is_converted_to_sales":
        return "Yes" if value is True or str(value).lower() == "true" else "No"

    return str(value)


def create_invoice_history(invoice, user, action, old_values, new_values, note=""):
    DealerInvoiceEditHistory.objects.create(
        invoice=invoice,
        edited_by=user,
        action=action,
        old_values=make_history_value_safe(old_values),
        new_values=make_history_value_safe(new_values),
        note=note,
    )


@login_required
def invoice_edit_history(request, invoice_id):
    invoice = get_object_or_404(
        DealerInvoice.objects.select_related("order", "order__dealer"),
        id=invoice_id
    )

    if not can_accountant_or_admin(request.user):
        messages.error(request, "Only Accountant or Admin can view invoice history.")
        return redirect("dealer_invoice_list")

    history_qs = invoice.edit_history.select_related("edited_by").order_by("-created_at")

    formatted_history = []

    for h in history_qs:
        old_values = clean_history_dict(h.old_values)
        new_values = clean_history_dict(h.new_values)

        all_keys = list(
            dict.fromkeys(
                list(old_values.keys()) + list(new_values.keys())
            )
        )

        formatted_changes = []

        for key in all_keys:
            old_value = old_values.get(key)
            new_value = new_values.get(key)

            formatted_changes.append({
                "label": get_history_label(key),
                "old": format_history_value(key, old_value),
                "new": format_history_value(key, new_value),
                "changed": old_value != new_value,
            })

        h.formatted_changes = formatted_changes
        h.changed_count = sum(1 for item in formatted_changes if item["changed"])

        formatted_history.append(h)

    return render(
        request,
        "core/invoice_edit_history.html",
        {
            "invoice": invoice,
            "history": formatted_history,
        }
    )


@login_required
def paid_sales_report(request):
    role = user_role(request.user)

    if not request.user.is_superuser and role not in ["ADMIN", "ACCOUNTANT"]:
        messages.error(request, "Only Admin or Accountant can view paid sales report.")
        return redirect("dashboard")

    today = timezone.localdate()
    period = request.GET.get("period", "this_month")

    def month_range(any_date):
        start = any_date.replace(day=1)

        if start.month == 12:
            end = date(start.year + 1, 1, 1) - timedelta(days=1)
        else:
            end = date(start.year, start.month + 1, 1) - timedelta(days=1)

        return start, end

    def current_financial_year_start(any_date):
        if any_date.month >= 4:
            return date(any_date.year, 4, 1)

        return date(any_date.year - 1, 4, 1)

    fy_start = current_financial_year_start(today)

    if period == "this_month":
        start_date, end_date = month_range(today)
        title = "This Month"

    elif period == "last_month":
        first_day_this_month = today.replace(day=1)
        last_month_day = first_day_this_month - timedelta(days=1)
        start_date, end_date = month_range(last_month_day)
        title = "Last Month"

    elif period == "this_year":
        start_date = fy_start
        end_date = date(fy_start.year + 1, 3, 31)
        title = f"Financial Year {fy_start.year}-{str(fy_start.year + 1)[-2:]}"

    elif period == "last_year":
        start_date = date(fy_start.year - 1, 4, 1)
        end_date = date(fy_start.year, 3, 31)
        title = f"Last Financial Year {start_date.year}-{str(end_date.year)[-2:]}"

    elif period == "q1":
        start_date = date(fy_start.year, 4, 1)
        end_date = date(fy_start.year, 6, 30)
        title = "Q1 - April to June"

    elif period == "q2":
        start_date = date(fy_start.year, 7, 1)
        end_date = date(fy_start.year, 9, 30)
        title = "Q2 - July to September"

    elif period == "q3":
        start_date = date(fy_start.year, 10, 1)
        end_date = date(fy_start.year, 12, 31)
        title = "Q3 - October to December"

    elif period == "q4":
        start_date = date(fy_start.year + 1, 1, 1)
        end_date = date(fy_start.year + 1, 3, 31)
        title = "Q4 - January to March"

    elif period == "custom":
        start_value = request.GET.get("start_date")
        end_value = request.GET.get("end_date")

        try:
            start_date = date.fromisoformat(start_value)
            end_date = date.fromisoformat(end_value)
        except Exception:
            start_date, end_date = month_range(today)

        title = "Custom Range"

    else:
        start_date, end_date = month_range(today)
        title = "This Month"

    payments = DealerInvoicePayment.objects.select_related(
        "invoice",
        "invoice__order",
        "invoice__order__dealer",
    ).filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
        net_received_amount__gt=0,
    )

    total_paid = payments.aggregate(
        total=Sum("net_received_amount")
    )["total"] or Decimal("0.00")

    total_discount = payments.aggregate(
        total=Sum("discount_amount")
    )["total"] or Decimal("0.00")

    total_entries = payments.count()

    parties_count = payments.values(
        "invoice__order__dealer_id"
    ).distinct().count()

    day_rows = payments.values("payment_date").annotate(
        total=Sum("net_received_amount")
    ).order_by("payment_date")

    day_map = {
        row["payment_date"]: float(row["total"] or 0)
        for row in day_rows
    }

    chart_labels = []
    chart_values = []

    day = start_date

    while day <= end_date:
        chart_labels.append(day.strftime("%d %b"))
        chart_values.append(day_map.get(day, 0))
        day += timedelta(days=1)

    dealer_rows = list(
        payments.values(
            "invoice__order__dealer__firm_name"
        ).annotate(
            total=Sum("net_received_amount")
        ).order_by("-total")[:10]
    )

    raw_mode_rows = list(
        payments.values(
            "payment_mode"
        ).annotate(
            total=Sum("net_received_amount")
        ).order_by("-total")
    )

    mode_rows = []
    mode_labels = []
    mode_values = []

    for row in raw_mode_rows:
        payment_mode = row["payment_mode"]
        amount = row["total"] or Decimal("0.00")

        if payment_mode == "UPI":
            mode_name = "UPI"
        elif payment_mode == "BANK":
            mode_name = "Bank Transfer"
        elif payment_mode == "CHEQUE":
            mode_name = "Cheque"
        else:
            mode_name = payment_mode or "Other"

        if total_paid > 0:
            percentage = (amount / total_paid) * Decimal("100.00")
        else:
            percentage = Decimal("0.00")

        mode_rows.append({
            "mode": mode_name,
            "total": amount,
            "percentage": percentage,
        })

        mode_labels.append(mode_name)
        mode_values.append(float(amount))

    recent_payments = payments.order_by(
        "-payment_date",
        "-created_at"
    )[:20]

    return render(
        request,
        "core/paid_sales_report.html",
        {
            "period": period,
            "title": title,
            "start_date": start_date,
            "end_date": end_date,
            "total_paid": total_paid,
            "total_discount": total_discount,
            "total_entries": total_entries,
            "parties_count": parties_count,
            "chart_labels": chart_labels,
            "chart_values": chart_values,
            "dealer_rows": dealer_rows,
            "mode_rows": mode_rows,
            "mode_labels": mode_labels,
            "mode_values": mode_values,
            "recent_payments": recent_payments,
            "has_paid_sales": total_paid > 0,
        }
    )



def can_manage_accounts(user):
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in ["ADMIN", "ACCOUNTANT"]


@login_required
def accounts_dashboard(request):
    if not can_manage_accounts(request.user):
        messages.error(request, "Only Admin or Accountant can access accounts.")
        return redirect("dashboard")

    today = timezone.localdate()
    period = request.GET.get("period", "this_month")

    def month_range(any_date):
        start = any_date.replace(day=1)

        if start.month == 12:
            end = date(start.year + 1, 1, 1) - timedelta(days=1)
        else:
            end = date(start.year, start.month + 1, 1) - timedelta(days=1)

        return start, end

    def financial_year_start(any_date):
        if any_date.month >= 4:
            return date(any_date.year, 4, 1)

        return date(any_date.year - 1, 4, 1)

    fy_start = financial_year_start(today)

    if period == "this_month":
        start_date, end_date = month_range(today)
        title = "This Month"

    elif period == "last_month":
        first_day_this_month = today.replace(day=1)
        last_month_day = first_day_this_month - timedelta(days=1)
        start_date, end_date = month_range(last_month_day)
        title = "Last Month"

    elif period == "this_year":
        start_date = fy_start
        end_date = date(fy_start.year + 1, 3, 31)
        title = f"Financial Year {fy_start.year}-{str(fy_start.year + 1)[-2:]}"

    elif period == "last_year":
        start_date = date(fy_start.year - 1, 4, 1)
        end_date = date(fy_start.year, 3, 31)
        title = f"Last Financial Year {start_date.year}-{str(end_date.year)[-2:]}"

    elif period == "q1":
        start_date = date(fy_start.year, 4, 1)
        end_date = date(fy_start.year, 6, 30)
        title = "Q1 - April to June"

    elif period == "q2":
        start_date = date(fy_start.year, 7, 1)
        end_date = date(fy_start.year, 9, 30)
        title = "Q2 - July to September"

    elif period == "q3":
        start_date = date(fy_start.year, 10, 1)
        end_date = date(fy_start.year, 12, 31)
        title = "Q3 - October to December"

    elif period == "q4":
        start_date = date(fy_start.year + 1, 1, 1)
        end_date = date(fy_start.year + 1, 3, 31)
        title = "Q4 - January to March"

    elif period == "custom":
        start_value = request.GET.get("start_date")
        end_value = request.GET.get("end_date")

        try:
            start_date = date.fromisoformat(start_value)
            end_date = date.fromisoformat(end_value)
        except Exception:
            start_date, end_date = month_range(today)

        title = "Custom Range"

    else:
        period = "this_month"
        start_date, end_date = month_range(today)
        title = "This Month"

    accounts = CompanyAccount.objects.order_by("account_name")

    total_balance = accounts.aggregate(
        total=Sum("current_balance")
    )["total"] or Decimal("0.00")

    payment_ins = CompanyPaymentIn.objects.select_related(
        "account",
        "dealer",
        "created_by",
    ).filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
    ).order_by("-payment_date", "-created_at")

    payment_outs = CompanyPaymentOut.objects.select_related(
        "account",
        "created_by",
    ).filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date,
    ).order_by("-payment_date", "-created_at")

    total_payment_in = payment_ins.aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    total_payment_out = payment_outs.aggregate(
        total=Sum("amount")
    )["total"] or Decimal("0.00")

    net_cash_flow = total_payment_in - total_payment_out

    return render(
        request,
        "core/accounts_dashboard.html",
        {
            "accounts": accounts,

            "period": period,
            "start_date": start_date,
            "end_date": end_date,
            "title": title,

            "total_balance": total_balance,
            "total_payment_in": total_payment_in,
            "total_payment_out": total_payment_out,
            "net_cash_flow": net_cash_flow,

            "recent_ins": payment_ins,
            "recent_outs": payment_outs,
        }
    )

    

@login_required
def company_account_create(request):
    if not can_manage_accounts(request.user):
        messages.error(request, "Only Admin or Accountant can add bank accounts.")
        return redirect("dashboard")

    if request.method == "POST":
        form = CompanyAccountForm(request.POST)

        if form.is_valid():
            account = form.save(commit=False)

            if not account.current_balance:
                account.current_balance = account.opening_balance

            account.save()

            messages.success(request, "Bank account added successfully.")
            return redirect("accounts_dashboard")
    else:
        form = CompanyAccountForm()

    return render(
        request,
        "core/company_account_form.html",
        {
            "form": form,
            "title": "Add Bank Account",
            "button_text": "Save Bank Account",
        }
    )


@login_required
@transaction.atomic
def company_payment_in_create(request):
    if not can_manage_accounts(request.user):
        messages.error(request, "Only Admin or Accountant can add payment in.")
        return redirect("dashboard")

    if request.method == "POST":
        form = CompanyPaymentInForm(request.POST, request.FILES)

        if form.is_valid():
            payment = form.save(commit=False)
            payment.created_by = request.user
            payment.full_clean()
            payment.save()

            account = CompanyAccount.objects.select_for_update().get(
                pk=payment.account_id
            )
            account.current_balance = account.current_balance + payment.amount
            account.save(update_fields=["current_balance", "updated_at"])

            messages.success(request, "Payment in saved successfully.")
            return redirect("accounts_dashboard")
    else:
        form = CompanyPaymentInForm()

    return render(
        request,
        "core/company_payment_in_form.html",
        {
            "form": form,
            "title": "Add Payment In",
            "button_text": "Save Payment In",
        }
    )


@login_required
@transaction.atomic
def company_payment_out_create(request):
    if not can_manage_accounts(request.user):
        messages.error(request, "Only Admin or Accountant can add payment out.")
        return redirect("dashboard")

    if request.method == "POST":
        form = CompanyPaymentOutForm(request.POST, request.FILES)

        if form.is_valid():
            payment = form.save(commit=False)
            payment.created_by = request.user
            payment.full_clean()
            payment.save()

            account = CompanyAccount.objects.select_for_update().get(
                pk=payment.account_id
            )
            account.current_balance = account.current_balance - payment.amount
            account.save(update_fields=["current_balance", "updated_at"])

            messages.success(request, "Payment out saved successfully.")
            return redirect("accounts_dashboard")
    else:
        form = CompanyPaymentOutForm()

    return render(
        request,
        "core/company_payment_out_form.html",
        {
            "form": form,
            "title": "Add Payment Out",
            "button_text": "Save Payment Out",
        }
    )



@login_required
def order_payment_difference_report(request):
    role = user_role(request.user)

    if not request.user.is_superuser and role not in ["ADMIN", "ACCOUNTANT"]:
        messages.error(request, "Only Admin or Accountant can view this report.")
        return redirect("dashboard")

    today = timezone.localdate()
    period = request.GET.get("period", "this_month")

    def month_range(any_date):
        start = any_date.replace(day=1)

        if start.month == 12:
            end = date(start.year + 1, 1, 1) - timedelta(days=1)
        else:
            end = date(start.year, start.month + 1, 1) - timedelta(days=1)

        return start, end

    def current_financial_year_start(any_date):
        if any_date.month >= 4:
            return date(any_date.year, 4, 1)
        return date(any_date.year - 1, 4, 1)

    fy_start = current_financial_year_start(today)

    if period == "this_month":
        start_date, end_date = month_range(today)
        title = "This Month"

    elif period == "last_month":
        first_day_this_month = today.replace(day=1)
        last_month_day = first_day_this_month - timedelta(days=1)
        start_date, end_date = month_range(last_month_day)
        title = "Last Month"

    elif period == "this_year":
        start_date = fy_start
        end_date = date(fy_start.year + 1, 3, 31)
        title = f"Financial Year {fy_start.year}-{str(fy_start.year + 1)[-2:]}"

    elif period == "last_year":
        start_date = date(fy_start.year - 1, 4, 1)
        end_date = date(fy_start.year, 3, 31)
        title = f"Last Financial Year {start_date.year}-{str(end_date.year)[-2:]}"

    elif period == "q1":
        start_date = date(fy_start.year, 4, 1)
        end_date = date(fy_start.year, 6, 30)
        title = "Q1 - April to June"

    elif period == "q2":
        start_date = date(fy_start.year, 7, 1)
        end_date = date(fy_start.year, 9, 30)
        title = "Q2 - July to September"

    elif period == "q3":
        start_date = date(fy_start.year, 10, 1)
        end_date = date(fy_start.year, 12, 31)
        title = "Q3 - October to December"

    elif period == "q4":
        start_date = date(fy_start.year + 1, 1, 1)
        end_date = date(fy_start.year + 1, 3, 31)
        title = "Q4 - January to March"

    elif period == "custom":
        start_value = request.GET.get("start_date")
        end_value = request.GET.get("end_date")

        try:
            start_date = date.fromisoformat(start_value)
            end_date = date.fromisoformat(end_value)
        except Exception:
            start_date, end_date = month_range(today)

        title = "Custom Range"

    else:
        start_date, end_date = month_range(today)
        title = "This Month"

    invoices = DealerInvoice.objects.select_related(
        "order",
        "order__dealer",
    ).filter(
        invoice_date__gte=start_date,
        invoice_date__lte=end_date,
    ).order_by("-invoice_date", "-id")

    total_order_amount = Decimal("0.00")
    total_final_payable = Decimal("0.00")
    total_received = Decimal("0.00")
    total_discount = Decimal("0.00")
    total_pending = Decimal("0.00")
    total_difference = Decimal("0.00")

    report_rows = []
    chart_labels = []
    order_values = []
    final_payable_values = []
    received_values = []
    pending_values = []

    for invoice in invoices:
        order_amount = invoice.total_amount or Decimal("0.00")
        discount_amount = invoice.discount_amount or Decimal("0.00")
        received_amount = invoice.paid_amount or Decimal("0.00")

        final_payable = invoice.final_payable_amount or Decimal("0.00")

        if final_payable <= 0:
            final_payable = order_amount - discount_amount

        if final_payable < 0:
            final_payable = Decimal("0.00")

        pending_amount = invoice.balance_amount or Decimal("0.00")

        if pending_amount <= 0:
            pending_amount = final_payable - received_amount

        if pending_amount < 0:
            pending_amount = Decimal("0.00")

        difference_amount = order_amount - received_amount

        if difference_amount < 0:
            difference_amount = Decimal("0.00")

        total_order_amount += order_amount
        total_final_payable += final_payable
        total_received += received_amount
        total_discount += discount_amount
        total_pending += pending_amount
        total_difference += difference_amount

        dealer_name = invoice.order.dealer.firm_name if invoice.order and invoice.order.dealer else "Dealer"

        report_rows.append({
            "invoice": invoice,
            "dealer_name": dealer_name,
            "order_amount": order_amount,
            "discount_amount": discount_amount,
            "final_payable": final_payable,
            "received_amount": received_amount,
            "pending_amount": pending_amount,
            "difference_amount": difference_amount,
        })

        chart_labels.append(invoice.invoice_number or f"Invoice {invoice.id}")
        order_values.append(float(order_amount))
        final_payable_values.append(float(final_payable))
        received_values.append(float(received_amount))
        pending_values.append(float(pending_amount))

    return render(
        request,
        "core/order_payment_difference_report.html",
        {
            "period": period,
            "title": title,
            "start_date": start_date,
            "end_date": end_date,
            "report_rows": report_rows,
            "total_order_amount": total_order_amount,
            "total_final_payable": total_final_payable,
            "total_received": total_received,
            "total_discount": total_discount,
            "total_pending": total_pending,
            "total_difference": total_difference,
            "chart_labels": chart_labels,
            "order_values": order_values,
            "final_payable_values": final_payable_values,
            "received_values": received_values,
            "pending_values": pending_values,
            "has_data": invoices.exists(),
        }
    )


def can_view_transport_copies(user):
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in [
        "ADMIN",
        "ACCOUNTANT",
        "WAREHOUSE_MANAGER",
        "PRODUCTION_INCHARGE",
    ]


def clean_indian_phone(phone):
    if not phone:
        return ""

    phone = "".join(ch for ch in str(phone) if ch.isdigit())

    if len(phone) == 10:
        return "91" + phone

    if len(phone) == 12 and phone.startswith("91"):
        return phone

    return phone


@login_required
def transport_copy_list(request):
    if not can_view_transport_copies(request.user):
        messages.error(request, "Only Admin, Accountant or Warehouse team can view transport copies.")
        return redirect("dashboard")

    dealer_id = request.GET.get("dealer", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()
    order_search = request.GET.get("order_id", "").strip()

    invoices = DealerInvoice.objects.select_related(
        "order",
        "order__dealer",
    ).exclude(
        invoice_number__isnull=True
    ).exclude(
        invoice_number=""
    ).order_by("-invoice_date", "-id")

    if dealer_id:
        invoices = invoices.filter(order__dealer_id=dealer_id)

    if from_date:
        try:
            invoices = invoices.filter(invoice_date__gte=date.fromisoformat(from_date))
        except Exception:
            pass

    if to_date:
        try:
            invoices = invoices.filter(invoice_date__lte=date.fromisoformat(to_date))
        except Exception:
            pass

    if order_search:
        search_q = Q(invoice_number__icontains=order_search)

        if order_search.isdigit():
            search_q |= Q(order__id=int(order_search))

        invoices = invoices.filter(search_q)

    invoices = invoices[:300]

    order_ids = [
        invoice.order_id
        for invoice in invoices
        if invoice.order_id
    ]

    dispatch_map = {}

    if order_ids:
        dispatches = DealerDispatch.objects.filter(
            order_id__in=order_ids
        ).order_by("-id")

        for dispatch in dispatches:
            if dispatch.order_id not in dispatch_map:
                dispatch_map[dispatch.order_id] = dispatch

    rows = []

    for invoice in invoices:
        order = invoice.order
        dealer = order.dealer if order else None
        dispatch = dispatch_map.get(order.id) if order else None

        copy_url = request.build_absolute_uri(
            reverse(
                "dealer_invoice_print",
                args=[invoice.id, "TRANSPORTER"]
            )
        )

        phone = ""

        if dispatch and dispatch.driver_phone:
            phone = dispatch.driver_phone
        elif dealer and dealer.phone:
            phone = dealer.phone

        whatsapp_phone = clean_indian_phone(phone)

        share_message = (
            f"Transport Copy\n"
            f"Invoice: {invoice.invoice_number}\n"
            f"Dealer: {dealer.firm_name if dealer else '-'}\n"
            f"Order ID: {order.id if order else '-'}\n\n"
            f"Open Transport Copy:\n{copy_url}"
        )

        whatsapp_url = ""

        if whatsapp_phone:
            whatsapp_url = f"https://wa.me/{whatsapp_phone}?text={quote(share_message)}"

        rows.append({
            "invoice": invoice,
            "order": order,
            "dealer": dealer,
            "dispatch": dispatch,
            "copy_url": copy_url,
            "whatsapp_url": whatsapp_url,
            "phone": phone,
        })

    dealers = Dealer.objects.all().order_by("firm_name")

    return render(
        request,
        "core/transport_copy_list.html",
        {
            "rows": rows,
            "dealers": dealers,
            "selected_dealer": dealer_id,
            "from_date": from_date,
            "to_date": to_date,
            "order_search": order_search,
        }
    )



def can_manage_sales_return(user):
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in ["ADMIN", "ACCOUNTANT"]


def get_user_role(user):
    profile = getattr(user, "profile", None)
    return profile.role if profile else ""


def get_logged_dealer(user):
    try:
        return Dealer.objects.get(user=user)
    except Dealer.DoesNotExist:
        return None


def safe_decimal(value):
    try:
        return Decimal(str(value or "0"))
    except Exception:
        return Decimal("0.00")


def amount_to_words_indian(number):
    number = int(round(float(number or 0)))

    if number == 0:
        return "Rupees Zero only"

    ones = [
        "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
        "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
        "Sixteen", "Seventeen", "Eighteen", "Nineteen"
    ]

    tens = [
        "", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"
    ]

    def two_digit_words(n):
        if n < 20:
            return ones[n]
        return (tens[n // 10] + " " + ones[n % 10]).strip()

    def three_digit_words(n):
        word = ""

        if n >= 100:
            word += ones[n // 100] + " Hundred "
            n = n % 100

        if n:
            word += two_digit_words(n)

        return word.strip()

    parts = []

    crore = number // 10000000
    number %= 10000000

    lakh = number // 100000
    number %= 100000

    thousand = number // 1000
    number %= 1000

    hundred = number

    if crore:
        parts.append(three_digit_words(crore) + " Crore")

    if lakh:
        parts.append(three_digit_words(lakh) + " Lakh")

    if thousand:
        parts.append(three_digit_words(thousand) + " Thousand")

    if hundred:
        parts.append(three_digit_words(hundred))

    return "Rupees " + " ".join(parts).strip() + " only"


def already_returned_boxes(order_item, exclude_credit_note=None):
    qs = SalesReturnCreditNoteItem.objects.filter(
        order_item=order_item,
        credit_note__status="RELEASED"
    )

    if exclude_credit_note:
        qs = qs.exclude(credit_note=exclude_credit_note)

    return qs.aggregate(total=Sum("return_quantity_boxes"))["total"] or 0


def update_invoice_after_credit_notes(invoice):
    total_credit = SalesReturnCreditNote.objects.filter(
        invoice=invoice,
        status="RELEASED"
    ).aggregate(
        total=Sum("grand_total")
    )["total"] or Decimal("0.00")

    invoice.return_credit_amount = total_credit

    base_amount = invoice.final_payable_amount or invoice.total_amount
    net_payable = base_amount - total_credit

    if net_payable < 0:
        net_payable = Decimal("0.00")

    invoice.net_payable_after_return = net_payable
    invoice.pending_amount = invoice.balance_amount

    invoice.save(update_fields=[
        "return_credit_amount",
        "net_payable_after_return",
        "pending_amount",
    ])


def apply_credit_note_to_targets(credit_note, reverse=False):
    amount = credit_note.grand_total or Decimal("0.00")

    if amount <= 0:
        return

    if reverse:
        change_amount = amount
    else:
        change_amount = -amount

    dealer = credit_note.dealer

    dealer.achieved_amount = max(
        Decimal("0.00"),
        (dealer.achieved_amount or Decimal("0.00")) + change_amount
    )
    dealer.save(update_fields=["achieved_amount"])

    sales_officer = dealer.created_by_sales_officer

    if sales_officer:
        target = EmployeeYearlyTarget.objects.filter(
            employee=sales_officer,
            year=credit_note.return_date.year
        ).first()

        if target:
            target.achieved_amount = max(
                Decimal("0.00"),
                (target.achieved_amount or Decimal("0.00")) + change_amount
            )
            target.save(update_fields=["achieved_amount"])


def save_credit_note_items_from_post(credit_note, request, exclude_credit_note=None):
    order_item_ids = request.POST.getlist("order_item_id[]")
    return_quantities = request.POST.getlist("return_quantity_boxes[]")
    taking_prices = request.POST.getlist("price_per_box[]")

    if not order_item_ids:
        raise ValidationError("Please add at least one returned product.")

    created_count = 0

    for index, order_item_id in enumerate(order_item_ids):
        order_item = get_object_or_404(
            DealerOrderItem.objects.select_related(
                "order",
                "product_pack",
                "product_pack__product",
            ),
            id=order_item_id,
            order=credit_note.order
        )

        try:
            return_qty = int(return_quantities[index])
        except Exception:
            return_qty = 0

        if return_qty <= 0:
            continue

        taking_price = safe_decimal(taking_prices[index])

        sold_qty = order_item.quantity_boxes
        returned_qty = already_returned_boxes(
            order_item,
            exclude_credit_note=exclude_credit_note
        )

        balance_qty = sold_qty - int(returned_qty)

        if return_qty > balance_qty:
            raise ValidationError(
                f"{order_item.product_name}: return quantity cannot be more than balance quantity {balance_qty}."
            )

        SalesReturnCreditNoteItem.objects.create(
            credit_note=credit_note,
            order_item=order_item,
            product_pack=order_item.product_pack,
            product_name_snapshot=order_item.product_name,
            pack_label_snapshot=order_item.pack_label,
            sold_quantity_boxes=sold_qty,
            return_quantity_boxes=return_qty,
            units_per_box=order_item.units_per_box,
            total_return_units=Decimal("0.00"),
            unit_name=order_item.unit_name,
            price_per_box=taking_price,
            gst_percent=order_item.gst_percent,
            hsn_code=order_item.hsn_code,
            batch_no=order_item.batch_no,
            mfg_date=order_item.mfg_date,
            expiry_date=order_item.expiry_date,
        )

        created_count += 1

    if created_count == 0:
        raise ValidationError("Please enter return quantity for at least one product.")


@login_required
def ajax_return_dealer_invoices(request):
    if not can_manage_sales_return(request.user):
        return JsonResponse({"success": False, "message": "Permission denied."}, status=403)

    dealer_id = request.GET.get("dealer_id")

    invoices = DealerInvoice.objects.select_related(
        "order",
        "order__dealer"
    ).filter(
        order__dealer_id=dealer_id
    ).order_by("-invoice_date", "-id")[:100]

    invoice_data = []

    for invoice in invoices:
        invoice_data.append({
            "id": invoice.id,
            "invoice_number": invoice.invoice_number,
            "invoice_date": invoice.invoice_date.strftime("%d %b %Y") if invoice.invoice_date else "",
            "order_id": invoice.order.id,
            "amount": str(invoice.total_amount),
            "final_payable": str(invoice.final_payable_amount or invoice.total_amount),
            "return_credit": str(invoice.return_credit_amount),
            "balance": str(invoice.balance_amount),
        })

    return JsonResponse({
        "success": True,
        "invoices": invoice_data,
    })


@login_required
def ajax_return_invoice_items(request):
    if not can_manage_sales_return(request.user):
        return JsonResponse({
            "success": False,
            "message": "Permission denied."
        }, status=403)

    invoice_id = request.GET.get("invoice_id")
    credit_note_id = request.GET.get("credit_note_id")

    if not invoice_id:
        return JsonResponse({
            "success": False,
            "message": "Invoice ID missing."
        })

    exclude_credit_note = None

    if credit_note_id:
        exclude_credit_note = SalesReturnCreditNote.objects.filter(
            id=credit_note_id
        ).first()

    invoice = get_object_or_404(
        DealerInvoice.objects.select_related("order", "order__dealer"),
        id=invoice_id
    )

    items = invoice.order.items.select_related(
        "product_pack",
        "product_pack__product",
    ).all()

    item_data = []

    for item in items:
        already_returned = already_returned_boxes(
            item,
            exclude_credit_note=exclude_credit_note
        )

        balance_qty = item.quantity_boxes - int(already_returned)

        if balance_qty <= 0:
            continue

        item_data.append({
            "order_item_id": item.id,
            "product_name": item.product_name,
            "pack_label": item.pack_label,
            "sold_quantity_boxes": item.quantity_boxes,
            "already_returned_boxes": int(already_returned),
            "balance_quantity_boxes": balance_qty,
            "units_per_box": item.units_per_box,
            "unit_name": item.unit_name,
            "price_per_box": str(item.box_price),
            "gst_percent": str(item.gst_percent),
            "hsn_code": item.hsn_code or "",
            "batch_no": item.batch_no or "",
            "mfg_date": item.mfg_date or "",
            "expiry_date": item.expiry_date or "",
        })

    return JsonResponse({
        "success": True,
        "items": item_data,
    })

@login_required
def sales_return_credit_note_list(request):
    role = get_user_role(request.user)

    if can_manage_sales_return(request.user):
        credit_notes = SalesReturnCreditNote.objects.select_related(
            "dealer",
            "invoice",
            "order",
            "created_by",
        ).prefetch_related("items")
        dealers = Dealer.objects.all().order_by("firm_name")
    elif role == "DEALER":
        dealer = get_logged_dealer(request.user)

        if not dealer:
            messages.error(request, "Dealer profile not found.")
            return redirect("dashboard")

        credit_notes = SalesReturnCreditNote.objects.select_related(
            "dealer",
            "invoice",
            "order",
            "created_by",
        ).prefetch_related("items").filter(
            dealer=dealer,
            status="RELEASED",
            show_to_dealer=True
        )
        dealers = []
    else:
        messages.error(request, "You do not have access to credit notes.")
        return redirect("dashboard")

    search = request.GET.get("search", "").strip()
    dealer_id = request.GET.get("dealer", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()

    if search:
        query = (
            Q(credit_note_number__icontains=search) |
            Q(invoice__invoice_number__icontains=search) |
            Q(dealer__firm_name__icontains=search) |
            Q(dealer__phone__icontains=search) |
            Q(dealer__dealer_code__icontains=search)
        )

        if search.isdigit():
            query |= Q(order__id=int(search))

        credit_notes = credit_notes.filter(query)

    if dealer_id and can_manage_sales_return(request.user):
        credit_notes = credit_notes.filter(dealer_id=dealer_id)

    if from_date:
        credit_notes = credit_notes.filter(return_date__gte=from_date)

    if to_date:
        credit_notes = credit_notes.filter(return_date__lte=to_date)

    total_credit = credit_notes.aggregate(
        total=Sum("grand_total")
    )["total"] or Decimal("0.00")

    return render(
        request,
        "core/sales_return_credit_note_list.html",
        {
            "credit_notes": credit_notes,
            "dealers": dealers,
            "search": search,
            "selected_dealer": dealer_id,
            "from_date": from_date,
            "to_date": to_date,
            "total_credit": total_credit,
            "can_manage": can_manage_sales_return(request.user),
            "role": role,
        }
    )


@login_required
@transaction.atomic
def sales_return_credit_note_create(request):
    if not can_manage_sales_return(request.user):
        messages.error(request, "Only Admin or Accountant can create sales return / credit note.")
        return redirect("dashboard")

    if request.method == "POST":
        selected_dealer = None

        if request.POST.get("dealer"):
            selected_dealer = Dealer.objects.filter(id=request.POST.get("dealer")).first()

        form = SalesReturnCreditNoteForm(request.POST, selected_dealer=selected_dealer)

        if form.is_valid():
            invoice = form.cleaned_data["invoice"]
            dealer = form.cleaned_data["dealer"]

            if invoice.order.dealer_id != dealer.id:
                form.add_error(None, "Selected invoice does not belong to selected dealer.")
            else:
                try:
                    credit_note = form.save(commit=False)
                    credit_note.invoice = invoice
                    credit_note.order = invoice.order
                    credit_note.created_by = request.user
                    credit_note.updated_by = request.user
                    credit_note.status = "RELEASED"
                    credit_note.save()

                    save_credit_note_items_from_post(credit_note, request)

                    credit_note.recalculate_totals(save=True)

                    apply_credit_note_to_targets(credit_note, reverse=False)
                    update_invoice_after_credit_notes(invoice)

                    messages.success(request, "Sales return / credit note created successfully.")
                    return redirect("sales_return_credit_note_list")

                except ValidationError as error:
                    form.add_error(None, error.message if hasattr(error, "message") else str(error))
    else:
        form = SalesReturnCreditNoteForm()

    return render(
        request,
        "core/sales_return_credit_note_form.html",
        {
            "form": form,
            "title": "Create Sales Return / Credit Note",
            "button_text": "Create Credit Note",
            "credit_note": None,
            "existing_items": [],
        }
    )


@login_required
@transaction.atomic
def sales_return_credit_note_edit(request, pk):
    if not can_manage_sales_return(request.user):
        messages.error(request, "Only Admin or Accountant can edit sales return / credit note.")
        return redirect("dashboard")

    credit_note = get_object_or_404(
        SalesReturnCreditNote.objects.select_related("dealer", "invoice", "order"),
        pk=pk
    )

    if request.method == "POST":
        selected_dealer = None

        if request.POST.get("dealer"):
            selected_dealer = Dealer.objects.filter(id=request.POST.get("dealer")).first()

        form = SalesReturnCreditNoteForm(
            request.POST,
            instance=credit_note,
            selected_dealer=selected_dealer
        )

        if form.is_valid():
            invoice = form.cleaned_data["invoice"]
            dealer = form.cleaned_data["dealer"]

            if invoice.order.dealer_id != dealer.id:
                form.add_error(None, "Selected invoice does not belong to selected dealer.")
            else:
                try:
                    old_invoice = credit_note.invoice

                    apply_credit_note_to_targets(credit_note, reverse=True)

                    credit_note.items.all().delete()

                    credit_note = form.save(commit=False)
                    credit_note.invoice = invoice
                    credit_note.order = invoice.order
                    credit_note.updated_by = request.user
                    credit_note.status = "RELEASED"
                    credit_note.save()

                    save_credit_note_items_from_post(
                        credit_note,
                        request,
                        exclude_credit_note=credit_note
                    )

                    credit_note.recalculate_totals(save=True)

                    apply_credit_note_to_targets(credit_note, reverse=False)
                    update_invoice_after_credit_notes(old_invoice)
                    update_invoice_after_credit_notes(invoice)

                    messages.success(request, "Credit note updated successfully.")
                    return redirect("sales_return_credit_note_list")

                except ValidationError as error:
                    form.add_error(None, error.message if hasattr(error, "message") else str(error))
                    apply_credit_note_to_targets(credit_note, reverse=False)
    else:
        form = SalesReturnCreditNoteForm(
            instance=credit_note,
            selected_dealer=credit_note.dealer
        )

    existing_items = []

    for item in credit_note.items.all():
        existing_items.append({
            "order_item_id": item.order_item_id,
            "return_quantity_boxes": item.return_quantity_boxes,
            "price_per_box": str(item.price_per_box),
        })

    return render(
        request,
        "core/sales_return_credit_note_form.html",
        {
            "form": form,
            "title": "Edit Sales Return / Credit Note",
            "button_text": "Update Credit Note",
            "credit_note": credit_note,
            "existing_items": existing_items,
        }
    )


@login_required
def sales_return_credit_note_print(request, pk):
    credit_note = get_object_or_404(
        SalesReturnCreditNote.objects.select_related(
            "dealer",
            "invoice",
            "order",
            "created_by",
        ).prefetch_related("items"),
        pk=pk
    )

    role = get_user_role(request.user)

    if role == "DEALER":
        dealer = get_logged_dealer(request.user)

        if (
            not dealer
            or dealer.id != credit_note.dealer_id
            or not credit_note.show_to_dealer
        ):
            messages.error(request, "This credit note is not available for dealer view.")
            return redirect("sales_return_credit_note_list")

    elif not can_manage_sales_return(request.user):
        messages.error(request, "You cannot view this credit note.")
        return redirect("dashboard")

    return render(
        request,
        "core/sales_return_credit_note_print.html",
        {
            "credit_note": credit_note,
            "dealer": credit_note.dealer,
            "invoice": credit_note.invoice,
            "order": credit_note.order,
            "items": credit_note.items.all(),
            "amount_words": amount_to_words_indian(credit_note.grand_total),
        }
    )



@login_required
def sales_return_credit_note_toggle_dealer_visibility(request, pk):
    if not can_manage_sales_return(request.user):
        messages.error(request, "Only Admin or Accountant can change dealer visibility.")
        return redirect("sales_return_credit_note_list")

    credit_note = get_object_or_404(SalesReturnCreditNote, pk=pk)

    credit_note.show_to_dealer = not credit_note.show_to_dealer
    credit_note.updated_by = request.user
    credit_note.save(update_fields=["show_to_dealer", "updated_by", "updated_at"])

    if credit_note.show_to_dealer:
        messages.success(request, "Credit note is now visible to dealer.")
    else:
        messages.success(request, "Credit note is now hidden from dealer.")

    return redirect("sales_return_credit_note_list")


@login_required
@transaction.atomic
def sales_return_credit_note_delete(request, pk):
    if not can_manage_sales_return(request.user):
        messages.error(request, "Only Admin or Accountant can delete credit notes.")
        return redirect("sales_return_credit_note_list")

    credit_note = get_object_or_404(
        SalesReturnCreditNote.objects.select_related("dealer", "invoice"),
        pk=pk
    )

    invoice = credit_note.invoice

    if request.method == "POST":
        if credit_note.status == "RELEASED":
            apply_credit_note_to_targets(credit_note, reverse=True)

        credit_note.delete()

        update_invoice_after_credit_notes(invoice)

        messages.success(request, "Credit note deleted successfully. Targets and invoice balance updated.")
        return redirect("sales_return_credit_note_list")

    return render(
        request,
        "core/sales_return_credit_note_delete.html",
        {
            "credit_note": credit_note,
        }
    )


def haversine_km(lat1, lon1, lat2, lon2):
    if not lat1 or not lon1 or not lat2 or not lon2:
        return Decimal("0.00")

    lat1 = float(lat1)
    lon1 = float(lon1)
    lat2 = float(lat2)
    lon2 = float(lon2)

    earth_radius_km = 6371

    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)

    a = (
        math.sin(d_lat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(d_lon / 2) ** 2
    )

    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    distance = earth_radius_km * c

    return Decimal(str(round(distance, 2)))


def calculate_attendance_distance(attendance):
    points = list(attendance.gps_points.all().order_by("recorded_at"))

    if len(points) >= 2:
        total = Decimal("0.00")

        for index in range(1, len(points)):
            previous = points[index - 1]
            current = points[index]

            total += haversine_km(
                previous.latitude,
                previous.longitude,
                current.latitude,
                current.longitude
            )

        return total.quantize(Decimal("0.01"))

    return haversine_km(
        attendance.clock_in_latitude,
        attendance.clock_in_longitude,
        attendance.clock_out_latitude,
        attendance.clock_out_longitude
    )


def get_employee_manager(user):
    profile = getattr(user, "profile", None)

    if profile and profile.manager:
        return profile.manager

    return None


def is_manager_of(manager_user, employee_user):
    employee_profile = getattr(employee_user, "profile", None)

    if not employee_profile:
        return False

    return employee_profile.manager_id == manager_user.id


def can_hr(user):
    profile = getattr(user, "profile", None)
    return user.is_superuser or profile.role in ["ADMIN", "HR"]


def can_accountant_admin(user):
    profile = getattr(user, "profile", None)
    return user.is_superuser or profile.role in ["ADMIN", "ACCOUNTANT"]

@login_required
def employee_attendance_dashboard(request):
    profile = getattr(request.user, "profile", None)
    tada_applicable = getattr(profile, "tada_applicable", True)

    today = timezone.localdate()

    my_open_attendance = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date=today,
        clock_out_at__isnull=True
    ).first()

    my_today_attendance = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date=today
    ).first()

    my_month_records = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date__year=today.year,
        attendance_date__month=today.month
    )

    if tada_applicable:
        monthly_km = my_month_records.aggregate(
            total=Sum("manager_approved_km")
        )["total"] or Decimal("0.00")

        monthly_ta = my_month_records.aggregate(
            total=Sum("ta_amount")
        )["total"] or Decimal("0.00")

        monthly_da = my_month_records.aggregate(
            total=Sum("da_amount")
        )["total"] or Decimal("0.00")
    else:
        monthly_km = Decimal("0.00")
        monthly_ta = Decimal("0.00")
        monthly_da = Decimal("0.00")

    pending_manager_count = 0
    pending_hr_count = 0

    manager_subordinates = UserProfile.objects.filter(
        manager=request.user
    ).values_list("user_id", flat=True)

    if manager_subordinates:
        pending_manager_count = EmployeeAttendance.objects.filter(
            employee_id__in=manager_subordinates,
            status="PENDING_MANAGER"
        ).count()

    if can_hr(request.user):
        pending_hr_count = EmployeeAttendance.objects.filter(
            status="MANAGER_APPROVED"
        ).count()

    return render(
        request,
        "core/employee_attendance_dashboard.html",
        {
            "my_open_attendance": my_open_attendance,
            "my_today_attendance": my_today_attendance,
            "monthly_km": monthly_km,
            "monthly_ta": monthly_ta,
            "monthly_da": monthly_da,
            "pending_manager_count": pending_manager_count,
            "pending_hr_count": pending_hr_count,
            "tada_applicable": tada_applicable,
        }
    )



@login_required
def employee_attendance_dashboard(request):
    profile = getattr(request.user, "profile", None)
    tada_applicable = getattr(profile, "tada_applicable", True)

    today = timezone.localdate()

    my_open_attendance = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date=today,
        clock_out_at__isnull=True
    ).first()

    my_today_attendance = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date=today
    ).first()

    my_month_records = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date__year=today.year,
        attendance_date__month=today.month
    )

    if tada_applicable:
        monthly_km = my_month_records.aggregate(
            total=Sum("manager_approved_km")
        )["total"] or Decimal("0.00")

        monthly_ta = my_month_records.aggregate(
            total=Sum("ta_amount")
        )["total"] or Decimal("0.00")

        monthly_da = my_month_records.aggregate(
            total=Sum("da_amount")
        )["total"] or Decimal("0.00")
    else:
        monthly_km = Decimal("0.00")
        monthly_ta = Decimal("0.00")
        monthly_da = Decimal("0.00")

    pending_manager_count = 0
    pending_hr_count = 0

    manager_subordinates = UserProfile.objects.filter(
        manager=request.user
    ).values_list("user_id", flat=True)

    if manager_subordinates:
        pending_manager_count = EmployeeAttendance.objects.filter(
            employee_id__in=manager_subordinates,
            status="PENDING_MANAGER"
        ).count()

    if can_hr(request.user):
        pending_hr_count = EmployeeAttendance.objects.filter(
            status="MANAGER_APPROVED"
        ).count()

    return render(
        request,
        "core/employee_attendance_dashboard.html",
        {
            "my_open_attendance": my_open_attendance,
            "my_today_attendance": my_today_attendance,
            "monthly_km": monthly_km,
            "monthly_ta": monthly_ta,
            "monthly_da": monthly_da,
            "pending_manager_count": pending_manager_count,
            "pending_hr_count": pending_hr_count,
            "tada_applicable": tada_applicable,
        }
    )

@login_required
def employee_clock_in(request):
    today = timezone.localdate()
    profile = getattr(request.user, "profile", None)
    tada_applicable = getattr(profile, "tada_applicable", True)

    existing_open = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date=today,
        clock_out_at__isnull=True
    ).first()

    if existing_open:
        messages.warning(request, "You are already clocked in.")
        return redirect("employee_attendance_dashboard")

    existing_today = EmployeeAttendance.objects.filter(
        employee=request.user,
        attendance_date=today
    ).first()

    if existing_today:
        messages.warning(request, "Attendance already completed for today.")
        return redirect("employee_attendance_dashboard")

    if request.method == "POST":
        latitude = request.POST.get("latitude")
        longitude = request.POST.get("longitude")
        odometer = request.POST.get("clock_in_odometer_reading")

        if not latitude or not longitude:
            messages.error(request, "Location is required for clock-in.")
            return redirect("employee_attendance_dashboard")

        if "can_employee_clockin_today" in globals():
            if not can_employee_clockin_today(request.user, today):
                messages.error(
                    request,
                    "Today is not your working day. Please request extra working day approval from your manager before clock-in."
                )
                return redirect("employee_attendance_dashboard")

        clock_in_odometer = None

        if tada_applicable:
            if not odometer:
                messages.error(request, "Starting odometer reading is required.")
                return redirect("employee_attendance_dashboard")

            try:
                clock_in_odometer = Decimal(str(odometer))
            except Exception:
                messages.error(request, "Please enter a valid starting odometer reading.")
                return redirect("employee_attendance_dashboard")

            if clock_in_odometer < 0:
                messages.error(request, "Starting odometer cannot be negative.")
                return redirect("employee_attendance_dashboard")

        attendance = EmployeeAttendance.objects.create(
            employee=request.user,
            attendance_date=today,
            clock_in_at=timezone.now(),
            clock_in_latitude=latitude,
            clock_in_longitude=longitude,
            clock_in_odometer_reading=clock_in_odometer,
            status="CLOCKED_IN",
        )

        EmployeeGPSTrack.objects.create(
            attendance=attendance,
            latitude=latitude,
            longitude=longitude
        )

        messages.success(request, "Clock-in completed successfully.")
        return redirect("employee_attendance_dashboard")

    return render(
        request,
        "core/employee_clock_in.html",
        {
            "tada_applicable": tada_applicable,
        }
    )

    
@login_required
def employee_gps_ping(request):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "Invalid request"}, status=400)

    attendance = EmployeeAttendance.objects.filter(
        employee=request.user,
        clock_out_at__isnull=True,
        status="CLOCKED_IN"
    ).first()

    if not attendance:
        return JsonResponse({"ok": False, "message": "No active clock-in"}, status=400)

    latitude = request.POST.get("latitude")
    longitude = request.POST.get("longitude")

    if not latitude or not longitude:
        return JsonResponse({"ok": False, "message": "Location missing"}, status=400)

    EmployeeGPSTrack.objects.create(
        attendance=attendance,
        latitude=latitude,
        longitude=longitude
    )

    return JsonResponse({"ok": True})


@login_required
def employee_clock_out(request):
    profile = getattr(request.user, "profile", None)
    tada_applicable = getattr(profile, "tada_applicable", True)

    attendance = EmployeeAttendance.objects.filter(
        employee=request.user,
        clock_out_at__isnull=True
    ).order_by("-clock_in_at").first()

    if not attendance:
        messages.error(request, "No active clock-in found.")
        return redirect("employee_attendance_dashboard")

    if request.method == "POST":
        latitude = request.POST.get("latitude")
        longitude = request.POST.get("longitude")

        if not latitude or not longitude:
            messages.error(request, "Location is required for clock-out.")
            return redirect("employee_clock_out")

        form = EmployeeClockOutForm(
            request.POST,
            request.FILES,
            instance=attendance,
            tada_applicable=tada_applicable
        )

        if form.is_valid():
            attendance = form.save(commit=False)

            attendance.clock_out_at = timezone.now()
            attendance.clock_out_latitude = latitude
            attendance.clock_out_longitude = longitude

            EmployeeGPSTrack.objects.create(
                attendance=attendance,
                latitude=latitude,
                longitude=longitude
            )

            if not tada_applicable:
                attendance.vehicle_type = None
                attendance.clock_out_odometer_reading = None
                attendance.public_vehicle_amount = Decimal("0.00")
                attendance.public_vehicle_bill = None
                attendance.system_distance_km = Decimal("0.00")
                attendance.manager_approved_km = Decimal("0.00")
                attendance.ta_rate_per_km = Decimal("0.00")
                attendance.ta_amount = Decimal("0.00")
                attendance.da_amount = Decimal("0.00")
                attendance.total_claim_amount = Decimal("0.00")
                attendance.dealers_connected_count = 0
                attendance.status = "ATTENDANCE_DONE"
                attendance.save()

                messages.success(request, "Clock-out completed successfully.")
                return redirect("employee_attendance_dashboard")

            start_odometer = attendance.clock_in_odometer_reading or Decimal("0.00")
            end_odometer = attendance.clock_out_odometer_reading or Decimal("0.00")

            if attendance.vehicle_type in ["BIKE", "CAR", "COMPANY"]:
                if end_odometer <= 0:
                    messages.error(request, "Ending odometer reading is required.")
                    return redirect("employee_clock_out")

                if end_odometer < start_odometer:
                    messages.error(request, "Ending odometer cannot be less than starting odometer.")
                    return redirect("employee_clock_out")

                distance = end_odometer - start_odometer
            else:
                distance = Decimal("0.00")

            attendance.system_distance_km = distance
            attendance.manager_approved_km = distance

            if attendance.vehicle_type == "BIKE":
                attendance.ta_rate_per_km = profile.bike_ta_per_km if profile else Decimal("0.00")
                attendance.ta_amount = distance * attendance.ta_rate_per_km

            elif attendance.vehicle_type == "CAR":
                attendance.ta_rate_per_km = profile.car_ta_per_km if profile else Decimal("0.00")
                attendance.ta_amount = distance * attendance.ta_rate_per_km

            elif attendance.vehicle_type == "PUBLIC":
                attendance.ta_rate_per_km = Decimal("0.00")
                attendance.ta_amount = attendance.public_vehicle_amount or Decimal("0.00")

            elif attendance.vehicle_type == "COMPANY":
                attendance.ta_rate_per_km = Decimal("0.00")
                attendance.ta_amount = Decimal("0.00")

            attendance.da_amount = profile.daily_da_amount if profile else Decimal("0.00")
            attendance.total_claim_amount = (
                (attendance.ta_amount or Decimal("0.00"))
                + (attendance.da_amount or Decimal("0.00"))
            )

            attendance.dealers_connected_count = 0
            attendance.status = "PENDING_MANAGER"
            attendance.save()

            messages.success(
                request,
                "Clock-out completed. TA/DA claim sent for manager approval."
            )
            return redirect("employee_attendance_dashboard")

    else:
        form = EmployeeClockOutForm(
            instance=attendance,
            tada_applicable=tada_applicable
        )

    return render(
        request,
        "core/employee_clock_out.html",
        {
            "form": form,
            "attendance": attendance,
            "tada_applicable": tada_applicable,
        }
    )
    



@login_required
def manager_attendance_approvals(request):
    subordinate_ids = UserProfile.objects.filter(
        manager=request.user
    ).values_list("user_id", flat=True)

    records = EmployeeAttendance.objects.select_related(
        "employee",
        "employee__profile"
    ).filter(
        employee_id__in=subordinate_ids,
        status="PENDING_MANAGER"
    )

    return render(
        request,
        "core/manager_attendance_approvals.html",
        {
            "records": records,
        }
    )


@login_required
def manager_approve_attendance(request, attendance_id):
    attendance = get_object_or_404(EmployeeAttendance, id=attendance_id)

    if not is_manager_of(request.user, attendance.employee):
        messages.error(request, "You are not allowed to approve this employee claim.")
        return redirect("manager_attendance_approvals")

    if request.method == "POST":
        form = ManagerAttendanceApprovalForm(request.POST, instance=attendance)

        if form.is_valid():
            attendance = form.save(commit=False)

            if attendance.manager_approved_km > attendance.system_distance_km:
                messages.error(request, "Approved KM cannot be more than system tracked KM.")
                return redirect("manager_approve_attendance", attendance_id=attendance.id)

            if attendance.vehicle_type == "BIKE":
                attendance.ta_amount = attendance.manager_approved_km * attendance.ta_rate_per_km

            elif attendance.vehicle_type == "CAR":
                attendance.ta_amount = attendance.manager_approved_km * attendance.ta_rate_per_km

            elif attendance.vehicle_type == "PUBLIC":
                attendance.ta_amount = attendance.public_vehicle_amount or Decimal("0.00")

            elif attendance.vehicle_type == "COMPANY":
                attendance.ta_amount = Decimal("0.00")

            attendance.total_claim_amount = attendance.ta_amount + attendance.da_amount
            attendance.manager_approved_by = request.user
            attendance.manager_approved_at = timezone.now()
            attendance.status = "MANAGER_APPROVED"
            attendance.save()

            messages.success(request, "Claim approved and sent to HR.")
            return redirect("manager_attendance_approvals")
    else:
        form = ManagerAttendanceApprovalForm(instance=attendance)

    return render(
        request,
        "core/manager_approve_attendance.html",
        {
            "attendance": attendance,
            "form": form,
        }
    )


@login_required
def hr_attendance_claims(request):
    if not can_hr(request.user):
        messages.error(request, "Only HR/Admin can view HR claims.")
        return redirect("dashboard")

    records = EmployeeAttendance.objects.select_related(
        "employee",
        "employee__profile",
        "manager_approved_by"
    ).filter(
        status="MANAGER_APPROVED"
    )

    return render(
        request,
        "core/hr_attendance_claims.html",
        {
            "records": records,
        }
    )


@login_required
def hr_approve_attendance(request, attendance_id):
    if not can_hr(request.user):
        messages.error(request, "Only HR/Admin can approve claims.")
        return redirect("dashboard")

    attendance = get_object_or_404(EmployeeAttendance, id=attendance_id)

    if request.method == "POST":
        form = HRClaimApprovalForm(request.POST, instance=attendance)

        if form.is_valid():
            attendance = form.save(commit=False)
            attendance.hr_approved_by = request.user
            attendance.hr_approved_at = timezone.now()
            attendance.status = "HR_APPROVED"
            attendance.save()

            messages.success(request, "Claim approved by HR. Accountant can release funds.")
            return redirect("hr_attendance_claims")
    else:
        form = HRClaimApprovalForm(instance=attendance)

    return render(
        request,
        "core/hr_approve_attendance.html",
        {
            "attendance": attendance,
            "form": form,
        }
    )


@login_required
def accountant_release_attendance_claim(request, attendance_id):
    if not can_accountant_admin(request.user):
        messages.error(request, "Only Accountant/Admin can release funds.")
        return redirect("dashboard")

    attendance = get_object_or_404(
        EmployeeAttendance,
        id=attendance_id,
        status="HR_APPROVED"
    )

    if request.method == "POST":
        attendance.released_by = request.user
        attendance.released_at = timezone.now()
        attendance.release_remarks = request.POST.get("release_remarks", "")
        attendance.status = "RELEASED"
        attendance.save()

        messages.success(request, "TA/DA amount released successfully.")
        return redirect("accountant_attendance_claims")

    return render(
        request,
        "core/accountant_release_attendance_claim.html",
        {
            "attendance": attendance,
        }
    )


@login_required
def accountant_attendance_claims(request):
    if not can_accountant_admin(request.user):
        messages.error(request, "Only Accountant/Admin can view claims.")
        return redirect("dashboard")

    today = timezone.localdate()

    period = request.GET.get("period", "THIS_MONTH")
    employee_id = request.GET.get("employee", "").strip()

    if period == "LAST_15":
        from_date = today - timedelta(days=14)
        to_date = today

    elif period == "LAST_MONTH":
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        from_date = last_day_last_month.replace(day=1)
        to_date = last_day_last_month

    elif period == "CUSTOM":
        from_date = parse_date(request.GET.get("from_date") or "")
        to_date = parse_date(request.GET.get("to_date") or "")

        if not from_date:
            from_date = today.replace(day=1)

        if not to_date:
            to_date = today

    else:
        period = "THIS_MONTH"
        from_date = today.replace(day=1)
        to_date = today

    records = EmployeeAttendance.objects.select_related(
        "employee",
        "employee__profile",
        "hr_approved_by"
    ).filter(
        status="HR_APPROVED",
        attendance_date__gte=from_date,
        attendance_date__lte=to_date,
    ).order_by(
        "employee__first_name",
        "employee__username",
        "attendance_date"
    )

    if employee_id:
        records = records.filter(employee_id=employee_id)

    employee_map = {}

    for record in records:
        emp = record.employee

        if emp.id not in employee_map:
            employee_map[emp.id] = {
                "employee": emp,
                "records": [],
                "claim_days": 0,
                "total_km": Decimal("0.00"),
                "total_ta": Decimal("0.00"),
                "total_da": Decimal("0.00"),
                "total_amount": Decimal("0.00"),
                "first_date": record.attendance_date,
                "last_date": record.attendance_date,
            }

        row = employee_map[emp.id]

        row["records"].append(record)
        row["claim_days"] += 1
        row["total_km"] += Decimal(str(record.manager_approved_km or 0))
        row["total_ta"] += Decimal(str(record.ta_amount or 0))
        row["total_da"] += Decimal(str(record.da_amount or 0))
        row["total_amount"] += Decimal(str(record.total_claim_amount or 0))

        if record.attendance_date < row["first_date"]:
            row["first_date"] = record.attendance_date

        if record.attendance_date > row["last_date"]:
            row["last_date"] = record.attendance_date

    rows = list(employee_map.values())

    total_pending_amount = sum(row["total_amount"] for row in rows)
    total_pending_days = sum(row["claim_days"] for row in rows)

    employees = User.objects.filter(
        attendance_records__status="HR_APPROVED"
    ).distinct().order_by("first_name", "username")

    recent_batches = EmployeeTadaFundReleaseBatch.objects.select_related(
        "employee",
        "released_by"
    ).order_by("-released_at", "-id")[:20]

    return render(
        request,
        "core/accountant_attendance_claims.html",
        {
            "rows": rows,
            "employees": employees,
            "employee_id": employee_id,
            "period": period,
            "from_date": from_date,
            "to_date": to_date,
            "total_pending_amount": total_pending_amount,
            "total_pending_days": total_pending_days,
            "recent_batches": recent_batches,
        }
    )

@login_required
@transaction.atomic
def accountant_bulk_release_attendance_claims(request):
    if not can_accountant_admin(request.user):
        messages.error(request, "Only Accountant/Admin can release funds.")
        return redirect("dashboard")

    if request.method != "POST":
        return redirect("accountant_attendance_claims")

    employee_id = request.POST.get("employee_id")
    from_date = parse_date(request.POST.get("from_date") or "")
    to_date = parse_date(request.POST.get("to_date") or "")
    remarks = request.POST.get("release_remarks", "").strip()

    if not employee_id or not from_date or not to_date:
        messages.error(request, "Employee and date range are required.")
        return redirect("accountant_attendance_claims")

    records = EmployeeAttendance.objects.select_for_update().select_related(
        "employee"
    ).filter(
        employee_id=employee_id,
        status="HR_APPROVED",
        attendance_date__gte=from_date,
        attendance_date__lte=to_date,
    ).order_by("attendance_date")

    if not records.exists():
        messages.error(request, "No HR approved TA/DA claims found for this employee and date range.")
        return redirect("accountant_attendance_claims")

    total_km = Decimal("0.00")
    total_ta = Decimal("0.00")
    total_da = Decimal("0.00")
    total_amount = Decimal("0.00")

    record_list = list(records)

    for record in record_list:
        total_km += Decimal(str(record.manager_approved_km or 0))
        total_ta += Decimal(str(record.ta_amount or 0))
        total_da += Decimal(str(record.da_amount or 0))
        total_amount += Decimal(str(record.total_claim_amount or 0))

    batch = EmployeeTadaFundReleaseBatch.objects.create(
        employee_id=employee_id,
        from_date=from_date,
        to_date=to_date,
        total_km=total_km,
        total_ta_amount=total_ta,
        total_da_amount=total_da,
        total_released_amount=total_amount,
        released_by=request.user,
        released_at=timezone.now(),
        remarks=remarks,
    )

    batch.attendance_records.set(record_list)

    for record in record_list:
        record.released_by = request.user
        record.released_at = timezone.now()
        record.release_remarks = remarks
        record.status = "RELEASED"
        record.save(update_fields=[
            "released_by",
            "released_at",
            "release_remarks",
            "status",
        ])

    messages.success(
        request,
        f"Released ₹{total_amount} TA/DA for {len(record_list)} day(s)."
    )

    return redirect("accountant_attendance_claims")


@login_required
def employee_leave_request_create(request):
    if request.method == "POST":
        form = EmployeeLeaveRequestForm(request.POST)

        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = request.user
            leave.deduction_days = leave.calculate_days()

            if leave.leave_type == "UNPAID":
                leave.salary_deduction_required = True

            leave.save()

            messages.success(request, "Leave request sent to your manager.")
            return redirect("employee_leave_list")
    else:
        form = EmployeeLeaveRequestForm()

    return render(
        request,
        "core/employee_leave_form.html",
        {
            "form": form,
            "title": "Apply Leave",
        }
    )


@login_required
def employee_leave_list(request):
    profile = getattr(request.user, "profile", None)

    if profile and profile.role in ["ADMIN", "HR", "ACCOUNTANT"] or request.user.is_superuser:
        leaves = EmployeeLeaveRequest.objects.select_related(
            "employee",
            "approved_by"
        ).all()
    else:
        leaves = EmployeeLeaveRequest.objects.select_related(
            "employee",
            "approved_by"
        ).filter(employee=request.user)

    return render(
        request,
        "core/employee_leave_list.html",
        {
            "leaves": leaves,
        }
    )


@login_required
def manager_leave_approvals(request):
    subordinate_ids = UserProfile.objects.filter(
        manager=request.user
    ).values_list("user_id", flat=True)

    leaves = EmployeeLeaveRequest.objects.select_related(
        "employee",
        "employee__profile"
    ).filter(
        employee_id__in=subordinate_ids,
        status="PENDING"
    )

    return render(
        request,
        "core/manager_leave_approvals.html",
        {
            "leaves": leaves,
        }
    )


@login_required
def manager_approve_leave(request, leave_id):
    leave = get_object_or_404(EmployeeLeaveRequest, id=leave_id)

    if not is_manager_of(request.user, leave.employee):
        messages.error(request, "You are not allowed to approve this leave.")
        return redirect("manager_leave_approvals")

    leave.status = "APPROVED"
    leave.approved_by = request.user
    leave.approved_at = timezone.now()
    leave.manager_remarks = request.POST.get("manager_remarks", "")
    leave.salary_deduction_required = True if leave.leave_type == "UNPAID" else False
    leave.deduction_days = leave.calculate_days()
    leave.save()

    messages.success(request, "Leave approved.")
    return redirect("manager_leave_approvals")


@login_required
def manager_reject_leave(request, leave_id):
    leave = get_object_or_404(EmployeeLeaveRequest, id=leave_id)

    if not is_manager_of(request.user, leave.employee):
        messages.error(request, "You are not allowed to reject this leave.")
        return redirect("manager_leave_approvals")

    leave.status = "REJECTED"
    leave.approved_by = request.user
    leave.approved_at = timezone.now()
    leave.manager_remarks = request.POST.get("manager_remarks", "")
    leave.salary_deduction_required = True
    leave.deduction_days = leave.calculate_days()
    leave.save()

    messages.success(request, "Leave rejected. Salary deduction marked.")
    return redirect("manager_leave_approvals")


@login_required
def hr_management_dashboard(request):
    user = request.user
    profile = getattr(user, "profile", None)
    role = profile.role if profile else ""

    if not (
        user.is_superuser
        or role in ["ADMIN", "HR", "ACCOUNTANT"]
    ):
        messages.error(request, "Only Admin, HR and Accountant can view HR dashboard.")
        return redirect("dashboard")

    today = timezone.localdate()
    selected_filter = request.GET.get("filter", "this_month")

    if selected_filter == "today":
        start_date = today
        end_date = today

    elif selected_filter == "last_month":
        first_day_this_month = today.replace(day=1)
        last_month_last_day = first_day_this_month - timedelta(days=1)
        start_date = last_month_last_day.replace(day=1)
        end_date = last_month_last_day

    elif selected_filter == "custom":
        start_date_raw = request.GET.get("start_date")
        end_date_raw = request.GET.get("end_date")

        try:
            start_date = timezone.datetime.strptime(start_date_raw, "%Y-%m-%d").date()
            end_date = timezone.datetime.strptime(end_date_raw, "%Y-%m-%d").date()
        except Exception:
            start_date = today.replace(day=1)
            end_date = today

    else:
        start_date = today.replace(day=1)
        end_date = today

    employee_profiles = UserProfile.objects.select_related("user").filter(
        user__is_active=True
    ).exclude(
        role="DEALER"
    )

    total_employees = employee_profiles.count()

    today_attendance_qs = EmployeeAttendance.objects.select_related(
        "employee",
        "employee__profile"
    ).filter(
        attendance_date=today
    )

    today_clocked_user_ids = set(
        today_attendance_qs.values_list("employee_id", flat=True)
    )

    clocked_in_today_count = len(today_clocked_user_ids)

    active_clockin_count = today_attendance_qs.filter(
        clock_out_at__isnull=True,
        status="CLOCKED_IN"
    ).count()

    on_time_count = 0
    late_count = 0
    no_timing_count = 0
    not_clocked_in_count = 0

    late_employees = []
    not_clocked_employees = []

    today_attendance_map = {
        attendance.employee_id: attendance
        for attendance in today_attendance_qs
    }

    for emp_profile in employee_profiles:
        attendance = today_attendance_map.get(emp_profile.user_id)

        if not attendance or not attendance.clock_in_at:
            not_clocked_in_count += 1
            not_clocked_employees.append(emp_profile)
            continue

        if not emp_profile.duty_start_time:
            no_timing_count += 1
            continue

        clock_in_time = timezone.localtime(attendance.clock_in_at).time()

        if clock_in_time <= emp_profile.duty_start_time:
            on_time_count += 1
        else:
            late_count += 1
            late_employees.append({
                "employee": emp_profile.user,
                "role": emp_profile.get_role_display(),
                "clock_in": attendance.clock_in_at,
                "duty_start": emp_profile.duty_start_time,
            })

    range_attendance_qs = EmployeeAttendance.objects.select_related(
        "employee",
        "employee__profile"
    ).filter(
        attendance_date__range=[start_date, end_date]
    )

    completed_attendance_qs = range_attendance_qs.filter(
        clock_out_at__isnull=False
    )

    total_km = completed_attendance_qs.aggregate(
        total=Sum("manager_approved_km")
    )["total"] or Decimal("0.00")

    total_system_km = completed_attendance_qs.aggregate(
        total=Sum("system_distance_km")
    )["total"] or Decimal("0.00")

    total_ta = completed_attendance_qs.aggregate(
        total=Sum("ta_amount")
    )["total"] or Decimal("0.00")

    total_da = completed_attendance_qs.aggregate(
        total=Sum("da_amount")
    )["total"] or Decimal("0.00")

    total_claim = completed_attendance_qs.aggregate(
        total=Sum("total_claim_amount")
    )["total"] or Decimal("0.00")

    total_dealer_visits = completed_attendance_qs.aggregate(
        total=Sum("dealers_connected_count")
    )["total"] or 0

    pending_manager_claims = EmployeeAttendance.objects.filter(
        status="PENDING_MANAGER"
    ).count()

    pending_hr_claims = EmployeeAttendance.objects.filter(
        status="MANAGER_APPROVED"
    ).count()

    pending_accountant_release = EmployeeAttendance.objects.filter(
        status="HR_APPROVED"
    ).count()

    released_claims = EmployeeAttendance.objects.filter(
        status="RELEASED",
        attendance_date__range=[start_date, end_date]
    ).count()

    leave_pending_count = EmployeeLeaveRequest.objects.filter(
        status="PENDING"
    ).count()

    unpaid_leave_count = EmployeeLeaveRequest.objects.filter(
        status="APPROVED",
        salary_deduction_required=True,
        from_date__lte=end_date,
        to_date__gte=start_date,
    ).count()

    top_performers = completed_attendance_qs.values(
        "employee_id",
        "employee__first_name",
        "employee__last_name",
        "employee__username",
        "employee__profile__role",
    ).annotate(
        total_km=Sum("manager_approved_km"),
        system_km=Sum("system_distance_km"),
        total_ta=Sum("ta_amount"),
        total_da=Sum("da_amount"),
        total_claim=Sum("total_claim_amount"),
        dealer_visits=Sum("dealers_connected_count"),
        working_days=Count("id"),
    ).order_by(
        "-dealer_visits",
        "-total_km"
    )[:10]

    top_km_performers = completed_attendance_qs.values(
        "employee_id",
        "employee__first_name",
        "employee__last_name",
        "employee__username",
    ).annotate(
        total_km=Sum("manager_approved_km"),
        dealer_visits=Sum("dealers_connected_count"),
    ).order_by("-total_km")[:8]

    best_field_officer = None
    if top_performers:
        best_field_officer = top_performers[0]

    daily_labels = []
    daily_clockins = []
    daily_late = []

    current_date = start_date

    while current_date <= end_date:
        day_records = EmployeeAttendance.objects.select_related(
            "employee",
            "employee__profile"
        ).filter(
            attendance_date=current_date
        )

        day_clockins = day_records.count()
        day_late_count = 0

        for record in day_records:
            emp_profile = getattr(record.employee, "profile", None)

            if emp_profile and emp_profile.duty_start_time and record.clock_in_at:
                clock_in_time = timezone.localtime(record.clock_in_at).time()

                if clock_in_time > emp_profile.duty_start_time:
                    day_late_count += 1

        daily_labels.append(current_date.strftime("%d %b"))
        daily_clockins.append(day_clockins)
        daily_late.append(day_late_count)

        current_date += timedelta(days=1)

    top_km_labels = []
    top_km_values = []
    top_visit_values = []

    for row in top_km_performers:
        full_name = (
            f"{row['employee__first_name']} {row['employee__last_name']}"
        ).strip()

        if not full_name:
            full_name = row["employee__username"]

        top_km_labels.append(full_name)
        top_km_values.append(float(row["total_km"] or 0))
        top_visit_values.append(int(row["dealer_visits"] or 0))

    attendance_pie_data = [
        on_time_count,
        late_count,
        not_clocked_in_count,
    ]

    claim_pie_data = [
        pending_manager_claims,
        pending_hr_claims,
        pending_accountant_release,
        released_claims,
    ]

    vehicle_summary = completed_attendance_qs.values(
        "vehicle_type"
    ).annotate(
        count=Count("id"),
        km=Sum("manager_approved_km"),
        amount=Sum("total_claim_amount"),
    ).order_by("vehicle_type")

    vehicle_labels = []
    vehicle_values = []

    for item in vehicle_summary:
        vehicle_labels.append(item["vehicle_type"] or "Not Set")
        vehicle_values.append(float(item["amount"] or 0))

    return render(
        request,
        "core/hr_management_dashboard.html",
        {
            "selected_filter": selected_filter,
            "start_date": start_date,
            "end_date": end_date,

            "total_employees": total_employees,
            "clocked_in_today_count": clocked_in_today_count,
            "active_clockin_count": active_clockin_count,
            "on_time_count": on_time_count,
            "late_count": late_count,
            "no_timing_count": no_timing_count,
            "not_clocked_in_count": not_clocked_in_count,

            "total_km": total_km,
            "total_system_km": total_system_km,
            "total_ta": total_ta,
            "total_da": total_da,
            "total_claim": total_claim,
            "total_dealer_visits": total_dealer_visits,

            "pending_manager_claims": pending_manager_claims,
            "pending_hr_claims": pending_hr_claims,
            "pending_accountant_release": pending_accountant_release,
            "released_claims": released_claims,

            "leave_pending_count": leave_pending_count,
            "unpaid_leave_count": unpaid_leave_count,

            "top_performers": top_performers,
            "best_field_officer": best_field_officer,
            "late_employees": late_employees,
            "not_clocked_employees": not_clocked_employees[:10],

            "daily_labels": daily_labels,
            "daily_clockins": daily_clockins,
            "daily_late": daily_late,

            "attendance_pie_data": attendance_pie_data,
            "claim_pie_data": claim_pie_data,

            "top_km_labels": top_km_labels,
            "top_km_values": top_km_values,
            "top_visit_values": top_visit_values,

            "vehicle_labels": vehicle_labels,
            "vehicle_values": vehicle_values,
        }
    )


def can_manage_payslips(user):
    profile = getattr(user, "profile", None)

    return (
        user.is_superuser
        or (profile and profile.role in ["ADMIN", "HR", "ACCOUNTANT"])
    )


@login_required
def payslip_list(request):
    if not can_manage_payslips(request.user):
        messages.error(request, "Only Admin, HR or Accountant can manage payslips.")
        return redirect("dashboard")

    status = request.GET.get("status", "DRAFT")
    month = request.GET.get("month")
    year = request.GET.get("year")

    payslips = EmployeePayslip.objects.select_related(
        "employee",
        "employee__profile",
        "generated_by",
        "released_by",
    ).all()

    if status:
        payslips = payslips.filter(status=status)

    if month:
        payslips = payslips.filter(month=month)

    if year:
        payslips = payslips.filter(year=year)

    return render(
        request,
        "core/payslip_list.html",
        {
            "payslips": payslips,
            "status": status,
            "month": month,
            "year": year,
        }
    )


@login_required
def payslip_generate(request):
    if not can_manage_payslips(request.user):
        messages.error(request, "Only Admin, HR or Accountant can generate payslips.")
        return redirect("dashboard")

    def safe_decimal(value):
        try:
            return Decimal(str(value or "0.00"))
        except Exception:
            return Decimal("0.00")

    if request.method == "POST":
        form = PayslipGenerateForm(request.POST)

        if form.is_valid():
            generate_type = form.cleaned_data["generate_type"]
            employee = form.cleaned_data.get("employee")
            month = int(form.cleaned_data["month"])
            year = int(form.cleaned_data["year"])

            if generate_type == "INDIVIDUAL":
                employees = [employee]
            else:
                employees = User.objects.filter(
                    is_active=True,
                    profile__isnull=False
                ).exclude(
                    profile__role="DEALER"
                ).select_related("profile")

            created_count = 0
            updated_count = 0
            skipped_count = 0

            for emp in employees:
                profile = getattr(emp, "profile", None)

                if not profile:
                    skipped_count += 1
                    continue

                existing = EmployeePayslip.objects.filter(
                    employee=emp,
                    month=month,
                    year=year,
                ).first()

                if existing and existing.status == "RELEASED":
                    skipped_count += 1
                    continue

                employee_code = profile.employee_code or emp.id
                payslip_number = f"PS-{year}{month:02d}-{employee_code}"

                salary = safe_decimal(profile.salary_amount)

                # -----------------------------
                # PF DEDUCTION CALCULATION
                # -----------------------------
                pf_deduction_amount = Decimal("0.00")
                pf_percentage = safe_decimal(
                    getattr(profile, "pf_deduction_percentage", Decimal("0.00"))
                )

                if getattr(profile, "pf_applicable", False):
                    pf_deduction_amount = (salary * pf_percentage) / Decimal("100.00")

                # -----------------------------
                # INSURANCE DEDUCTION
                # -----------------------------
                insurance_deduction_amount = Decimal("0.00")

                if getattr(profile, "insurance_applicable", False):
                    insurance_deduction_amount = safe_decimal(
                        getattr(profile, "insurance_deduction_amount", Decimal("0.00"))
                    )

                # -----------------------------
                # UPDATE EXISTING DRAFT PAYSLIP
                # -----------------------------
                if existing:
                    existing.payslip_number = payslip_number
                    existing.basic_salary = salary

                    existing.allowance_amount = existing.allowance_amount or Decimal("0.00")
                    existing.leave_deduction_amount = existing.leave_deduction_amount or Decimal("0.00")
                    existing.other_deduction_amount = existing.other_deduction_amount or Decimal("0.00")

                    existing.pf_deduction_amount = pf_deduction_amount
                    existing.insurance_deduction_amount = insurance_deduction_amount

                    # Employee payroll snapshot
                    existing.employee_dob = getattr(profile, "date_of_birth", None)
                    existing.employee_pan_number = getattr(profile, "pan_number", "")

                    existing.employee_bank_name = getattr(profile, "bank_name", "")
                    existing.employee_bank_account_number = getattr(profile, "bank_account_number", "")
                    existing.employee_bank_ifsc_code = getattr(profile, "bank_ifsc_code", "")
                    existing.employee_bank_account_holder_name = getattr(profile, "bank_account_holder_name", "")

                    existing.employee_pf_uan_number = getattr(profile, "pf_uan_number", "")
                    existing.employee_pf_percentage = pf_percentage

                    existing.employee_insurance_provider_name = getattr(profile, "insurance_provider_name", "")
                    existing.employee_insurance_policy_number = getattr(profile, "insurance_policy_number", "")

                    existing.generated_by = request.user
                    existing.status = "DRAFT"

                    existing.save()
                    updated_count += 1

                # -----------------------------
                # CREATE NEW DRAFT PAYSLIP
                # -----------------------------
                else:
                    EmployeePayslip.objects.create(
                        employee=emp,
                        payslip_number=payslip_number,
                        month=month,
                        year=year,

                        basic_salary=salary,
                        allowance_amount=Decimal("0.00"),
                        leave_deduction_amount=Decimal("0.00"),
                        other_deduction_amount=Decimal("0.00"),

                        pf_deduction_amount=pf_deduction_amount,
                        insurance_deduction_amount=insurance_deduction_amount,

                        # Employee payroll snapshot
                        employee_dob=getattr(profile, "date_of_birth", None),
                        employee_pan_number=getattr(profile, "pan_number", ""),

                        employee_bank_name=getattr(profile, "bank_name", ""),
                        employee_bank_account_number=getattr(profile, "bank_account_number", ""),
                        employee_bank_ifsc_code=getattr(profile, "bank_ifsc_code", ""),
                        employee_bank_account_holder_name=getattr(profile, "bank_account_holder_name", ""),

                        employee_pf_uan_number=getattr(profile, "pf_uan_number", ""),
                        employee_pf_percentage=pf_percentage,

                        employee_insurance_provider_name=getattr(profile, "insurance_provider_name", ""),
                        employee_insurance_policy_number=getattr(profile, "insurance_policy_number", ""),

                        generated_by=request.user,
                        status="DRAFT",
                    )

                    created_count += 1

            messages.success(
                request,
                f"Payslips prepared for review. Created: {created_count}, Updated: {updated_count}, Skipped released: {skipped_count}."
            )

            return redirect(
                f"{reverse('payslip_list')}?status=DRAFT&month={month}&year={year}"
            )

    else:
        form = PayslipGenerateForm()

    return render(
        request,
        "core/payslip_generate.html",
        {
            "form": form,
            "title": "Generate Payslips",
        }
    )


@login_required
def payslip_detail(request, payslip_id):
    payslip = get_object_or_404(
        EmployeePayslip.objects.select_related(
            "employee",
            "employee__profile",
            "generated_by",
            "released_by",
        ),
        id=payslip_id
    )

    if not can_manage_payslips(request.user):
        if payslip.employee != request.user or payslip.status != "RELEASED":
            messages.error(request, "You cannot view this payslip.")
            return redirect("dashboard")

    return render(
        request,
        "core/payslip_detail.html",
        {
            "payslip": payslip,
        }
    )


@login_required
def payslip_edit(request, payslip_id):
    if not can_manage_payslips(request.user):
        messages.error(request, "Only Admin, HR or Accountant can edit payslips.")
        return redirect("dashboard")

    payslip = get_object_or_404(EmployeePayslip, id=payslip_id)

    if payslip.status == "RELEASED":
        messages.error(request, "Released payslip cannot be edited.")
        return redirect("payslip_detail", payslip_id=payslip.id)

    if request.method == "POST":
        form = PayslipEditForm(request.POST, instance=payslip)

        if form.is_valid():
            form.save()
            messages.success(request, "Payslip updated. Please review before release.")
            return redirect("payslip_detail", payslip_id=payslip.id)
    else:
        form = PayslipEditForm(instance=payslip)

    return render(
        request,
        "core/payslip_edit.html",
        {
            "form": form,
            "payslip": payslip,
            "title": "Edit Payslip",
        }
    )


@login_required
def payslip_release(request, payslip_id):
    if not can_manage_payslips(request.user):
        messages.error(request, "Only Admin, HR or Accountant can release payslips.")
        return redirect("dashboard")

    payslip = get_object_or_404(EmployeePayslip, id=payslip_id)

    if request.method == "POST":
        payslip.status = "RELEASED"
        payslip.released_by = request.user
        payslip.released_at = timezone.now()
        payslip.save()

        messages.success(request, "Payslip released successfully.")
        return redirect("payslip_detail", payslip_id=payslip.id)

    return redirect("payslip_detail", payslip_id=payslip.id)


@login_required
def payslip_release_bulk(request):
    if not can_manage_payslips(request.user):
        messages.error(request, "Only Admin, HR or Accountant can release payslips.")
        return redirect("dashboard")

    if request.method == "POST":
        month = request.POST.get("month")
        year = request.POST.get("year")

        if not month or not year:
            messages.error(request, "Month and year are required for bulk release.")
            return redirect("payslip_list")

        payslips = EmployeePayslip.objects.filter(
            month=month,
            year=year,
            status="DRAFT",
        )

        count = payslips.count()

        payslips.update(
            status="RELEASED",
            released_by=request.user,
            released_at=timezone.now(),
        )

        messages.success(request, f"{count} payslip(s) released successfully.")
        return redirect(f"{reverse('payslip_list')}?status=RELEASED&month={month}&year={year}")

    return redirect("payslip_list")


@login_required
def my_payslips(request):
    payslips = EmployeePayslip.objects.filter(
        employee=request.user,
        status="RELEASED",
    ).order_by("-year", "-month")

    return render(
        request,
        "core/my_payslips.html",
        {
            "payslips": payslips,
        }
    )


@login_required
def payslip_print(request, payslip_id):
    payslip = get_object_or_404(
        EmployeePayslip.objects.select_related(
            "employee",
            "employee__profile",
            "generated_by",
            "released_by",
        ),
        id=payslip_id
    )

    if not can_manage_payslips(request.user):
        if payslip.employee != request.user or payslip.status != "RELEASED":
            messages.error(request, "You cannot print this payslip.")
            return redirect("dashboard")

    return render(
        request,
        "core/payslip_print.html",
        {
            "payslip": payslip,
        }
    )


@login_required
def employee_tada_calendar(request):
    user = request.user
    profile = getattr(user, "profile", None)

    if not profile or profile.role == "DEALER":
        messages.error(request, "TA/DA calendar is available only for employees.")
        return redirect("dashboard")

    today = timezone.localdate()

    try:
        selected_month = int(request.GET.get("month", today.month))
        selected_year = int(request.GET.get("year", today.year))
    except Exception:
        selected_month = today.month
        selected_year = today.year

    if selected_month < 1 or selected_month > 12:
        selected_month = today.month

    month_start = date(selected_year, selected_month, 1)
    last_day = py_calendar.monthrange(selected_year, selected_month)[1]
    month_end = date(selected_year, selected_month, last_day)

    previous_month = selected_month - 1
    previous_year = selected_year

    if previous_month == 0:
        previous_month = 12
        previous_year -= 1

    next_month = selected_month + 1
    next_year = selected_year

    if next_month == 13:
        next_month = 1
        next_year += 1

    approved_statuses = [
        "MANAGER_APPROVED",
        "HR_APPROVED",
        "RELEASED",
    ]

    records = EmployeeAttendance.objects.filter(
        employee=user,
        attendance_date__range=[month_start, month_end],
        status__in=approved_statuses,
    ).order_by("attendance_date")

    records_by_day = {
        record.attendance_date: record
        for record in records
    }

    calendar_obj = py_calendar.Calendar(firstweekday=6)
    raw_weeks = calendar_obj.monthdatescalendar(selected_year, selected_month)

    calendar_weeks = []

    for week in raw_weeks:
        week_days = []

        for day_date in week:
            record = records_by_day.get(day_date)
            is_current_month = day_date.month == selected_month

            is_modified = False
            dot_class = ""
            status_text = ""

            if record:
                system_km = record.system_distance_km or Decimal("0.00")
                approved_km = record.manager_approved_km or Decimal("0.00")

                if approved_km != system_km:
                    is_modified = True
                    dot_class = "orange"
                    status_text = "Manager Modified"
                else:
                    dot_class = "green"
                    status_text = "Approved"

            week_days.append({
                "date": day_date,
                "day": day_date.day,
                "is_current_month": is_current_month,
                "record": record,
                "has_record": True if record else False,
                "is_modified": is_modified,
                "dot_class": dot_class,
                "status_text": status_text,
            })

        calendar_weeks.append(week_days)

    month_summary = records.aggregate(
        total_system_km=Sum("system_distance_km"),
        total_approved_km=Sum("manager_approved_km"),
        total_ta=Sum("ta_amount"),
        total_da=Sum("da_amount"),
        total_claim=Sum("total_claim_amount"),
        total_dealer_visits=Sum("dealers_connected_count"),
    )

    total_system_km = month_summary["total_system_km"] or Decimal("0.00")
    total_approved_km = month_summary["total_approved_km"] or Decimal("0.00")
    total_ta = month_summary["total_ta"] or Decimal("0.00")
    total_da = month_summary["total_da"] or Decimal("0.00")
    total_claim = month_summary["total_claim"] or Decimal("0.00")
    total_dealer_visits = month_summary["total_dealer_visits"] or 0

    released_amount = records.filter(
        status="RELEASED"
    ).aggregate(
        total=Sum("total_claim_amount")
    )["total"] or Decimal("0.00")

    pending_release_amount = total_claim - released_amount

    modified_count = 0
    green_count = 0

    for record in records:
        system_km = record.system_distance_km or Decimal("0.00")
        approved_km = record.manager_approved_km or Decimal("0.00")

        if approved_km != system_km:
            modified_count += 1
        else:
            green_count += 1

    pending_manager_count = EmployeeAttendance.objects.filter(
        employee=user,
        status="PENDING_MANAGER"
    ).count()

    pending_hr_count = EmployeeAttendance.objects.filter(
        employee=user,
        status="MANAGER_APPROVED"
    ).count()

    pending_accountant_count = EmployeeAttendance.objects.filter(
        employee=user,
        status="HR_APPROVED"
    ).count()

    return render(
        request,
        "core/employee_tada_calendar.html",
        {
            "profile": profile,

            "selected_month": selected_month,
            "selected_year": selected_year,
            "month_name": py_calendar.month_name[selected_month],

            "previous_month": previous_month,
            "previous_year": previous_year,
            "next_month": next_month,
            "next_year": next_year,

            "calendar_weeks": calendar_weeks,
            "records": records,

            "total_system_km": total_system_km,
            "total_approved_km": total_approved_km,
            "total_ta": total_ta,
            "total_da": total_da,
            "total_claim": total_claim,
            "released_amount": released_amount,
            "pending_release_amount": pending_release_amount,
            "total_dealer_visits": total_dealer_visits,

            "green_count": green_count,
            "modified_count": modified_count,

            "pending_manager_count": pending_manager_count,
            "pending_hr_count": pending_hr_count,
            "pending_accountant_count": pending_accountant_count,

            "today": today,
        }
    )



def get_weekday_work_field(work_date):
    weekday = work_date.weekday()

    fields = {
        0: "work_monday",
        1: "work_tuesday",
        2: "work_wednesday",
        3: "work_thursday",
        4: "work_friday",
        5: "work_saturday",
        6: "work_sunday",
    }

    return fields.get(weekday)


def is_employee_regular_working_day(user, work_date):
    profile = getattr(user, "profile", None)

    if not profile:
        return False

    field_name = get_weekday_work_field(work_date)

    if not field_name:
        return False

    return bool(getattr(profile, field_name, False))


def has_extra_workday_manager_approval(user, work_date):
    return EmployeeExtraWorkDayRequest.objects.filter(
        employee=user,
        work_date=work_date,
        status__in=["MANAGER_APPROVED_PENDING_HR", "HR_APPROVED"]
    ).exists()


def can_employee_clockin_today(user, work_date):
    if is_employee_regular_working_day(user, work_date):
        return True

    if has_extra_workday_manager_approval(user, work_date):
        return True

    return False


def get_employee_manager_user(employee_user):
    profile = getattr(employee_user, "profile", None)

    if profile and profile.manager:
        return profile.manager

    return None


def is_manager_of_employee(manager_user, employee_user):
    employee_profile = getattr(employee_user, "profile", None)

    if not employee_profile:
        return False

    return employee_profile.manager_id == manager_user.id


def can_hr_admin(user):
    profile = getattr(user, "profile", None)
    return user.is_superuser or (profile and profile.role in ["ADMIN", "HR"])


@login_required
def employee_extra_workday_list(request):
    requests_qs = EmployeeExtraWorkDayRequest.objects.select_related(
        "employee",
        "manager_approved_by",
        "hr_approved_by",
    ).filter(
        employee=request.user
    )

    return render(
        request,
        "core/employee_extra_workday_list.html",
        {
            "requests_qs": requests_qs,
        }
    )


@login_required
def employee_extra_workday_create(request):
    profile = getattr(request.user, "profile", None)

    if not profile or profile.role == "DEALER":
        messages.error(request, "Only employees can request extra working day.")
        return redirect("dashboard")

    if request.method == "POST":
        form = EmployeeExtraWorkDayRequestForm(request.POST)

        if form.is_valid():
            work_date = form.cleaned_data["work_date"]

            if is_employee_regular_working_day(request.user, work_date):
                messages.error(
                    request,
                    "Selected date is already your regular working day. Extra request is not required."
                )
                return redirect("employee_extra_workday_create")

            manager = get_employee_manager_user(request.user)

            if not manager:
                messages.error(
                    request,
                    "Manager is not assigned for your profile. Please contact Admin/HR."
                )
                return redirect("employee_extra_workday_create")

            request_obj = form.save(commit=False)
            request_obj.employee = request.user
            request_obj.status = "PENDING_MANAGER"
            request_obj.save()

            messages.success(
                request,
                "Extra working day request sent to your concerned manager."
            )
            return redirect("employee_extra_workday_list")
    else:
        form = EmployeeExtraWorkDayRequestForm()

    return render(
        request,
        "core/employee_extra_workday_form.html",
        {
            "form": form,
            "title": "Request Extra Working Day",
        }
    )


@login_required
def manager_extra_workday_requests(request):
    subordinate_ids = UserProfile.objects.filter(
        manager=request.user
    ).values_list("user_id", flat=True)

    requests_qs = EmployeeExtraWorkDayRequest.objects.select_related(
        "employee",
        "employee__profile",
    ).filter(
        employee_id__in=subordinate_ids,
        status="PENDING_MANAGER"
    )

    return render(
        request,
        "core/manager_extra_workday_requests.html",
        {
            "requests_qs": requests_qs,
        }
    )


@login_required
def manager_approve_extra_workday(request, request_id):
    extra_request = get_object_or_404(
        EmployeeExtraWorkDayRequest,
        id=request_id,
        status="PENDING_MANAGER"
    )

    if not is_manager_of_employee(request.user, extra_request.employee):
        messages.error(request, "You are not allowed to approve this request.")
        return redirect("manager_extra_workday_requests")

    if request.method == "POST":
        form = ExtraWorkDayManagerApprovalForm(request.POST, instance=extra_request)

        if form.is_valid():
            extra_request = form.save(commit=False)
            extra_request.status = "MANAGER_APPROVED_PENDING_HR"
            extra_request.manager_approved_by = request.user
            extra_request.manager_approved_at = timezone.now()
            extra_request.save()

            messages.success(
                request,
                "Extra working day approved. Employee can clock-in on that day. Request sent to HR for extra leave addition."
            )
            return redirect("manager_extra_workday_requests")
    else:
        form = ExtraWorkDayManagerApprovalForm(instance=extra_request)

    return render(
        request,
        "core/manager_extra_workday_approve.html",
        {
            "form": form,
            "extra_request": extra_request,
        }
    )


@login_required
def manager_reject_extra_workday(request, request_id):
    extra_request = get_object_or_404(
        EmployeeExtraWorkDayRequest,
        id=request_id,
        status="PENDING_MANAGER"
    )

    if not is_manager_of_employee(request.user, extra_request.employee):
        messages.error(request, "You are not allowed to reject this request.")
        return redirect("manager_extra_workday_requests")

    if request.method == "POST":
        extra_request.status = "REJECTED"
        extra_request.manager_approved_by = request.user
        extra_request.manager_approved_at = timezone.now()
        extra_request.manager_remarks = request.POST.get("manager_remarks", "")
        extra_request.save()

        messages.success(request, "Extra working day request rejected.")
        return redirect("manager_extra_workday_requests")

    return redirect("manager_extra_workday_requests")


@login_required
def hr_extra_workday_requests(request):
    if not can_hr_admin(request.user):
        messages.error(request, "Only HR/Admin can view this section.")
        return redirect("dashboard")

    requests_qs = EmployeeExtraWorkDayRequest.objects.select_related(
        "employee",
        "employee__profile",
        "manager_approved_by",
    ).filter(
        status="MANAGER_APPROVED_PENDING_HR"
    )

    return render(
        request,
        "core/hr_extra_workday_requests.html",
        {
            "requests_qs": requests_qs,
        }
    )


@login_required
def hr_approve_extra_workday(request, request_id):
    if not can_hr_admin(request.user):
        messages.error(request, "Only HR/Admin can approve extra leave.")
        return redirect("dashboard")

    extra_request = get_object_or_404(
        EmployeeExtraWorkDayRequest,
        id=request_id,
        status="MANAGER_APPROVED_PENDING_HR"
    )

    if request.method == "POST":
        form = ExtraWorkDayHRApprovalForm(request.POST, instance=extra_request)

        if form.is_valid():
            extra_request = form.save(commit=False)
            extra_request.status = "HR_APPROVED"
            extra_request.hr_approved_by = request.user
            extra_request.hr_approved_at = timezone.now()

            if not extra_request.extra_leave_added:
                profile = extra_request.employee.profile
                profile.earned_extra_other_leaves += 1
                profile.save()

                extra_request.extra_leave_added = True

            extra_request.save()

            messages.success(
                request,
                "HR approved. One extra other leave added to employee leave balance."
            )
            return redirect("hr_extra_workday_requests")
    else:
        form = ExtraWorkDayHRApprovalForm(instance=extra_request)

    return render(
        request,
        "core/hr_extra_workday_approve.html",
        {
            "form": form,
            "extra_request": extra_request,
        }
    )


@login_required
def employee_leave_balance(request):
    profile = getattr(request.user, "profile", None)

    if not profile or profile.role == "DEALER":
        messages.error(request, "Leave balance is available only for employees.")
        return redirect("dashboard")

    today = timezone.localdate()

    try:
        selected_year = int(request.GET.get("year", today.year))
    except Exception:
        selected_year = today.year

    year_start = timezone.datetime(selected_year, 1, 1).date()
    year_end = timezone.datetime(selected_year, 12, 31).date()

    approved_leaves = EmployeeLeaveRequest.objects.filter(
        employee=request.user,
        status="APPROVED",
        from_date__lte=year_end,
        to_date__gte=year_start,
    )

    used_paid = Decimal("0.00")
    used_sick = Decimal("0.00")
    used_other = Decimal("0.00")
    used_unpaid = Decimal("0.00")

    for leave in approved_leaves:
        days = leave.calculate_days()

        if leave.leave_type == "PAID":
            used_paid += days
        elif leave.leave_type == "SICK":
            used_sick += days
        elif leave.leave_type == "OTHER":
            used_other += days
        elif leave.leave_type == "UNPAID":
            used_unpaid += days

    paid_total = Decimal(str(profile.paid_leaves_per_year or 0))
    sick_total = Decimal(str(profile.sick_leaves_per_year or 0))
    other_base = Decimal(str(profile.other_leaves_per_year or 0))
    earned_extra = Decimal(str(profile.earned_extra_other_leaves or 0))
    other_total = other_base + earned_extra

    paid_left = max(paid_total - used_paid, Decimal("0.00"))
    sick_left = max(sick_total - used_sick, Decimal("0.00"))
    other_left = max(other_total - used_other, Decimal("0.00"))

    pending_leaves = EmployeeLeaveRequest.objects.filter(
        employee=request.user,
        status="PENDING",
    ).count()

    chart_labels = [
        "Paid Used",
        "Paid Left",
        "Sick Used",
        "Sick Left",
        "Other Used",
        "Other Left",
        "Unpaid Used",
    ]

    chart_values = [
        float(used_paid),
        float(paid_left),
        float(used_sick),
        float(sick_left),
        float(used_other),
        float(other_left),
        float(used_unpaid),
    ]

    return render(
        request,
        "core/employee_leave_balance.html",
        {
            "selected_year": selected_year,

            "paid_total": paid_total,
            "paid_used": used_paid,
            "paid_left": paid_left,

            "sick_total": sick_total,
            "sick_used": used_sick,
            "sick_left": sick_left,

            "other_base": other_base,
            "earned_extra": earned_extra,
            "other_total": other_total,
            "other_used": used_other,
            "other_left": other_left,

            "unpaid_used": used_unpaid,
            "pending_leaves": pending_leaves,

            "chart_labels": chart_labels,
            "chart_values": chart_values,
        }
    )


def is_asset_manager_user(user):
    profile = getattr(user, "profile", None)

    if user.is_superuser:
        return True

    if not profile:
        return False

    return profile.role in ["ADMIN", "ASSET_MANAGER"]


def generate_asset_code(category):
    prefix = category.code_prefix.upper().strip()

    last_asset = AssetItem.objects.filter(
        asset_code__startswith=f"{prefix}-"
    ).order_by("-id").first()

    if last_asset:
        try:
            last_number = int(last_asset.asset_code.split("-")[-1])
        except Exception:
            last_number = AssetItem.objects.filter(
                asset_code__startswith=f"{prefix}-"
            ).count()
    else:
        last_number = 0

    next_number = last_number + 1

    while True:
        asset_code = f"{prefix}-{next_number:04d}"

        if not AssetItem.objects.filter(asset_code=asset_code).exists():
            return asset_code

        next_number += 1


def save_camera_image_from_base64(image_data):
    if not image_data:
        return None

    if ";base64," not in image_data:
        return None

    format_part, image_string = image_data.split(";base64,")
    ext = format_part.split("/")[-1]

    file_name = f"asset_handover_{uuid.uuid4().hex}.{ext}"

    return ContentFile(
        base64.b64decode(image_string),
        name=file_name
    )


@login_required
def asset_list(request):
    if not is_asset_manager_user(request.user):
        messages.error(request, "Only Admin or Asset Manager can view assets.")
        return redirect("dashboard")

    search = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    category_id = request.GET.get("category", "").strip()

    assignment_qs = AssetAssignment.objects.select_related(
        "employee",
        "employee__profile",
        "assigned_by",
        "returned_by",
    ).order_by("-assigned_at")

    assets = AssetItem.objects.select_related(
        "category",
        "current_employee",
        "current_employee__profile",
    ).prefetch_related(
        Prefetch("assignments", queryset=assignment_qs)
    ).all()

    if search:
        assets = assets.filter(
            Q(asset_code__icontains=search) |
            Q(name__icontains=search) |
            Q(brand__icontains=search) |
            Q(model_number__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(current_employee__username__icontains=search) |
            Q(current_employee__first_name__icontains=search) |
            Q(current_employee__last_name__icontains=search)
        )

    if status:
        assets = assets.filter(status=status)

    if category_id:
        assets = assets.filter(category_id=category_id)

    asset_rows = []

    for asset in assets:
        history = list(asset.assignments.all())

        active_assignment = None
        last_returned_assignment = None

        for assignment in history:
            if assignment.is_active and active_assignment is None:
                active_assignment = assignment

            if not assignment.is_active and last_returned_assignment is None:
                last_returned_assignment = assignment

        asset_rows.append({
            "asset": asset,
            "active_assignment": active_assignment,
            "last_returned_assignment": last_returned_assignment,
            "history_count": len(history),
        })

    categories = AssetCategory.objects.filter(is_active=True)

    return render(request, "core/asset_list.html", {
        "asset_rows": asset_rows,
        "categories": categories,
        "search": search,
        "status": status,
        "category_id": category_id,
    })

@login_required
def asset_detail(request, asset_id):
    if not is_asset_manager_user(request.user):
        messages.error(request, "Only Admin or Asset Manager can view asset history.")
        return redirect("dashboard")

    asset = get_object_or_404(
        AssetItem.objects.select_related(
            "category",
            "current_employee",
            "current_employee__profile",
            "created_by",
        ),
        id=asset_id
    )

    assignments = AssetAssignment.objects.filter(
        asset=asset
    ).select_related(
        "employee",
        "employee__profile",
        "assigned_by",
        "returned_by",
    ).order_by("-assigned_at")

    active_assignment = assignments.filter(is_active=True).first()
    past_assignments = assignments.filter(is_active=False)

    return render(request, "core/asset_detail.html", {
        "asset": asset,
        "assignments": assignments,
        "active_assignment": active_assignment,
        "past_assignments": past_assignments,
    })



@login_required
def asset_create_bulk(request):
    if not is_asset_manager_user(request.user):
        messages.error(request, "Only Admin or Asset Manager can create assets.")
        return redirect("dashboard")

    categories = AssetCategory.objects.filter(is_active=True)

    if request.method == "POST":
        category_id = request.POST.get("category")
        new_category_name = request.POST.get("new_category_name", "").strip()
        new_category_prefix = request.POST.get("new_category_prefix", "").strip()

        name = request.POST.get("name", "").strip()
        brand = request.POST.get("brand", "").strip()
        model_number = request.POST.get("model_number", "").strip()
        serial_number = request.POST.get("serial_number", "").strip()
        description = request.POST.get("description", "").strip()
        purchase_date = request.POST.get("purchase_date") or None
        purchase_value = request.POST.get("purchase_value") or None

        try:
            quantity = int(request.POST.get("quantity") or 1)
        except ValueError:
            quantity = 1

        if quantity < 1:
            quantity = 1

        if quantity > 500:
            quantity = 500

        if not name:
            messages.error(request, "Please enter asset name.")
            return redirect("asset_create_bulk")

        if new_category_name and new_category_prefix:
            category, created = AssetCategory.objects.get_or_create(
                code_prefix=new_category_prefix.upper().strip(),
                defaults={
                    "name": new_category_name,
                    "description": "",
                    "is_active": True,
                }
            )
        else:
            category = AssetCategory.objects.filter(id=category_id).first()

        if not category:
            messages.error(request, "Please select or create asset category.")
            return redirect("asset_create_bulk")

        created_count = 0

        for i in range(quantity):
            asset_code = generate_asset_code(category)

            AssetItem.objects.create(
                category=category,
                asset_code=asset_code,
                name=name,
                brand=brand,
                model_number=model_number,
                serial_number=serial_number if quantity == 1 else "",
                description=description,
                purchase_date=purchase_date,
                purchase_value=purchase_value,
                status="AVAILABLE",
                created_by=request.user,
            )

            created_count += 1

        messages.success(request, f"{created_count} asset(s) created successfully.")
        return redirect("asset_list")

    return render(request, "core/asset_create_bulk.html", {
        "categories": categories,
    })




@login_required
def asset_assign(request, asset_id):
    if not is_asset_manager_user(request.user):
        messages.error(request, "Only Admin or Asset Manager can assign assets.")
        return redirect("dashboard")

    asset = get_object_or_404(
        AssetItem.objects.select_related("category", "current_employee"),
        id=asset_id
    )

    employees = User.objects.filter(
        is_active=True
    ).exclude(
        profile__role="DEALER"
    ).select_related("profile").order_by(
        "first_name",
        "last_name",
        "username"
    )

    if request.method == "POST":
        employee_id = request.POST.get("employee")
        handover_note = request.POST.get("handover_note", "").strip()
        camera_image_data = request.POST.get("camera_image_data")

        employee = User.objects.filter(id=employee_id, is_active=True).first()

        if not employee:
            messages.error(request, "Please select employee.")
            return redirect("asset_assign", asset_id=asset.id)

        if asset.status == "ASSIGNED":
            messages.error(request, "This asset is already assigned.")
            return redirect("asset_list")

        image_file = save_camera_image_from_base64(camera_image_data)

        if not image_file:
            messages.error(request, "Please capture asset handover photo using camera.")
            return redirect("asset_assign", asset_id=asset.id)

        AssetAssignment.objects.create(
            asset=asset,
            employee=employee,
            assigned_by=request.user,
            handover_image=image_file,
            handover_note=handover_note,
            is_active=True,
        )

        asset.current_employee = employee
        asset.status = "ASSIGNED"
        asset.save(update_fields=["current_employee", "status", "updated_at"])

        messages.success(request, f"{asset.asset_code} assigned successfully.")
        return redirect("asset_list")

    return render(request, "core/asset_assign.html", {
        "asset": asset,
        "employees": employees,
    })




@login_required
def my_assets(request):
    assignments = AssetAssignment.objects.filter(
        employee=request.user,
        is_active=True
    ).select_related(
        "asset",
        "asset__category",
        "assigned_by"
    ).order_by("-assigned_at")

    return render(request, "core/my_assets.html", {
        "assignments": assignments,
    })


@login_required
def employee_assets_detail(request, employee_id):
    if not is_asset_manager_user(request.user):
        messages.error(request, "Only Admin or Asset Manager can view employee assets.")
        return redirect("dashboard")

    employee = get_object_or_404(
        User.objects.select_related("profile"),
        id=employee_id,
        is_active=True
    )

    active_assignments = AssetAssignment.objects.filter(
        employee=employee,
        is_active=True
    ).select_related(
        "asset",
        "asset__category",
        "assigned_by"
    ).order_by("-assigned_at")

    returned_assignments = AssetAssignment.objects.filter(
        employee=employee,
        is_active=False
    ).select_related(
        "asset",
        "asset__category",
        "assigned_by",
        "returned_by"
    ).order_by("-returned_at")

    return render(request, "core/employee_assets_detail.html", {
        "employee": employee,
        "active_assignments": active_assignments,
        "returned_assignments": returned_assignments,
    })



@login_required
def asset_return(request, assignment_id):
    if not is_asset_manager_user(request.user):
        messages.error(request, "Only Admin or Asset Manager can return assets.")
        return redirect("dashboard")

    assignment = get_object_or_404(
        AssetAssignment.objects.select_related("asset", "employee"),
        id=assignment_id,
        is_active=True
    )

    if request.method == "POST":
        return_note = request.POST.get("return_note", "").strip()

        assignment.is_active = False
        assignment.returned_at = timezone.now()
        assignment.returned_by = request.user
        assignment.return_note = return_note
        assignment.save()

        asset = assignment.asset
        asset.status = "AVAILABLE"
        asset.current_employee = None
        asset.save(update_fields=["status", "current_employee", "updated_at"])

        messages.success(
            request,
            f"{asset.asset_code} returned successfully and is now available."
        )

        return redirect("asset_detail", asset_id=asset.id)

    return render(request, "core/asset_return.html", {
        "assignment": assignment,
    })
    
    
def get_user_role(user):
    profile = getattr(user, "profile", None)
    return profile.role if profile else ""


def get_logged_in_dealer(user):
    return Dealer.objects.filter(user=user).first()


def dealer_point_balance(dealer):
    credit = DealerPointLedger.objects.filter(
        dealer=dealer,
        transaction_type="CREDIT"
    ).aggregate(total=Sum("points"))["total"] or 0

    debit = DealerPointLedger.objects.filter(
        dealer=dealer,
        transaction_type="DEBIT"
    ).aggregate(total=Sum("points"))["total"] or 0

    return credit - debit


def can_view_redemption_request(user, redemption):
    role = get_user_role(user)

    if user.is_superuser or role == "ADMIN":
        return True

    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        return redemption.status == "PENDING_SALES_OFFICER" and redemption.sales_officer_id == user.id

    if role == "ASM":
        return redemption.status == "PENDING_ASM" and redemption.asm_id == user.id

    if role == "ACCOUNTANT":
        return redemption.status == "PENDING_ACCOUNTANT"

    return False


@login_required
def dealer_farmer_data(request):
    role = get_user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealer can upload farmer data.")
        return redirect("dashboard")

    dealer = get_logged_in_dealer(request.user)

    if not dealer:
        messages.error(request, "Dealer profile not found.")
        return redirect("dashboard")

    setting = DealerPointSetting.get_settings()

    if request.method == "POST":
        farmer_name = request.POST.get("farmer_name", "").strip()
        mobile_number = request.POST.get("mobile_number", "").strip()
        place = request.POST.get("place", "").strip()

        if not farmer_name or not mobile_number or not place:
            messages.error(request, "Farmer name, mobile number and place are required.")
            return redirect("dealer_farmer_data")

        try:
            mobile_number = normalize_indian_mobile(mobile_number)
        except ValidationError as e:
            messages.error(request, e.messages[0])
            return redirect("dealer_farmer_data")

        if FarmerData.objects.filter(mobile_number=mobile_number).exists():
            messages.error(
                request,
                "This farmer mobile number already exists. Duplicate farmer data is not allowed."
            )
            return redirect("dealer_farmer_data")

        try:
            with transaction.atomic():
                farmer = FarmerData.objects.create(
                    dealer=dealer,
                    farmer_name=farmer_name,
                    mobile_number=mobile_number,
                    place=place,
                    points_awarded=setting.farmer_points,
                    created_by=request.user,
                )

                DealerPointLedger.objects.create(
                    dealer=dealer,
                    transaction_type="CREDIT",
                    points=setting.farmer_points,
                    farmer_data=farmer,
                    note=f"Farmer data uploaded: {farmer.farmer_name}"
                )

            messages.success(
                request,
                f"Farmer data uploaded successfully. {setting.farmer_points} points added."
            )
            return redirect("dealer_farmer_data")

        except IntegrityError:
            messages.error(request, "This mobile number already exists.")
            return redirect("dealer_farmer_data")

    farmers = FarmerData.objects.filter(dealer=dealer).order_by("-created_at")
    balance_points = dealer_point_balance(dealer)
    point_value = Decimal(balance_points) * setting.rupees_per_point

    return render(request, "core/dealer_farmer_data.html", {
        "dealer": dealer,
        "farmers": farmers,
        "setting": setting,
        "balance_points": balance_points,
        "point_value": point_value,
    })


@login_required
def dealer_points_dashboard(request):
    role = get_user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealer can view points dashboard.")
        return redirect("dashboard")

    dealer = get_logged_in_dealer(request.user)

    if not dealer:
        messages.error(request, "Dealer profile not found.")
        return redirect("dashboard")

    setting = DealerPointSetting.get_settings()
    balance_points = dealer_point_balance(dealer)
    point_value = Decimal(balance_points) * setting.rupees_per_point

    farmers_count = FarmerData.objects.filter(dealer=dealer).count()

    product_rules = DealerPointProductRule.objects.filter(
        is_active=True
    ).select_related(
        "product",
        "pack_size"
    ).order_by("points_required")

    redemptions = DealerPointRedemptionRequest.objects.filter(
        dealer=dealer
    ).select_related(
        "product_rule",
        "product_rule__product",
        "sales_officer",
        "asm",
        "accountant",
    ).order_by("-created_at")

    return render(request, "core/dealer_points_dashboard.html", {
        "dealer": dealer,
        "setting": setting,
        "balance_points": balance_points,
        "point_value": point_value,
        "farmers_count": farmers_count,
        "product_rules": product_rules,
        "redemptions": redemptions,
    })


@login_required
def dealer_redemption_create(request):
    role = get_user_role(request.user)

    if role != "DEALER":
        messages.error(request, "Only dealer can create redemption request.")
        return redirect("dashboard")

    dealer = get_logged_in_dealer(request.user)

    if not dealer:
        messages.error(request, "Dealer profile not found.")
        return redirect("dashboard")

    setting = DealerPointSetting.get_settings()
    balance_points = dealer_point_balance(dealer)

    product_rules = DealerPointProductRule.objects.filter(
        is_active=True
    ).select_related("product", "pack_size")

    if request.method == "POST":
        redemption_type = request.POST.get("redemption_type")
        note = request.POST.get("note", "").strip()

        sales_officer = dealer.created_by_sales_officer
        asm = dealer.concerned_asm

        if not sales_officer:
            messages.error(request, "Concerned Sales Officer is not assigned to this dealer.")
            return redirect("dealer_redemption_create")

        if not asm:
            messages.error(request, "Concerned ASM is not assigned to this dealer.")
            return redirect("dealer_redemption_create")

        if redemption_type == "MONEY":
            try:
                points_requested = int(request.POST.get("points_requested") or 0)
            except ValueError:
                points_requested = 0

            if points_requested < setting.minimum_money_redemption_points:
                messages.error(
                    request,
                    f"Minimum money redemption is {setting.minimum_money_redemption_points} points."
                )
                return redirect("dealer_redemption_create")

            if points_requested > balance_points:
                messages.error(request, "You do not have enough points.")
                return redirect("dealer_redemption_create")

            money_amount = Decimal(points_requested) * setting.rupees_per_point

            DealerPointRedemptionRequest.objects.create(
                dealer=dealer,
                requested_by=request.user,
                redemption_type="MONEY",
                points_requested=points_requested,
                money_amount=money_amount,
                sales_officer=sales_officer,
                asm=asm,
                status="PENDING_SALES_OFFICER",
                rejection_reason=note,
            )

            messages.success(request, "Money redemption request sent to Sales Officer.")
            return redirect("dealer_points_dashboard")

        if redemption_type == "PRODUCT":
            rule_id = request.POST.get("product_rule")

            try:
                request_quantity = int(request.POST.get("request_quantity") or 1)
            except ValueError:
                request_quantity = 1

            if request_quantity < 1:
                request_quantity = 1

            rule = DealerPointProductRule.objects.filter(id=rule_id, is_active=True).first()

            if not rule:
                messages.error(request, "Please select valid product redemption option.")
                return redirect("dealer_redemption_create")

            points_requested = rule.points_required * request_quantity
            requested_free_quantity = rule.free_quantity * request_quantity

            if points_requested < setting.minimum_product_redemption_points:
                messages.error(
                    request,
                    f"Minimum product redemption is {setting.minimum_product_redemption_points} points."
                )
                return redirect("dealer_redemption_create")

            if points_requested > balance_points:
                messages.error(request, "You do not have enough points.")
                return redirect("dealer_redemption_create")

            DealerPointRedemptionRequest.objects.create(
                dealer=dealer,
                requested_by=request.user,
                redemption_type="PRODUCT",
                points_requested=points_requested,
                product_rule=rule,
                request_quantity=request_quantity,
                requested_free_quantity=requested_free_quantity,
                sales_officer=sales_officer,
                asm=asm,
                status="PENDING_SALES_OFFICER",
                rejection_reason=note,
            )

            messages.success(request, "Product redemption request sent to Sales Officer.")
            return redirect("dealer_points_dashboard")

        messages.error(request, "Invalid redemption type.")
        return redirect("dealer_redemption_create")

    return render(request, "core/dealer_redemption_create.html", {
        "dealer": dealer,
        "setting": setting,
        "balance_points": balance_points,
        "product_rules": product_rules,
    })


@login_required
def dealer_redemption_approvals(request):
    role = get_user_role(request.user)

    requests = DealerPointRedemptionRequest.objects.select_related(
        "dealer",
        "product_rule",
        "product_rule__product",
        "sales_officer",
        "asm",
        "accountant",
    )

    if request.user.is_superuser or role == "ADMIN":
        requests = requests.all()

    elif role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        requests = requests.filter(
            status="PENDING_SALES_OFFICER",
            sales_officer=request.user
        )

    elif role == "ASM":
        requests = requests.filter(
            status="PENDING_ASM",
            asm=request.user
        )

    elif role == "ACCOUNTANT":
        requests = requests.filter(
            status="PENDING_ACCOUNTANT"
        )

    else:
        messages.error(request, "You do not have permission to view redemption approvals.")
        return redirect("dashboard")

    return render(request, "core/dealer_redemption_approvals.html", {
        "requests": requests.order_by("-created_at"),
    })


@login_required
def dealer_redemption_approve(request, request_id):
    redemption = get_object_or_404(
        DealerPointRedemptionRequest.objects.select_related(
            "dealer",
            "sales_officer",
            "asm",
            "product_rule",
            "product_rule__product",
        ),
        id=request_id
    )

    if not can_view_redemption_request(request.user, redemption):
        messages.error(request, "You do not have permission to approve this request.")
        return redirect("dashboard")

    with transaction.atomic():
        redemption = DealerPointRedemptionRequest.objects.select_for_update().get(id=request_id)

        if redemption.status == "PENDING_SALES_OFFICER":
            redemption.status = "PENDING_ASM"
            redemption.sales_officer_approved_at = timezone.now()
            redemption.save(update_fields=["status", "sales_officer_approved_at"])
            messages.success(request, "Request approved and sent to ASM.")
            return redirect("dealer_redemption_approvals")

        if redemption.status == "PENDING_ASM":
            redemption.status = "PENDING_ACCOUNTANT"
            redemption.asm_approved_at = timezone.now()
            redemption.save(update_fields=["status", "asm_approved_at"])
            messages.success(request, "Request approved and sent to Accountant.")
            return redirect("dealer_redemption_approvals")

        if redemption.status == "PENDING_ACCOUNTANT":
            current_balance = dealer_point_balance(redemption.dealer)

            if redemption.points_requested > current_balance:
                messages.error(request, "Dealer does not have enough points now.")
                return redirect("dealer_redemption_approvals")

            redemption.status = "APPROVED"
            redemption.accountant = request.user
            redemption.accountant_approved_at = timezone.now()
            redemption.save(update_fields=["status", "accountant", "accountant_approved_at"])

            DealerPointLedger.objects.create(
                dealer=redemption.dealer,
                transaction_type="DEBIT",
                points=redemption.points_requested,
                redemption_request=redemption,
                note=f"Redemption approved: {redemption.get_redemption_type_display()}"
            )

            messages.success(request, "Redemption approved. Points reduced from dealer account.")
            return redirect("dealer_redemption_approvals")

    messages.error(request, "Invalid request status.")
    return redirect("dealer_redemption_approvals")


@login_required
def dealer_redemption_reject(request, request_id):
    redemption = get_object_or_404(DealerPointRedemptionRequest, id=request_id)

    if not can_view_redemption_request(request.user, redemption):
        messages.error(request, "You do not have permission to reject this request.")
        return redirect("dashboard")

    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()

        redemption.status = "REJECTED"
        redemption.rejected_by = request.user
        redemption.rejected_at = timezone.now()
        redemption.rejection_reason = reason
        redemption.save(update_fields=["status", "rejected_by", "rejected_at", "rejection_reason"])

        messages.success(request, "Redemption request rejected.")
        return redirect("dealer_redemption_approvals")

    return render(request, "core/dealer_redemption_reject.html", {
        "redemption": redemption,
    })


@login_required
def dealer_point_admin_settings(request):
    profile = getattr(request.user, "profile", None)

    if not (request.user.is_superuser or (profile and profile.role == "ADMIN")):
        messages.error(request, "Only Admin can manage dealer point settings.")
        return redirect("dashboard")

    setting = DealerPointSetting.get_settings()

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "save_settings":
            try:
                setting.farmer_points = int(request.POST.get("farmer_points") or 5)
                setting.rupees_per_point = Decimal(request.POST.get("rupees_per_point") or "1")
                setting.minimum_money_redemption_points = int(request.POST.get("minimum_money_redemption_points") or 100)
                setting.minimum_product_redemption_points = int(request.POST.get("minimum_product_redemption_points") or 100)
                setting.is_active = True
                setting.save()

                messages.success(request, "Dealer point settings updated successfully.")
                return redirect("dealer_point_admin_settings")

            except Exception:
                messages.error(request, "Invalid point setting values.")
                return redirect("dealer_point_admin_settings")

        if action == "add_product_rule":
            product_id = request.POST.get("product")
            pack_size_id = request.POST.get("pack_size") or None

            try:
                points_required = int(request.POST.get("points_required") or 0)
                free_quantity = int(request.POST.get("free_quantity") or 1)
            except ValueError:
                messages.error(request, "Invalid product redemption values.")
                return redirect("dealer_point_admin_settings")

            product = Product.objects.filter(id=product_id).first()

            if not product:
                messages.error(request, "Please select product.")
                return redirect("dealer_point_admin_settings")

            pack_size = None
            if pack_size_id:
                pack_size = ProductPackSize.objects.filter(id=pack_size_id).first()

            if points_required <= 0:
                messages.error(request, "Points required must be greater than zero.")
                return redirect("dealer_point_admin_settings")

            DealerPointProductRule.objects.create(
                product=product,
                pack_size=pack_size,
                points_required=points_required,
                free_quantity=free_quantity,
                is_active=True,
            )

            messages.success(request, "Product redemption rule added successfully.")
            return redirect("dealer_point_admin_settings")

        if action == "toggle_product_rule":
            rule_id = request.POST.get("rule_id")
            rule = DealerPointProductRule.objects.filter(id=rule_id).first()

            if rule:
                rule.is_active = not rule.is_active
                rule.save(update_fields=["is_active"])
                messages.success(request, "Product redemption rule status updated.")

            return redirect("dealer_point_admin_settings")

    products = Product.objects.all().order_by("name")
    pack_sizes = ProductPackSize.objects.select_related("product").all().order_by("product__name", "id")

    product_rules = DealerPointProductRule.objects.select_related(
        "product",
        "pack_size"
    ).order_by("product__name", "points_required")

    return render(request, "core/dealer_point_admin_settings.html", {
        "setting": setting,
        "products": products,
        "pack_sizes": pack_sizes,
        "product_rules": product_rules,
    })


def can_view_all_farmer_data(user):
    profile = getattr(user, "profile", None)
    role = getattr(profile, "role", "")

    return user.is_superuser or role in ["ADMIN", "ACCOUNTANT"]


def get_filtered_farmer_data_queryset(request):
    q = request.GET.get("q", "").strip()
    dealer_id = request.GET.get("dealer", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()

    farmers = FarmerData.objects.select_related(
        "dealer",
        "created_by",
        "dealer__created_by_sales_officer",
        "dealer__concerned_asm",
    ).order_by("-created_at")

    if q:
        farmers = farmers.filter(
            Q(farmer_name__icontains=q) |
            Q(mobile_number__icontains=q) |
            Q(place__icontains=q) |
            Q(dealer__firm_name__icontains=q) |
            Q(dealer__owner_name__icontains=q)
        )

    if dealer_id:
        farmers = farmers.filter(dealer_id=dealer_id)

    if from_date:
        parsed_from_date = parse_date(from_date)
        if parsed_from_date:
            farmers = farmers.filter(created_at__date__gte=parsed_from_date)

    if to_date:
        parsed_to_date = parse_date(to_date)
        if parsed_to_date:
            farmers = farmers.filter(created_at__date__lte=parsed_to_date)

    filters = {
        "q": q,
        "dealer_id": dealer_id,
        "from_date": from_date,
        "to_date": to_date,
    }

    return farmers, filters


@login_required
def farmer_data_admin_list(request):
    if not can_view_all_farmer_data(request.user):
        messages.error(request, "Only Admin or Accountant can view all farmer data.")
        return redirect("dashboard")

    farmers, filters = get_filtered_farmer_data_queryset(request)

    dealers = Dealer.objects.all().order_by("firm_name")

    total_farmers = farmers.count()
    total_points = farmers.aggregate(total=Sum("points_awarded"))["total"] or 0
    unique_dealers = farmers.values("dealer_id").distinct().count()

    return render(request, "core/farmer_data_admin_list.html", {
        "farmers": farmers,
        "dealers": dealers,
        "filters": filters,
        "total_farmers": total_farmers,
        "total_points": total_points,
        "unique_dealers": unique_dealers,
    })


@login_required
def farmer_data_admin_export(request, export_type):
    if not can_view_all_farmer_data(request.user):
        messages.error(request, "Only Admin or Accountant can download farmer data.")
        return redirect("dashboard")

    farmers, filters = get_filtered_farmer_data_queryset(request)

    if export_type == "excel":
        return export_farmer_data_excel(farmers)

    if export_type == "pdf":
        return export_farmer_data_pdf(farmers)

    messages.error(request, "Invalid export type.")
    return redirect("farmer_data_admin_list")


def export_farmer_data_excel(farmers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Farmer Data"

    headers = [
        "S.No",
        "Dealer Firm",
        "Dealer Owner",
        "Farmer Name",
        "Mobile Number",
        "Place",
        "Points Awarded",
        "Sales Officer",
        "ASM",
        "Uploaded Date",
    ]

    ws.append(headers)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for index, farmer in enumerate(farmers, start=1):
        dealer = farmer.dealer

        sales_officer = "-"
        asm = "-"

        if dealer and dealer.created_by_sales_officer:
            sales_officer = dealer.created_by_sales_officer.get_full_name() or dealer.created_by_sales_officer.username

        if dealer and dealer.concerned_asm:
            asm = dealer.concerned_asm.get_full_name() or dealer.concerned_asm.username

        ws.append([
            index,
            dealer.firm_name if dealer else "-",
            getattr(dealer, "owner_name", "-") if dealer else "-",
            farmer.farmer_name,
            farmer.mobile_number,
            farmer.place,
            farmer.points_awarded,
            sales_officer,
            asm,
            timezone.localtime(farmer.created_at).strftime("%d-%m-%Y %I:%M %p") if farmer.created_at else "-",
        ])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = left_align

    column_widths = {
        1: 8,
        2: 28,
        3: 22,
        4: 24,
        5: 16,
        6: 26,
        7: 16,
        8: 24,
        9: 24,
        10: 22,
    }

    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    today = timezone.localdate().strftime("%d_%m_%Y")
    response["Content-Disposition"] = f'attachment; filename="farmer_data_{today}.xlsx"'

    wb.save(response)
    return response


def export_farmer_data_pdf(farmers):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#0F172A"),
        alignment=1,
        spaceAfter=10,
    )

    small_style = ParagraphStyle(
        "SmallStyle",
        parent=styles["BodyText"],
        fontSize=7,
        leading=9,
        wordWrap="CJK",
    )

    elements = []

    elements.append(Paragraph("Farmer Data Report", title_style))
    elements.append(Paragraph(
        f"Generated on: {timezone.localtime(timezone.now()).strftime('%d-%m-%Y %I:%M %p')}",
        small_style
    ))
    elements.append(Spacer(1, 8))

    data = [[
        "S.No",
        "Dealer",
        "Farmer",
        "Mobile",
        "Place",
        "Points",
        "Sales Officer",
        "ASM",
        "Date",
    ]]

    for index, farmer in enumerate(farmers, start=1):
        dealer = farmer.dealer

        sales_officer = "-"
        asm = "-"

        if dealer and dealer.created_by_sales_officer:
            sales_officer = dealer.created_by_sales_officer.get_full_name() or dealer.created_by_sales_officer.username

        if dealer and dealer.concerned_asm:
            asm = dealer.concerned_asm.get_full_name() or dealer.concerned_asm.username

        data.append([
            str(index),
            Paragraph(escape(dealer.firm_name if dealer else "-"), small_style),
            Paragraph(escape(farmer.farmer_name), small_style),
            farmer.mobile_number,
            Paragraph(escape(farmer.place), small_style),
            str(farmer.points_awarded),
            Paragraph(escape(sales_officer), small_style),
            Paragraph(escape(asm), small_style),
            timezone.localtime(farmer.created_at).strftime("%d-%m-%Y") if farmer.created_at else "-",
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            12 * mm,
            40 * mm,
            34 * mm,
            26 * mm,
            38 * mm,
            18 * mm,
            36 * mm,
            36 * mm,
            24 * mm,
        ],
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))

    elements.append(table)

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    today = timezone.localdate().strftime("%d_%m_%Y")
    response["Content-Disposition"] = f'attachment; filename="farmer_data_{today}.pdf"'

    return response





def dealer_visit_role(user):
    profile = getattr(user, "profile", None)
    return profile.role if profile else ""


def is_dealer_visit_allowed_user(user):
    role = dealer_visit_role(user)

    return (
        user.is_superuser
        or role in [
            "ADMIN",
            "HR",
            "ACCOUNTANT",
            "STATE_HEAD",
            "REGIONAL_MANAGER",
            "ASM",
            "SALES_OFFICER_SENIOR",
            "SALES_OFFICER_JUNIOR",
        ]
    )


def get_subordinate_user_ids_for_dealer_visits(manager_user):
    collected = set()
    current_manager_ids = {manager_user.id}

    for _ in range(6):
        direct_ids = set(
            UserProfile.objects.filter(
                manager_id__in=current_manager_ids
            ).exclude(
                role="DEALER"
            ).values_list("user_id", flat=True)
        )

        new_ids = direct_ids - collected

        if not new_ids:
            break

        collected.update(new_ids)
        current_manager_ids = new_ids

    return collected


def dealer_visit_queryset_for_user(user):
    role = dealer_visit_role(user)

    visits = DealerVisit.objects.select_related(
        "dealer",
        "sales_officer",
        "sales_officer__profile",
        "dealer__state",
        "dealer__district",
    ).all()

    if user.is_superuser or role in ["ADMIN", "HR", "ACCOUNTANT"]:
        return visits

    if role in ["STATE_HEAD", "REGIONAL_MANAGER", "ASM"]:
        subordinate_ids = get_subordinate_user_ids_for_dealer_visits(user)
        visible_ids = subordinate_ids | {user.id}
        return visits.filter(sales_officer_id__in=visible_ids)

    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        return visits.filter(sales_officer=user)

    return visits.none()


def get_visit_font(size):
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]

    for path in font_paths:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass

    return ImageFont.load_default()


def generate_dealer_visit_stamped_image(visit):
    if not visit.visit_image:
        return

    try:
        visit.visit_image.open("rb")
        image = Image.open(visit.visit_image)
        image = ImageOps.exif_transpose(image).convert("RGB")
    except Exception:
        return

    max_width = 1600

    if image.width > max_width:
        ratio = max_width / float(image.width)
        new_height = int(float(image.height) * ratio)
        image = image.resize((max_width, new_height))

    width, height = image.size

    overlay_height = max(230, int(height * 0.22))
    overlay_y = height - overlay_height

    overlay = Image.new("RGBA", (width, overlay_height), (0, 0, 0, 165))
    image_rgba = image.convert("RGBA")
    image_rgba.paste(overlay, (0, overlay_y), overlay)

    draw = ImageDraw.Draw(image_rgba)

    title_font = get_visit_font(max(28, int(width * 0.030)))
    text_font = get_visit_font(max(20, int(width * 0.020)))
    small_font = get_visit_font(max(18, int(width * 0.017)))

    sales_name = visit.sales_officer.get_full_name() or visit.sales_officer.username
    dealer_name = visit.dealer.firm_name

    dealer_address = visit.dealer.firm_address or visit.dealer.owner_address or ""
    dealer_state = visit.dealer.state.name if visit.dealer.state else ""
    dealer_district = visit.dealer.district.name if visit.dealer.district else ""

    address_parts = [x for x in [dealer_address, dealer_district, dealer_state] if x]
    address_text = ", ".join(address_parts)

    if len(address_text) > 95:
        address_text = address_text[:95] + "..."

    lat_text = str(visit.latitude or "-")
    long_text = str(visit.longitude or "-")

    visit_datetime = f"{visit.visit_date.strftime('%d/%m/%Y')} {visit.visit_time.strftime('%I:%M %p')}"

    x = int(width * 0.035)
    y = overlay_y + 18

    draw.text((x, y), dealer_name, font=title_font, fill=(255, 255, 255, 255))
    y += int(width * 0.040)

    draw.text(
        (x, y),
        f"Reason: {visit.get_reason_display()}",
        font=text_font,
        fill=(245, 245, 245, 255)
    )
    y += int(width * 0.030)

    draw.text(
        (x, y),
        f"Sales Officer: {sales_name}",
        font=text_font,
        fill=(245, 245, 245, 255)
    )
    y += int(width * 0.030)

    draw.text(
        (x, y),
        f"Date: {visit_datetime}",
        font=text_font,
        fill=(245, 245, 245, 255)
    )
    y += int(width * 0.030)

    draw.text(
        (x, y),
        f"Lat {lat_text}  Long {long_text}",
        font=small_font,
        fill=(245, 245, 245, 255)
    )
    y += int(width * 0.028)

    if address_text:
        draw.text(
            (x, y),
            address_text,
            font=small_font,
            fill=(230, 230, 230, 255)
        )

    output = BytesIO()
    final_image = image_rgba.convert("RGB")
    final_image.save(output, format="JPEG", quality=92)

    filename = f"dealer_visit_{visit.id}.jpg"

    visit.stamped_image.save(
        filename,
        ContentFile(output.getvalue()),
        save=False
    )

    visit.save(update_fields=["stamped_image"])


@login_required
def dealer_visit_list(request):
    if not is_dealer_visit_allowed_user(request.user):
        messages.error(request, "You do not have permission to view dealer visits.")
        return redirect("dashboard")

    search = request.GET.get("q", "").strip()
    dealer_id = request.GET.get("dealer", "").strip()
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()

    visits = dealer_visit_queryset_for_user(request.user)

    if search:
        visits = visits.filter(
            Q(dealer__firm_name__icontains=search)
            | Q(dealer__owner_name__icontains=search)
            | Q(dealer__phone__icontains=search)
            | Q(sales_officer__username__icontains=search)
            | Q(sales_officer__first_name__icontains=search)
            | Q(sales_officer__last_name__icontains=search)
            | Q(reason__icontains=search)
            | Q(note__icontains=search)
        )

    if dealer_id:
        visits = visits.filter(dealer_id=dealer_id)

    if from_date:
        visits = visits.filter(visit_date__gte=from_date)

    if to_date:
        visits = visits.filter(visit_date__lte=to_date)

    role = dealer_visit_role(request.user)

    dealers = Dealer.objects.filter(
        approval_status="APPROVED",
        is_active=True
    ).order_by("firm_name")

    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        dealers = dealers.filter(created_by_sales_officer=request.user)

    elif role == "ASM":
        dealers = dealers.filter(concerned_asm=request.user)

    today = timezone.localdate()

    today_count = visits.filter(visit_date=today).count()
    month_count = visits.filter(
        visit_date__year=today.year,
        visit_date__month=today.month
    ).count()

    return render(
        request,
        "core/dealer_visit_list.html",
        {
            "visits": visits,
            "dealers": dealers,
            "search": search,
            "dealer_id": dealer_id,
            "from_date": from_date,
            "to_date": to_date,
            "today_count": today_count,
            "month_count": month_count,
            "role": role,
        }
    )

def calculate_distance_meters(lat1, lon1, lat2, lon2):
    lat1 = float(lat1)
    lon1 = float(lon1)
    lat2 = float(lat2)
    lon2 = float(lon2)

    earth_radius = 6371000

    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)

    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lon2 - lon1)

    a = (
        math.sin(d_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    )

    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    return earth_radius * c

@login_required
def dealer_visit_create(request):
    role = dealer_visit_role(request.user)

    if role not in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        messages.error(request, "Only Sales Officers can create dealer visits.")
        return redirect("dealer_visit_list")

    if request.method == "POST":
        form = DealerVisitForm(
            request.POST,
            request_user=request.user
        )

        latitude = request.POST.get("latitude")
        longitude = request.POST.get("longitude")
        gps_accuracy = request.POST.get("gps_accuracy")
        captured_image_data = request.POST.get("captured_image_data")

        if form.is_valid():
            selected_dealer = form.cleaned_data.get("dealer")

            if not selected_dealer:
                messages.error(request, "Please select dealer.")
                return redirect("dealer_visit_create")

            if not latitude or not longitude:
                messages.error(request, "Location is required. Please allow GPS permission.")
                return redirect("dealer_visit_create")

            if not gps_accuracy:
                messages.error(request, "GPS accuracy missing. Please capture location again.")
                return redirect("dealer_visit_create")

            if not captured_image_data:
                messages.error(request, "Please capture dealer visit photo.")
                return redirect("dealer_visit_create")

            try:
                latitude_value = Decimal(str(latitude))
                longitude_value = Decimal(str(longitude))
                gps_accuracy_value = Decimal(str(gps_accuracy))
            except Exception:
                messages.error(request, "Invalid GPS location. Please try again.")
                return redirect("dealer_visit_create")

            if gps_accuracy_value > Decimal("25"):
                messages.error(
                    request,
                    f"GPS accuracy is weak ({gps_accuracy_value} meters). "
                    "Please stand near dealer shop or open area and capture again."
                )
                return redirect("dealer_visit_create")

            if not selected_dealer.latitude or not selected_dealer.longitude:
                messages.error(
                    request,
                    "Dealer exact GPS location is not saved. Please update dealer GPS location first."
                )
                return redirect("dealer_visit_create")

            distance_meters = calculate_distance_meters(
                selected_dealer.latitude,
                selected_dealer.longitude,
                latitude_value,
                longitude_value
            )

            allowed_radius = selected_dealer.visit_radius_meters or 15

            if distance_meters > allowed_radius:
                messages.error(
                    request,
                    f"Visit rejected. You are {distance_meters:.2f} meters away from selected dealer. "
                    f"Allowed only within {allowed_radius} meters. "
                    "Please go to the correct dealer shop and capture again."
                )
                return redirect("dealer_visit_create")

            visit = form.save(commit=False)
            visit.sales_officer = request.user
            visit.latitude = latitude_value
            visit.longitude = longitude_value
            visit.gps_accuracy_meters = gps_accuracy_value
            visit.distance_from_dealer_meters = Decimal(str(round(distance_meters, 2)))
            visit.location_verified = True
            visit.visit_time = timezone.localtime().time()
            visit.save()

            try:
                format_text, image_string = captured_image_data.split(";base64,")
                image_bytes = base64.b64decode(image_string)

                visit.visit_image.save(
                    f"dealer_visit_original_{visit.id}.jpg",
                    ContentFile(image_bytes),
                    save=False
                )

                visit.stamped_image.save(
                    f"dealer_visit_stamped_{visit.id}.jpg",
                    ContentFile(image_bytes),
                    save=True
                )

            except Exception:
                messages.error(request, "Captured image could not be saved. Please try again.")
                visit.delete()
                return redirect("dealer_visit_create")

            messages.success(
                request,
                f"Dealer visit saved successfully. Distance from dealer: {distance_meters:.2f} meters."
            )
            return redirect("dealer_visit_list")

    else:
        form = DealerVisitForm(request_user=request.user)

    dealer_data = {}

    for dealer in form.fields["dealer"].queryset:
        address_parts = []

        if dealer.firm_address:
            address_parts.append(dealer.firm_address)

        if dealer.district:
            address_parts.append(dealer.district.name)

        if dealer.state:
            address_parts.append(dealer.state.name)

        dealer_data[str(dealer.id)] = {
            "name": dealer.firm_name,
            "owner": dealer.owner_name or "",
            "phone": dealer.phone or "",
            "address": ", ".join(address_parts),
            "code": dealer.dealer_code or "",
            "latitude": str(dealer.latitude or ""),
            "longitude": str(dealer.longitude or ""),
            "visit_radius_meters": dealer.visit_radius_meters or 15,
        }

    return render(
        request,
        "core/dealer_visit_form.html",
        {
            "form": form,
            "dealer_data": dealer_data,
        }
    )


    

@login_required
def dealer_visit_detail(request, visit_id):
    visit = get_object_or_404(
        dealer_visit_queryset_for_user(request.user),
        id=visit_id
    )

    return render(
        request,
        "core/dealer_visit_detail.html",
        {
            "visit": visit,
        }
    )



@login_required
def dealer_credit_score_list(request):
    profile = getattr(request.user, "profile", None)
    role = profile.role if profile else ""

    if not request.user.is_superuser and role not in ["ADMIN", "ACCOUNTANT"]:
        messages.error(request, "Only Admin or Accountant can update dealer credit score.")
        return redirect("dashboard")

    search = request.GET.get("q", "").strip()

    dealers = Dealer.objects.select_related(
        "user",
        "state",
        "district",
        "credit_score_updated_by",
    ).order_by("firm_name")

    if search:
        dealers = dealers.filter(
            Q(firm_name__icontains=search) |
            Q(owner_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(dealer_code__icontains=search)
        )

    dealer_rows = []

    for dealer in dealers:
        pending_amount = DealerInvoice.objects.filter(
            order__dealer=dealer,
            pending_amount__gt=0
        ).aggregate(
            total=Sum("pending_amount")
        )["total"] or Decimal("0.00")

        dealer_rows.append({
            "dealer": dealer,
            "pending_amount": pending_amount,
        })

    return render(request, "core/dealer_credit_score_list.html", {
        "dealer_rows": dealer_rows,
        "search": search,
    })


@login_required
@transaction.atomic
def dealer_credit_score_update(request, dealer_id):
    profile = getattr(request.user, "profile", None)
    role = profile.role if profile else ""

    if not request.user.is_superuser and role not in ["ADMIN", "ACCOUNTANT"]:
        messages.error(request, "Only Admin or Accountant can update dealer credit score.")
        return redirect("dashboard")

    dealer = get_object_or_404(Dealer, id=dealer_id)

    old_score = dealer.credit_score or 650

    if request.method == "POST":
        form = DealerCreditScoreForm(request.POST, instance=dealer)

        if form.is_valid():
            dealer_obj = form.save(commit=False)

            new_score = dealer_obj.credit_score
            change_points = new_score - old_score

            dealer_obj.credit_score_change_points = change_points
            dealer_obj.credit_score_updated_by = request.user
            dealer_obj.credit_score_updated_at = timezone.now()
            dealer_obj.save()

            DealerCreditScoreHistory.objects.create(
                dealer=dealer_obj,
                old_score=old_score,
                new_score=new_score,
                change_points=change_points,
                note=dealer_obj.credit_score_note,
                updated_by=request.user,
            )

            messages.success(request, "Dealer credit score updated successfully.")
            return redirect("dealer_credit_score_list")

    else:
        form = DealerCreditScoreForm(instance=dealer)

    return render(request, "core/dealer_credit_score_form.html", {
        "form": form,
        "dealer": dealer,
        "old_score": old_score,
    })


@login_required
def dealer_my_credit_score(request):
    profile = getattr(request.user, "profile", None)
    role = profile.role if profile else ""

    if role != "DEALER":
        messages.error(request, "Only dealer can view this page.")
        return redirect("dashboard")

    dealer = Dealer.objects.filter(user=request.user).first()

    if not dealer:
        messages.error(request, "Dealer profile not found.")
        return redirect("dashboard")

    due_amount = DealerInvoice.objects.filter(
        order__dealer=dealer,
        pending_amount__gt=0
    ).aggregate(
        total=Sum("pending_amount")
    )["total"] or Decimal("0.00")

    pending_invoice_count = DealerInvoice.objects.filter(
        order__dealer=dealer,
        pending_amount__gt=0
    ).count()

    score = dealer.credit_score or 650

    if score >= 800:
        label = "Excellent"
    elif score >= 700:
        label = "Good"
    elif score >= 600:
        label = "Average"
    else:
        label = "Needs Improvement"

    score_percent = ((score - 300) / 600) * 100
    score_percent = max(0, min(score_percent, 100))
    needle_deg = -90 + (score_percent * 1.8)

    credit_summary = {
        "score": score,
        "change_points": dealer.credit_score_change_points or 0,
        "label": label,
        "due_amount": due_amount,
        "pending_invoice_count": pending_invoice_count,
        "needle_deg": round(needle_deg, 2),
        "updated_at": dealer.credit_score_updated_at,
        "note": dealer.credit_score_note,
    }

    return render(request, "core/dealer_my_credit_score.html", {
        "dealer": dealer,
        "credit_summary": credit_summary,
    })




# ==========================================================
# PURCHASE DASHBOARD FILTER + ACTION MENU VIEWS
# Paste this at the bottom of core/views.py
#
# IMPORTANT:
# If you already have purchase_dashboard, purchase_payment_out_create,
# purchase_return_create, purchase_entry_create functions with same names,
# replace those old functions with these updated ones.
# ==========================================================

from decimal import Decimal
import csv

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone



# ----------------------------------------------------------
# COMMON PERMISSION
# ----------------------------------------------------------
def purchase_user_allowed(user):
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in [
        "ADMIN",
        "INVENTORY_MANAGER",
        "ACCOUNTANT",
    ]


def purchase_permission_check(request):
    if not purchase_user_allowed(request.user):
        messages.error(request, "Only Admin or Inventory Manager can access Purchase Management.")
        return False
    return True


# ----------------------------------------------------------
# COMMON HELPERS
# ----------------------------------------------------------

def get_filtered_purchase_bills(request):
    bills = PurchaseBill.objects.select_related("party").order_by("-bill_date", "-id")

    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    selected_party = (request.GET.get("party") or "").strip()
    selected_status = (request.GET.get("status") or "").strip()
    search = (request.GET.get("q") or "").strip()

    if from_date:
        bills = bills.filter(bill_date__gte=from_date)

    if to_date:
        bills = bills.filter(bill_date__lte=to_date)

    if selected_party:
        bills = bills.filter(party_id=selected_party)

    if selected_status:
        bills = bills.filter(payment_status=selected_status)

    if search:
        bills = bills.filter(
            Q(bill_number__icontains=search) |
            Q(party__party_name__icontains=search)
        )

    return bills, {
        "from_date": from_date,
        "to_date": to_date,
        "selected_party": selected_party,
        "selected_status": selected_status,
        "search": search,
    }


def get_or_create_simple_purchase_party(form):
    party = form.cleaned_data.get("party")

    if party:
        return party

    party_name = (form.cleaned_data.get("new_party_name") or "").strip()
    party_phone = (form.cleaned_data.get("party_phone") or "").strip()
    party_gst = (form.cleaned_data.get("party_gst") or "").strip()

    party = PurchaseParty.objects.filter(party_name__iexact=party_name).first()

    if party:
        changed = False

        if party_phone and not party.phone:
            party.phone = party_phone
            changed = True

        if party_gst and not party.gst_number:
            party.gst_number = party_gst
            changed = True

        if changed:
            party.save()

        return party

    return PurchaseParty.objects.create(
        party_name=party_name,
        phone=party_phone,
        gst_number=party_gst,
        party_type="SUPPLIER",
        is_active=True,
    )


def get_or_create_simple_purchase_type(form):
    purchase_type = form.cleaned_data.get("purchase_type")

    if purchase_type:
        return purchase_type

    type_name = (form.cleaned_data.get("new_purchase_type") or "").strip()

    purchase_type = PurchaseType.objects.filter(name__iexact=type_name).first()

    if purchase_type:
        return purchase_type

    return PurchaseType.objects.create(
        name=type_name,
        is_active=True,
    )




def build_purchase_formset_initial(bill):
    initial_rows = []

    for item in bill.items.select_related("inventory_item").all():
        inventory_item = item.inventory_item

        initial_rows.append({
            "item": inventory_item,
            "new_item_name": "",
            "sku_code": getattr(inventory_item, "sku_code", ""),
            "item_type": getattr(inventory_item, "item_type", "RAW_MATERIAL"),
            "unit": item.unit,
            "quantity": item.quantity,
            "purchase_price": item.purchase_price,
            "gst_percent": item.gst_percent,
        })

    if not initial_rows:
        initial_rows.append({})

    return initial_rows


from django.db.models import Sum, Q


def gst_report_permission_check(request):
    if not request.user.is_authenticated:
        return False

    if request.user.is_superuser:
        return True

    profile = getattr(request.user, "profile", None)

    if not profile:
        return False

    return profile.role in ["ADMIN", "ACCOUNTANT", "INVENTORY_MANAGER"]


def gstr2b_report(request):
    if not gst_report_permission_check(request):
        return redirect("dashboard")

    search = (request.GET.get("q") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()

    bills = PurchaseBill.objects.select_related(
        "party",
        "created_by",
    ).filter(
        add_to_gstr2b=True
    ).order_by("-bill_date", "-id")

    if search:
        bills = bills.filter(
            Q(bill_number__icontains=search) |
            Q(party__party_name__icontains=search) |
            Q(party__gst_number__icontains=search)
        )

    if from_date:
        bills = bills.filter(bill_date__gte=from_date)

    if to_date:
        bills = bills.filter(bill_date__lte=to_date)

    totals = bills.aggregate(
        taxable_total=Sum("subtotal_amount"),
        gst_total=Sum("gst_amount"),
        bill_total=Sum("total_amount"),
    )

    return render(request, "core/gstr2b_report.html", {
        "bills": bills,
        "search": search,
        "from_date": from_date,
        "to_date": to_date,
        "taxable_total": totals["taxable_total"] or Decimal("0.00"),
        "gst_total": totals["gst_total"] or Decimal("0.00"),
        "bill_total": totals["bill_total"] or Decimal("0.00"),
    })

from django.db.models import Sum, Q
from decimal import Decimal


def gst_sales_report_permission_check(request):
    if not request.user.is_authenticated:
        return False

    if request.user.is_superuser:
        return True

    profile = getattr(request.user, "profile", None)

    if not profile:
        return False

    return profile.role in ["ADMIN", "ACCOUNTANT"]


def gstr1_report(request):
    if not gst_sales_report_permission_check(request):
        return redirect("dashboard")

    search = (request.GET.get("q") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()

    invoices = DealerInvoice.objects.select_related(
        "order",
        "order__dealer",
        "released_by",
    ).filter(
        add_to_gstr1=True
    ).order_by("-invoice_date", "-id")

    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(order__dealer__firm_name__icontains=search) |
            Q(order__dealer__owner_name__icontains=search) |
            Q(order__dealer__gst_number__icontains=search)
        )

    if from_date:
        invoices = invoices.filter(invoice_date__gte=from_date)

    if to_date:
        invoices = invoices.filter(invoice_date__lte=to_date)

    totals = invoices.aggregate(
        taxable_total=Sum("subtotal_amount"),
        gst_total=Sum("gst_amount"),
        invoice_total=Sum("total_amount"),
    )

    return render(request, "core/gstr1_report.html", {
        "invoices": invoices,
        "search": search,
        "from_date": from_date,
        "to_date": to_date,
        "taxable_total": totals["taxable_total"] or Decimal("0.00"),
        "gst_total": totals["gst_total"] or Decimal("0.00"),
        "invoice_total": totals["invoice_total"] or Decimal("0.00"),
    })



from decimal import Decimal
from django.db.models import Sum, Q
from django.shortcuts import render, redirect


def gstr3b_report(request):
    if not gst_sales_report_permission_check(request):
        return redirect("dashboard")

    search = (request.GET.get("q") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()

    sales_invoices = DealerInvoice.objects.select_related(
        "order",
        "order__dealer",
        "released_by",
    ).filter(
        add_to_gstr1=True
    ).order_by("-invoice_date", "-id")

    purchase_bills = PurchaseBill.objects.select_related(
        "party",
        "created_by",
    ).filter(
        add_to_gstr2b=True
    ).order_by("-bill_date", "-id")

    if from_date:
        sales_invoices = sales_invoices.filter(invoice_date__gte=from_date)
        purchase_bills = purchase_bills.filter(bill_date__gte=from_date)

    if to_date:
        sales_invoices = sales_invoices.filter(invoice_date__lte=to_date)
        purchase_bills = purchase_bills.filter(bill_date__lte=to_date)

    if search:
        sales_invoices = sales_invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(order__dealer__firm_name__icontains=search) |
            Q(order__dealer__gst_number__icontains=search)
        )

        purchase_bills = purchase_bills.filter(
            Q(bill_number__icontains=search) |
            Q(party__party_name__icontains=search) |
            Q(party__gst_number__icontains=search)
        )

    sales_totals = sales_invoices.aggregate(
        taxable_total=Sum("subtotal_amount"),
        gst_total=Sum("gst_amount"),
        invoice_total=Sum("total_amount"),
    )

    purchase_totals = purchase_bills.aggregate(
        taxable_total=Sum("subtotal_amount"),
        gst_total=Sum("gst_amount"),
        bill_total=Sum("total_amount"),
    )

    output_taxable = sales_totals["taxable_total"] or Decimal("0.00")
    output_gst = sales_totals["gst_total"] or Decimal("0.00")
    output_total = sales_totals["invoice_total"] or Decimal("0.00")

    input_taxable = purchase_totals["taxable_total"] or Decimal("0.00")
    input_gst = purchase_totals["gst_total"] or Decimal("0.00")
    input_total = purchase_totals["bill_total"] or Decimal("0.00")

    gst_payable = output_gst - input_gst

    if gst_payable > 0:
        payable_amount = gst_payable
        credit_carried_forward = Decimal("0.00")
        result_status = "PAYABLE"
    elif gst_payable < 0:
        payable_amount = Decimal("0.00")
        credit_carried_forward = abs(gst_payable)
        result_status = "CREDIT"
    else:
        payable_amount = Decimal("0.00")
        credit_carried_forward = Decimal("0.00")
        result_status = "NIL"

    return render(request, "core/gstr3b_report.html", {
        "sales_invoices": sales_invoices,
        "purchase_bills": purchase_bills,

        "search": search,
        "from_date": from_date,
        "to_date": to_date,

        "output_taxable": output_taxable,
        "output_gst": output_gst,
        "output_total": output_total,

        "input_taxable": input_taxable,
        "input_gst": input_gst,
        "input_total": input_total,

        "gst_payable": gst_payable,
        "payable_amount": payable_amount,
        "credit_carried_forward": credit_carried_forward,
        "result_status": result_status,
    })

    

# ----------------------------------------------------------
# DASHBOARD WITH FILTERS, TOTALS, THREE DOTS
# ----------------------------------------------------------

def purchase_dashboard(request):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bills, filters = get_filtered_purchase_bills(request)

    total_amount = bills.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")
    paid_amount = bills.aggregate(total=Sum("paid_amount"))["total"] or Decimal("0.00")
    unpaid_amount = bills.aggregate(total=Sum("balance_amount"))["total"] or Decimal("0.00")

    parties = PurchaseParty.objects.filter(is_active=True).order_by("party_name")

    total_inventory_items = InventoryItem.objects.filter(is_active=True).count()
    total_parties = parties.count()
    unpaid_bill_count = bills.filter(balance_amount__gt=0).count()

    context = {
        "bills": bills,
        "recent_bills": bills[:15],
        "parties": parties,

        "paid_amount": paid_amount,
        "unpaid_amount": unpaid_amount,
        "total_amount": total_amount,

        "total_inventory_items": total_inventory_items,
        "total_parties": total_parties,
        "unpaid_bill_count": unpaid_bill_count,
    }
    context.update(filters)

    return render(request, "core/purchase_dashboard.html", context)


# ----------------------------------------------------------
# EXCEL REPORT - CSV FORMAT OPENS IN EXCEL
# ----------------------------------------------------------

def purchase_dashboard_excel(request):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bills, filters = get_filtered_purchase_bills(request)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="purchase_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Date",
        "Invoice No",
        "Party Name",
        "Payment Type",
        "Total Amount",
        "Paid Amount",
        "Balance Due",
        "Status",
    ])

    for bill in bills:
        writer.writerow([
            bill.bill_date.strftime("%d/%m/%Y") if bill.bill_date else "",
            bill.bill_number,
            bill.party.party_name if bill.party else "",
            "Arvis Fertilizers",
            bill.total_amount,
            bill.paid_amount,
            bill.balance_amount,
            bill.get_payment_status_display(),
        ])

    writer.writerow([])
    writer.writerow(["Paid", bills.aggregate(total=Sum("paid_amount"))["total"] or Decimal("0.00")])
    writer.writerow(["Unpaid", bills.aggregate(total=Sum("balance_amount"))["total"] or Decimal("0.00")])
    writer.writerow(["Total", bills.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00")])

    return response


# ----------------------------------------------------------
# PURCHASE ENTRY CREATE
# Existing item uses saved stock unit automatically.
# New item uses selected unit.
# ----------------------------------------------------------

@transaction.atomic
def purchase_entry_create(request):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    if request.method == "POST":
        form = SimplePurchaseEntryForm(request.POST)
        formset = SimplePurchaseItemFormSet(request.POST, prefix="items")

        if form.is_valid() and formset.is_valid():
            valid_rows = get_valid_purchase_item_rows(formset)

            if not valid_rows:
                messages.error(request, "Please add at least one purchase item.")
                return render(request, "core/purchase_entry_form.html", {
                    "form": form,
                    "formset": formset,
                    "title": "New Purchase Entry",
                })

            try:
                party = get_or_create_simple_purchase_party(form)
                purchase_type = get_or_create_simple_purchase_type(form)
                is_asset_bill = is_asset_purchase_type(purchase_type)

                add_to_gstr2b = form.cleaned_data.get("add_to_gstr2b", False)

                bill = PurchaseBill.objects.create(
                    bill_number=form.cleaned_data["bill_number"],
                    bill_date=form.cleaned_data["bill_date"],
                    party=party,
                    note=form.cleaned_data.get("note"),
                    add_to_gstr2b=add_to_gstr2b,
                    gstr2b_added_at=timezone.now() if add_to_gstr2b else None,
                    created_by=request.user,
                )

                total_assets_created = 0

                for row in valid_rows:
                    inventory_item = get_or_create_simple_inventory_item(row, purchase_type)

                    if is_asset_bill:
                        line_unit = "PIECE"
                    else:
                        line_unit = inventory_item.stock_unit if row.get("item") else row["unit"]

                    PurchaseBillItem.objects.create(
                        bill=bill,
                        inventory_item=inventory_item,
                        quantity=row["quantity"],
                        unit=line_unit,
                        purchase_price=row["purchase_price"],
                        gst_percent=row.get("gst_percent") or Decimal("0.00"),
                    )

                    if is_asset_bill:
                        total_assets_created += create_assets_from_purchase_row(
                            row=row,
                            bill=bill,
                            request_user=request.user,
                        )

                bill.recalculate_totals()

                if is_asset_bill:
                    # Asset purchase should NOT add KG/LITRE/Raw Material stock.
                    bill.stock_added = False
                    bill.save(update_fields=["stock_added"])
                    messages.success(
                        request,
                        f"Asset purchase saved. {total_assets_created} asset(s) created in Asset Management."
                    )
                else:
                    bill.apply_stock_in(user=request.user)
                    messages.success(request, "Purchase entry saved and stock added.")

                return redirect("purchase_dashboard")

            except ValidationError as error:
                messages.error(request, error)
                transaction.set_rollback(True)

            except Exception as error:
                messages.error(request, f"Purchase entry could not be saved: {error}")
                transaction.set_rollback(True)

    else:
        form = SimplePurchaseEntryForm(
            initial={
                "bill_date": timezone.localdate(),
            }
        )
        formset = SimplePurchaseItemFormSet(prefix="items")

    return render(request, "core/purchase_entry_form.html", {
        "form": form,
        "formset": formset,
        "title": "New Purchase Entry",
    })






# ----------------------------------------------------------
# PRINT PURCHASE ENTRY
# ----------------------------------------------------------

def purchase_entry_print(request, bill_id):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bill = get_object_or_404(
        PurchaseBill.objects.select_related("party", "created_by"),
        id=bill_id,
    )

    items = bill.items.select_related("inventory_item").all()

    return render(
        request,
        "core/purchase_entry_print.html",
        {
            "bill": bill,
            "items": items,
        }
    )


# ==========================================================
# PURCHASE THREE DOT ACTIONS - VIEWS
# Paste this at the bottom of core/views.py
# ==========================================================

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone


def purchase_entry_detail(request, bill_id):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bill = get_object_or_404(
        PurchaseBill.objects.select_related("party", "created_by"),
        id=bill_id,
    )

    items = bill.items.select_related("inventory_item").all()

    item_rows = []
    subtotal_amount = Decimal("0.00")
    gst_total = Decimal("0.00")
    grand_total = Decimal("0.00")

    for item in items:
        qty = item.quantity or Decimal("0.00")
        rate = item.purchase_price or Decimal("0.00")
        gst_percent = item.gst_percent or Decimal("0.00")

        taxable_amount = qty * rate
        gst_amount = (taxable_amount * gst_percent) / Decimal("100.00")
        line_total = taxable_amount + gst_amount

        subtotal_amount += taxable_amount
        gst_total += gst_amount
        grand_total += line_total

        item_rows.append({
            "item": item,
            "taxable_amount": taxable_amount,
            "gst_amount": gst_amount,
            "line_total": line_total,
        })

    payments = PurchasePaymentOut.objects.filter(
        bill=bill
    ).select_related("party").order_by("-payment_date", "-id")

    return render(
        request,
        "core/purchase_entry_detail.html",
        {
            "bill": bill,
            "items": items,
            "item_rows": item_rows,
            "payments": payments,
            "subtotal_amount": subtotal_amount,
            "gst_total": gst_total,
            "grand_total": grand_total,
        }
    )

    
@transaction.atomic
def purchase_entry_delete(request, bill_id):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bill = get_object_or_404(PurchaseBill, id=bill_id)

    if request.method != "POST":
        messages.error(request, "Invalid delete request.")
        return redirect("purchase_dashboard")

    try:
        # If stock was already added, reverse stock before delete.
        if getattr(bill, "stock_added", False):
            for old_item in bill.items.select_related("inventory_item").all():
                old_item.inventory_item.reduce_stock(
                    old_item.quantity,
                    old_item.unit,
                    user=request.user,
                    reference_type="PURCHASE_DELETE_REVERSE",
                    reference_id=bill.id,
                    note=f"Reverse stock before deleting purchase bill {bill.bill_number}",
                )

        bill_number = bill.bill_number
        bill.delete()

        messages.success(request, f"Purchase bill {bill_number} deleted successfully.")
        return redirect("purchase_dashboard")

    except ValidationError as error:
        messages.error(request, error)
        transaction.set_rollback(True)

    except Exception as error:
        messages.error(
            request,
            f"Could not delete this bill: {error}. "
            f"If this stock is already used in production, create Purchase Return instead."
        )
        transaction.set_rollback(True)

    return redirect("purchase_dashboard")


@transaction.atomic
def purchase_entry_duplicate(request, bill_id):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    old_bill = get_object_or_404(
        PurchaseBill.objects.select_related("party"),
        id=bill_id,
    )

    try:
        today_code = timezone.localdate().strftime("%d%m%Y")
        new_bill_number = f"{old_bill.bill_number}-COPY-{today_code}"

        counter = 1
        original_new_bill_number = new_bill_number

        while PurchaseBill.objects.filter(bill_number=new_bill_number).exists():
            counter += 1
            new_bill_number = f"{original_new_bill_number}-{counter}"

        new_bill = PurchaseBill.objects.create(
            bill_number=new_bill_number,
            bill_date=timezone.localdate(),
            party=old_bill.party,
            note=f"Duplicated from bill {old_bill.bill_number}",
            created_by=request.user,
        )

        for old_item in old_bill.items.select_related("inventory_item").all():
            PurchaseBillItem.objects.create(
                bill=new_bill,
                inventory_item=old_item.inventory_item,
                quantity=old_item.quantity,
                unit=old_item.unit,
                purchase_price=old_item.purchase_price,
                gst_percent=old_item.gst_percent,
            )

        new_bill.recalculate_totals()

        messages.success(
            request,
            f"Bill duplicated as {new_bill.bill_number}. Please review/edit and save stock if needed."
        )
        return redirect("purchase_entry_edit", bill_id=new_bill.id)

    except Exception as error:
        messages.error(request, f"Could not duplicate bill: {error}")
        transaction.set_rollback(True)
        return redirect("purchase_dashboard")


def purchase_entry_history(request, bill_id):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bill = get_object_or_404(
        PurchaseBill.objects.select_related("party", "created_by"),
        id=bill_id,
    )

    items = bill.items.select_related("inventory_item").all()

    payments = PurchasePaymentOut.objects.filter(
        bill=bill
    ).select_related("party").order_by("-payment_date", "-id")

    returns = PurchaseReturnOrder.objects.filter(
        bill=bill
    ).select_related("party", "created_by").order_by("-return_date", "-id")

    return render(
        request,
        "core/purchase_entry_history.html",
        {
            "bill": bill,
            "items": items,
            "payments": payments,
            "returns": returns,
        }
    )


# ----------------------------------------------------------
# PAYMENT OUT WITH ?bill= PRESELECT
# ----------------------------------------------------------

@transaction.atomic
def purchase_payment_out_create(request):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bill_id = request.GET.get("bill")

    if request.method == "POST":
        form = SimplePaymentOutForm(request.POST, request.FILES)

        if form.is_valid():
            try:
                payment = form.save(commit=False)
                payment.party = payment.bill.party
                payment.created_by = request.user
                payment.full_clean()
                payment.save()

                messages.success(
                    request,
                    f"Payment saved. Bill status: {payment.bill.get_payment_status_display()}."
                )
                return redirect("purchase_dashboard")

            except ValidationError as error:
                messages.error(request, error)
                transaction.set_rollback(True)

    else:
        initial = {
            "payment_date": timezone.localdate(),
        }

        if bill_id:
            initial["bill"] = bill_id

        form = SimplePaymentOutForm(initial=initial)

    recent_payments = PurchasePaymentOut.objects.select_related(
        "party",
        "bill"
    ).order_by("-id")[:15]

    return render(
        request,
        "core/purchase_payment_out_form.html",
        {
            "form": form,
            "recent_payments": recent_payments,
            "title": "Payment Out",
        }
    )


# ----------------------------------------------------------
# PURCHASE RETURN WITH ?bill= PRESELECT
# ----------------------------------------------------------

@transaction.atomic
def purchase_return_create(request):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bill_id = request.GET.get("bill")

    if request.method == "POST":
        form = SimplePurchaseReturnForm(request.POST)
        formset = SimpleReturnItemFormSet(request.POST, prefix="items")

        if form.is_valid() and formset.is_valid():
            valid_rows = []

            for item_form in formset:
                if item_form.cleaned_data.get("DELETE"):
                    continue

                inventory_item = item_form.cleaned_data.get("inventory_item")
                quantity = item_form.cleaned_data.get("quantity")

                if not inventory_item and not quantity:
                    continue

                valid_rows.append(item_form.cleaned_data)

            if not valid_rows:
                messages.error(request, "Please add at least one return item.")
                return render(
                    request,
                    "core/purchase_return_form.html",
                    {
                        "form": form,
                        "formset": formset,
                        "title": "Purchase Return / Debit Note",
                    }
                )

            try:
                bill = form.cleaned_data["bill"]

                purchase_return = PurchaseReturnOrder.objects.create(
                    party=bill.party,
                    bill=bill,
                    return_date=form.cleaned_data["return_date"],
                    reason=form.cleaned_data["reason"],
                    created_by=request.user,
                )

                for row in valid_rows:
                    PurchaseReturnItem.objects.create(
                        purchase_return=purchase_return,
                        inventory_item=row["inventory_item"],
                        quantity=row["quantity"],
                        unit=row["unit"],
                        return_price=row["return_price"],
                        gst_percent=row.get("gst_percent") or Decimal("0.00"),
                    )

                purchase_return.recalculate_totals()
                purchase_return.apply_return_stock(user=request.user)

                messages.success(
                    request,
                    f"Debit note created and stock reduced. Debit Note: {purchase_return.debit_note_number}."
                )
                return redirect("purchase_dashboard")

            except ValidationError as error:
                messages.error(request, error)
                transaction.set_rollback(True)

            except Exception as error:
                messages.error(request, f"Return could not be saved: {error}")
                transaction.set_rollback(True)

    else:
        initial = {
            "return_date": timezone.localdate(),
        }

        if bill_id:
            initial["bill"] = bill_id

        form = SimplePurchaseReturnForm(initial=initial)
        formset = SimpleReturnItemFormSet(prefix="items")

    return render(
        request,
        "core/purchase_return_form.html",
        {
            "form": form,
            "formset": formset,
            "title": "Purchase Return / Debit Note",
        }
    )

@transaction.atomic
def production_make_product_create(request):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    if request.method == "POST":
        form = SimpleMakeProductForm(request.POST)
        formset = SimpleMakeProductItemFormSet(request.POST, prefix="items")

        if form.is_valid() and formset.is_valid():
            valid_rows = []

            for item_form in formset:
                if item_form.cleaned_data.get("DELETE"):
                    continue

                inventory_item = item_form.cleaned_data.get("inventory_item")
                quantity = item_form.cleaned_data.get("quantity")

                if not inventory_item and not quantity:
                    continue

                valid_rows.append(item_form.cleaned_data)

            if not valid_rows:
                messages.error(request, "Please add at least one used raw/packing item.")
                return render(
                    request,
                    "core/production_make_product_form.html",
                    {
                        "form": form,
                        "formset": formset,
                        "title": "Production / Repacking",
                    }
                )

            try:
                output_pack = form.cleaned_data["output_product_pack"]
                output_boxes = form.cleaned_data["output_boxes"]

                recipe = ProductionRecipe.objects.create(
                    name=form.cleaned_data["production_name"],
                    output_product_pack=output_pack,
                    output_boxes=output_boxes,
                    note=form.cleaned_data.get("note"),
                    is_active=True,
                )

                for row in valid_rows:
                    ProductionRecipeItem.objects.create(
                        recipe=recipe,
                        inventory_item=row["inventory_item"],
                        quantity_required=row["quantity"],
                        unit=row["unit"],
                    )

                batch = ProductionBatch.objects.create(
                    recipe=recipe,
                    output_product_pack=output_pack,
                    output_boxes=output_boxes,
                    production_date=timezone.localdate(),
                    note=form.cleaned_data.get("note"),
                    created_by=request.user,
                )

                batch.apply_production(user=request.user)

                messages.success(
                    request,
                    f"Production completed. Raw stock reduced and final product stock increased by {output_boxes}."
                )
                return redirect("purchase_dashboard")

            except ValidationError as error:
                messages.error(request, error)
                transaction.set_rollback(True)

            except Exception as error:
                messages.error(request, f"Production could not be saved: {error}")
                transaction.set_rollback(True)

    else:
        form = SimpleMakeProductForm()
        formset = SimpleMakeProductItemFormSet(prefix="items")

    return render(
        request,
        "core/production_make_product_form.html",
        {
            "form": form,
            "formset": formset,
            "title": "Production / Repacking",
        }
    )






# ==========================================================
# MAIN WAREHOUSE STOCK TRANSFER - VIEWS
# Paste this at the bottom of core/views.py
# ==========================================================

from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import redirect, render
from .forms import ProductStockTransferForm, get_default_main_warehouse
from .models import ProductPackSize, ProductStockTransfer


def stock_transfer_allowed(user):
    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)
    role = profile.role if profile else ""

    return role in ["ADMIN", "INVENTORY_MANAGER"]


def copy_or_get_target_product_pack(source_pack, to_warehouse):
    target_pack = ProductPackSize.objects.filter(
        product=source_pack.product,
        warehouse=to_warehouse,
        pack_size=source_pack.pack_size,
        unit=source_pack.unit,
        packing_type=source_pack.packing_type,
        units_per_box=source_pack.units_per_box,
    ).first()

    if target_pack:
        return target_pack

    return ProductPackSize.objects.create(
        product=source_pack.product,
        warehouse=to_warehouse,
        pack_size=source_pack.pack_size,
        unit=source_pack.unit,
        packing_type=source_pack.packing_type,
        units_per_box=source_pack.units_per_box,
        mrp_per_unit=source_pack.mrp_per_unit,
        sale_price_per_unit=source_pack.sale_price_per_unit,
        purchase_price_per_unit=source_pack.purchase_price_per_unit,
        box_sale_price=source_pack.box_sale_price,
        stock_boxes=Decimal("0.000"),
        is_active=True,
    )


@login_required
def product_stock_transfer_list(request):
    if not stock_transfer_allowed(request.user):
        messages.error(request, "Only Admin or Inventory Manager can transfer stock.")
        return redirect("dashboard")

    main = get_default_main_warehouse()

    main_packs = ProductPackSize.objects.select_related(
        "product",
        "warehouse",
    ).filter(
        warehouse=main,
        is_active=True,
    ).order_by("product__name", "pack_size") if main else ProductPackSize.objects.none()

    transfers = ProductStockTransfer.objects.select_related(
        "product",
        "product_pack",
        "from_warehouse",
        "to_warehouse",
        "created_by",
    ).order_by("-created_at")[:50]

    return render(
        request,
        "core/product_stock_transfer_list.html",
        {
            "main": main,
            "main_packs": main_packs,
            "transfers": transfers,
        }
    )


@login_required
@transaction.atomic
def product_stock_transfer_create(request):
    if not stock_transfer_allowed(request.user):
        messages.error(request, "Only Admin or Inventory Manager can transfer stock.")
        return redirect("dashboard")

    if request.method == "POST":
        form = ProductStockTransferForm(request.POST)

        if form.is_valid():
            source_pack = form.cleaned_data["product_pack"]
            to_warehouse = form.cleaned_data["to_warehouse"]
            quantity = form.cleaned_data["quantity_boxes"]

            try:
                source_pack = ProductPackSize.objects.select_for_update().get(id=source_pack.id)

                available = source_pack.stock_boxes or Decimal("0.000")

                if quantity > available:
                    messages.error(request, f"Only {available} available in Main Warehouse.")
                    transaction.set_rollback(True)
                    return render(
                        request,
                        "core/product_stock_transfer_form.html",
                        {"form": form, "source_pack": source_pack}
                    )

                before_main_stock = available
                after_main_stock = available - quantity

                target_pack = copy_or_get_target_product_pack(source_pack, to_warehouse)
                target_pack = ProductPackSize.objects.select_for_update().get(id=target_pack.id)

                source_pack.stock_boxes = after_main_stock
                source_pack.save(update_fields=["stock_boxes"])

                target_pack.stock_boxes = (target_pack.stock_boxes or Decimal("0.000")) + quantity
                target_pack.save(update_fields=["stock_boxes"])

                ProductStockTransfer.objects.create(
                    product_pack=source_pack,
                    product=source_pack.product,
                    from_warehouse=source_pack.warehouse,
                    to_warehouse=to_warehouse,
                    quantity_boxes=quantity,
                    before_main_stock=before_main_stock,
                    after_main_stock=after_main_stock,
                    note=form.cleaned_data.get("note"),
                    created_by=request.user,
                )

                messages.success(
                    request,
                    f"Stock transferred successfully. Main warehouse remaining stock: {after_main_stock}."
                )
                return redirect("product_stock_transfer_list")

            except Exception as error:
                messages.error(request, f"Stock transfer failed: {error}")
                transaction.set_rollback(True)

    else:
        form = ProductStockTransferForm()

    return render(
        request,
        "core/product_stock_transfer_form.html",
        {"form": form}
    )


# ----------------------------------------------------------
# PURCHASE STOCK USAGE HISTORY
# ----------------------------------------------------------

from decimal import Decimal
from django.db.models import Q, Sum
from django.shortcuts import render, redirect
from django.utils import timezone






# ==========================================================
# ASSET PURCHASE SUPPORT
# If Purchase Type is Asset:
# - Do not add stock in KG/LITRE inventory
# - Create AssetItem records automatically
# - Keep purchase bill for payment/accounting
# ==========================================================

from decimal import Decimal
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.shortcuts import redirect, render
from django.utils import timezone

from .models import (
    PurchaseBill,
    PurchaseBillItem,
    InventoryItem,
    AssetCategory,
    AssetItem,
)
from .forms import SimplePurchaseEntryForm, SimplePurchaseItemFormSet


def is_asset_purchase_type(purchase_type):
    if not purchase_type:
        return False

    name = str(getattr(purchase_type, "name", "") or "").strip().lower()

    return (
        name == "asset"
        or name == "assets"
        or "asset" in name
        or "fixed asset" in name
        or "company asset" in name
    )


def get_inventory_fallback_item_type():
    try:
        field = InventoryItem._meta.get_field("item_type")
        allowed = [choice[0] for choice in field.choices]

        if "ASSET" in allowed:
            return "ASSET"

        if "OTHER" in allowed:
            return "OTHER"

        if "RAW_MATERIAL" in allowed:
            return "RAW_MATERIAL"
    except Exception:
        pass

    return "RAW_MATERIAL"


def get_or_create_asset_category():
    category = AssetCategory.objects.filter(code_prefix__iexact="AST").first()

    if category:
        return category

    category = AssetCategory.objects.filter(name__iexact="Purchased Assets").first()

    if category:
        return category

    return AssetCategory.objects.create(
        name="Purchased Assets",
        code_prefix="AST",
        description="Assets purchased from Purchase Management",
        is_active=True,
    )


def generate_purchase_asset_code(category):
    prefix = category.code_prefix.upper().strip()

    last_asset = AssetItem.objects.filter(
        asset_code__startswith=f"{prefix}-"
    ).order_by("-id").first()

    if last_asset:
        try:
            last_number = int(last_asset.asset_code.split("-")[-1])
        except Exception:
            last_number = AssetItem.objects.filter(
                asset_code__startswith=f"{prefix}-"
            ).count()
    else:
        last_number = 0

    next_number = last_number + 1

    while True:
        asset_code = f"{prefix}-{next_number:04d}"

        if not AssetItem.objects.filter(asset_code=asset_code).exists():
            return asset_code

        next_number += 1


def get_or_create_simple_inventory_item(row_data, purchase_type):
    selected_item = row_data.get("item")

    # Normal purchase can use existing inventory item
    if selected_item and not is_asset_purchase_type(purchase_type):
        return selected_item

    item_name = (row_data.get("new_item_name") or "").strip()
    sku_code = (row_data.get("sku_code") or "").strip()
    purchase_price = row_data.get("purchase_price") or Decimal("0.00")

    if is_asset_purchase_type(purchase_type):
        item_type = get_inventory_fallback_item_type()
        unit = "PIECE"
    else:
        item_type = row_data.get("item_type") or "RAW_MATERIAL"
        unit = row_data.get("unit") or "PIECE"

    existing = InventoryItem.objects.filter(
        name__iexact=item_name,
        purchase_type=purchase_type,
    ).first()

    if existing:
        return existing

    return InventoryItem.objects.create(
        purchase_type=purchase_type,
        item_type=item_type,
        name=item_name,
        sku_code=sku_code,
        stock_unit=unit,
        stock_quantity=Decimal("0.000"),
        average_purchase_price=purchase_price,
        minimum_stock_alert=Decimal("0.000"),
        is_active=True,
    )


def create_assets_from_purchase_row(row, bill, request_user):
    asset_name = (row.get("new_item_name") or "").strip()
    quantity = row.get("quantity") or Decimal("0")
    purchase_price = row.get("purchase_price") or Decimal("0.00")
    gst_percent = row.get("gst_percent") or Decimal("0.00")

    if not asset_name:
        raise ValidationError("Asset name is required.")

    if quantity <= 0:
        raise ValidationError(f"{asset_name}: quantity must be greater than zero.")

    if quantity != quantity.to_integral_value():
        raise ValidationError(f"{asset_name}: asset quantity must be whole number like 1, 2, 3.")

    quantity_count = int(quantity)

    category = get_or_create_asset_category()

    gst_amount_per_asset = (purchase_price * gst_percent) / Decimal("100.00")
    purchase_value_per_asset = purchase_price + gst_amount_per_asset

    created_count = 0

    for i in range(quantity_count):
        asset_code = generate_purchase_asset_code(category)

        AssetItem.objects.create(
            category=category,
            asset_code=asset_code,
            name=asset_name,
            brand="",
            model_number="",
            serial_number="",
            description=f"Created from purchase bill {bill.bill_number}",
            purchase_date=bill.bill_date,
            purchase_value=purchase_value_per_asset,
            status="AVAILABLE",
            created_by=request_user,
        )

        created_count += 1

    return created_count


# ==========================================================
# FIXED PURCHASE ENTRY EDIT
# Fixes TransactionManagementError by using atomic only inside try block
# ==========================================================

from django.core.exceptions import ValidationError
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from decimal import Decimal


def purchase_error_text(error):
    if hasattr(error, "messages"):
        return " ".join(error.messages)
    return str(error)


def safe_reduce_purchase_stock(inventory_item, quantity, unit, user, bill):
    """
    Some reduce_stock functions accept reference_type/reference_id/note.
    Some accept only quantity, unit, user.
    This supports both.
    """
    try:
        inventory_item.reduce_stock(
            quantity,
            unit,
            user=user,
            reference_type="PURCHASE_EDIT_REVERSE",
            reference_id=bill.id,
            note=f"Reverse stock before editing purchase bill {bill.bill_number}",
        )
    except TypeError:
        inventory_item.reduce_stock(
            quantity,
            unit,
            user=user,
        )


def purchase_entry_edit(request, bill_id):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    bill = get_object_or_404(
        PurchaseBill.objects.select_related("party"),
        id=bill_id,
    )

    first_item = bill.items.select_related("inventory_item__purchase_type").first()
    existing_purchase_type = None

    if first_item and first_item.inventory_item:
        existing_purchase_type = first_item.inventory_item.purchase_type

    if request.method == "POST":
        form = SimplePurchaseEntryForm(request.POST)
        formset = SimplePurchaseItemFormSet(request.POST, prefix="items")

        if form.is_valid() and formset.is_valid():
            try:
                valid_rows = get_valid_purchase_item_rows(formset)

                if not valid_rows:
                    messages.error(request, "Please add at least one purchase item.")
                    return render(request, "core/purchase_entry_form.html", {
                        "form": form,
                        "formset": formset,
                        "title": f"View/Edit Purchase Entry - {bill.bill_number}",
                        "editing_bill": bill,
                    })

                with transaction.atomic():
                    bill_locked = PurchaseBill.objects.select_for_update().get(id=bill.id)

                    # Reverse old stock first, only if already added.
                    if getattr(bill_locked, "stock_added", False):
                        for old_item in bill_locked.items.select_related("inventory_item").all():
                            safe_reduce_purchase_stock(
                                inventory_item=old_item.inventory_item,
                                quantity=old_item.quantity,
                                unit=old_item.unit,
                                user=request.user,
                                bill=bill_locked,
                            )

                        bill_locked.stock_added = False
                        bill_locked.save(update_fields=["stock_added"])

                    party = get_or_create_simple_purchase_party(form)
                    purchase_type = get_or_create_simple_purchase_type(form)

                    add_to_gstr2b = form.cleaned_data.get("add_to_gstr2b", False)

                    bill_locked.party = party
                    bill_locked.bill_number = form.cleaned_data["bill_number"]
                    bill_locked.bill_date = form.cleaned_data["bill_date"]
                    bill_locked.note = form.cleaned_data.get("note")

                    # GSTR-2B option
                    bill_locked.add_to_gstr2b = add_to_gstr2b

                    if add_to_gstr2b and not bill_locked.gstr2b_added_at:
                        bill_locked.gstr2b_added_at = timezone.now()

                    if not add_to_gstr2b:
                        bill_locked.gstr2b_added_at = None

                    bill_locked.save()

                    # Delete old bill items and create new edited items
                    bill_locked.items.all().delete()

                    is_asset_bill = False
                    try:
                        is_asset_bill = is_asset_purchase_type(purchase_type)
                    except Exception:
                        is_asset_bill = False

                    for row in valid_rows:
                        inventory_item = get_or_create_simple_inventory_item(row, purchase_type)

                        if is_asset_bill:
                            line_unit = "PIECE"
                        else:
                            line_unit = inventory_item.stock_unit if row.get("item") else row["unit"]

                        PurchaseBillItem.objects.create(
                            bill=bill_locked,
                            inventory_item=inventory_item,
                            quantity=row["quantity"],
                            unit=line_unit,
                            purchase_price=row["purchase_price"],
                            gst_percent=row.get("gst_percent") or Decimal("0.00"),
                        )

                    # Recalculate taxable, GST, total, balance
                    bill_locked.recalculate_totals()

                    if is_asset_bill:
                        bill_locked.stock_added = False
                        bill_locked.save(update_fields=["stock_added"])
                    else:
                        bill_locked.apply_stock_in(user=request.user)

                messages.success(request, "Purchase entry updated successfully.")
                return redirect("purchase_dashboard")

            except ValidationError as error:
                messages.error(request, purchase_error_text(error))

            except Exception as error:
                messages.error(request, f"Purchase entry could not be updated: {error}")

    else:
        initial = {
            "party": bill.party,
            "bill_number": bill.bill_number,
            "bill_date": bill.bill_date,
            "note": bill.note,

            # GSTR-2B checked status while editing
            "add_to_gstr2b": getattr(bill, "add_to_gstr2b", False),
        }

        if existing_purchase_type:
            initial["purchase_type"] = existing_purchase_type

        form = SimplePurchaseEntryForm(initial=initial)

        item_initial = []

        is_asset_edit_bill = False
        try:
            is_asset_edit_bill = is_asset_purchase_type(existing_purchase_type)
        except Exception:
            is_asset_edit_bill = False

        for item in bill.items.select_related("inventory_item").all():
            inventory_item = item.inventory_item

            item_initial.append({
                "item": None if is_asset_edit_bill else inventory_item,
                "new_item_name": inventory_item.name if is_asset_edit_bill else "",
                "sku_code": getattr(inventory_item, "sku_code", ""),
                "item_type": getattr(inventory_item, "item_type", ""),
                "unit": "PIECE" if is_asset_edit_bill else item.unit,
                "quantity": item.quantity,
                "purchase_price": item.purchase_price,
                "gst_percent": item.gst_percent,
            })

        formset = SimplePurchaseItemFormSet(
            prefix="items",
            initial=item_initial,
        )

    return render(request, "core/purchase_entry_form.html", {
        "form": form,
        "formset": formset,
        "title": f"View/Edit Purchase Entry - {bill.bill_number}",
        "editing_bill": bill,
    })


# ==========================================================
# FIX ASSET PURCHASE EDIT SAVE
# Asset edit can save even if item comes as existing item
# ==========================================================

from django.core.exceptions import ValidationError
from decimal import Decimal


def get_or_create_simple_inventory_item(row_data, purchase_type):
    selected_item = row_data.get("item")

    # IMPORTANT FIX:
    # In edit mode, asset item comes as selected existing item.
    # So return selected item directly.
    if selected_item:
        return selected_item

    item_name = (row_data.get("new_item_name") or "").strip()
    sku_code = (row_data.get("sku_code") or "").strip()
    purchase_price = row_data.get("purchase_price") or Decimal("0.00")

    if not item_name:
        if is_asset_purchase_type(purchase_type):
            raise ValidationError("Asset Name is required.")
        raise ValidationError("Item Name is required.")

    if is_asset_purchase_type(purchase_type):
        item_type = get_inventory_fallback_item_type()
        unit = "PIECE"
    else:
        item_type = row_data.get("item_type") or "RAW_MATERIAL"
        unit = row_data.get("unit") or "PIECE"

    existing = InventoryItem.objects.filter(
        name__iexact=item_name,
        purchase_type=purchase_type,
    ).first()

    if existing:
        return existing

    return InventoryItem.objects.create(
        purchase_type=purchase_type,
        item_type=item_type,
        name=item_name,
        sku_code=sku_code,
        stock_unit=unit,
        stock_quantity=Decimal("0.000"),
        average_purchase_price=purchase_price,
        minimum_stock_alert=Decimal("0.000"),
        is_active=True,
    )


def create_assets_from_purchase_row(row, bill, request_user):
    selected_item = row.get("item")

    asset_name = (row.get("new_item_name") or "").strip()

    # IMPORTANT FIX:
    # During edit, selected existing item can hold asset name.
    if not asset_name and selected_item:
        asset_name = selected_item.name

    quantity = row.get("quantity") or Decimal("0")
    purchase_price = row.get("purchase_price") or Decimal("0.00")
    gst_percent = row.get("gst_percent") or Decimal("0.00")

    if not asset_name:
        raise ValidationError("Asset name is required.")

    if quantity <= 0:
        raise ValidationError(f"{asset_name}: quantity must be greater than zero.")

    if quantity != quantity.to_integral_value():
        raise ValidationError(f"{asset_name}: asset quantity must be whole number like 1, 2, 3.")

    quantity_count = int(quantity)

    category = get_or_create_asset_category()

    gst_amount_per_asset = (purchase_price * gst_percent) / Decimal("100.00")
    purchase_value_per_asset = purchase_price + gst_amount_per_asset

    created_count = 0

    for i in range(quantity_count):
        asset_code = generate_purchase_asset_code(category)

        AssetItem.objects.create(
            category=category,
            asset_code=asset_code,
            name=asset_name,
            brand="",
            model_number="",
            serial_number="",
            description=f"Created from purchase bill {bill.bill_number}",
            purchase_date=bill.bill_date,
            purchase_value=purchase_value_per_asset,
            status="AVAILABLE",
            created_by=request_user,
        )

        created_count += 1

    return created_count

# ==========================================================
# SMART PRODUCTION FORMULA - VIEWS
# ==========================================================

from decimal import Decimal, ROUND_HALF_UP

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render


def smart_decimal(value, default="0.000"):
    try:
        if value is None or value == "":
            return Decimal(default)
        return Decimal(str(value))
    except Exception:
        return Decimal(default)


def q2(value):
    return smart_decimal(value, "0.00").quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP
    )


def q3(value):
    return smart_decimal(value, "0.000").quantize(
        Decimal("0.001"),
        rounding=ROUND_HALF_UP
    )


def smart_production_allowed(user):
    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in [
        "ADMIN",
        "INVENTORY_MANAGER",
    ]


def smart_production_check(request):
    if not smart_production_allowed(request.user):
        messages.error(
            request,
            "Only Admin or Inventory Manager can access production."
        )
        return False

    return True


def inventory_rate(item):
    return q2(getattr(item, "average_purchase_price", 0) or 0)


def reduce_inventory_item_stock(item, qty):
    qty = q3(qty)
    current = q3(getattr(item, "stock_quantity", 0) or 0)

    if current < qty:
        raise ValidationError(
            f"{item.name} stock not enough. Available {current}, Required {qty}"
        )

    item.stock_quantity = current - qty
    item.save(update_fields=["stock_quantity"])


def add_product_pack_boxes(pack, boxes):
    boxes = int(boxes or 0)

    if boxes <= 0:
        return

    pack.stock_boxes = int(pack.stock_boxes or 0) + boxes
    pack.save(update_fields=["stock_boxes"])


def get_formula_initial_raw(formula):
    rows = []

    for item in formula.raw_materials.select_related("inventory_item").all():
        rows.append({
            "inventory_item": item.inventory_item,
            "quantity_required": item.quantity_required,
            "unit": item.unit,
        })

    if not rows:
        rows.append({})

    return rows


def get_formula_initial_pack(formula):
    rows = []

    for item in formula.pack_materials.select_related(
        "product_pack",
        "inventory_item"
    ).all():
        rows.append({
            "product_pack": item.product_pack,
            "inventory_item": item.inventory_item,
            "usage_basis": item.usage_basis,
            "quantity_required": item.quantity_required,
            "unit": item.unit,
        })

    if not rows:
        rows.append({})

    return rows


def get_formula_screen_packing_materials(formula=None):
    """
    Shows all packing-like materials in quick checkbox list and validates them in form.
    It includes:
    - item_type = PACKING_MATERIAL
    - common packing names like Bottle, Cap, Label, Outer Box, Box, Carton, Pouch, Bag
    - existing formula packing items while editing
    """

    packing_name_filter = (
        Q(name__icontains="bottle") |
        Q(name__icontains="cap") |
        Q(name__icontains="label") |
        Q(name__icontains="outer") |
        Q(name__icontains="box") |
        Q(name__icontains="carton") |
        Q(name__icontains="pouch") |
        Q(name__icontains="bag") |
        Q(name__icontains="sticker")
    )

    qs = InventoryItem.objects.filter(
        Q(is_active=True, item_type="PACKING_MATERIAL") |
        Q(is_active=True) & packing_name_filter
    )

    if formula:
        existing_ids = formula.pack_materials.values_list(
            "inventory_item_id",
            flat=True
        )

        qs = InventoryItem.objects.filter(
            Q(id__in=existing_ids) |
            Q(is_active=True, item_type="PACKING_MATERIAL") |
            Q(is_active=True) & packing_name_filter
        )

    return qs.distinct().order_by("name")
    

def product_formula_list(request):
    if not smart_production_check(request):
        return redirect("dashboard")

    formulas = ProductProductionFormula.objects.select_related(
        "product"
    ).prefetch_related(
        "raw_materials",
        "pack_materials"
    ).order_by("product__name")

    return render(request, "core/product_formula_list.html", {
        "formulas": formulas,
    })


def product_formula_create(request):
    if not smart_production_check(request):
        return redirect("dashboard")

    selected_product = None
    packing_materials = get_formula_screen_packing_materials()

    if request.method == "POST":
        selected_product_id = request.POST.get("product")

        if selected_product_id:
            selected_product = Product.objects.filter(
                id=selected_product_id,
                is_active=True
            ).first()

        form = ProductFormulaForm(request.POST)

        raw_formset = ProductFormulaRawMaterialFormSet(
            request.POST,
            prefix="raw"
        )

        pack_formset = ProductFormulaPackMaterialFormSet(
            request.POST,
            prefix="pack",
            form_kwargs={
                "product": selected_product,
                "packing_material_queryset": packing_materials,
            }
        )

        if form.is_valid() and raw_formset.is_valid() and pack_formset.is_valid():
            try:
                with transaction.atomic():
                    formula = form.save(commit=False)
                    formula.created_by = request.user
                    formula.save()

                    save_formula_rows(formula, raw_formset, pack_formset)

                messages.success(request, "Production formula created successfully.")
                return redirect("product_formula_list")

            except ValidationError as error:
                messages.error(request, error)

            except Exception as error:
                messages.error(request, f"Formula could not be saved: {error}")

    else:
        selected_product_id = request.GET.get("product_id")

        if selected_product_id:
            selected_product = Product.objects.filter(
                id=selected_product_id,
                is_active=True
            ).first()

        form = ProductFormulaForm(
            initial={"product": selected_product.id} if selected_product else None
        )

        raw_formset = ProductFormulaRawMaterialFormSet(
            prefix="raw"
        )

        pack_formset = ProductFormulaPackMaterialFormSet(
            prefix="pack",
            form_kwargs={
                "product": selected_product,
                "packing_material_queryset": packing_materials,
            }
        )

    return render(request, "core/product_formula_form.html", {
        "form": form,
        "raw_formset": raw_formset,
        "pack_formset": pack_formset,
        "title": "Create Production Formula",
        "selected_product": selected_product,
        "packing_materials": packing_materials,
    })


def product_formula_edit(request, formula_id):
    if not smart_production_check(request):
        return redirect("dashboard")

    formula = get_object_or_404(
        ProductProductionFormula.objects.select_related("product"),
        id=formula_id
    )

    selected_product = formula.product
    packing_materials = get_formula_screen_packing_materials(formula)

    if request.method == "POST":
        selected_product_id = request.POST.get("product") or formula.product_id

        selected_product = Product.objects.filter(
            id=selected_product_id,
            is_active=True
        ).first() or formula.product

        form = ProductFormulaForm(
            request.POST,
            instance=formula
        )

        raw_formset = ProductFormulaRawMaterialFormSet(
            request.POST,
            prefix="raw"
        )

        pack_formset = ProductFormulaPackMaterialFormSet(
            request.POST,
            prefix="pack",
            form_kwargs={
                "product": selected_product,
                "packing_material_queryset": packing_materials,
            }
        )

        if form.is_valid() and raw_formset.is_valid() and pack_formset.is_valid():
            try:
                with transaction.atomic():
                    formula = form.save()
                    save_formula_rows(formula, raw_formset, pack_formset)

                messages.success(request, "Production formula updated successfully.")
                return redirect("product_formula_list")

            except ValidationError as error:
                messages.error(request, error)

            except Exception as error:
                messages.error(request, f"Formula could not be updated: {error}")

    else:
        form = ProductFormulaForm(instance=formula)

        raw_formset = ProductFormulaRawMaterialFormSet(
            prefix="raw",
            initial=get_formula_initial_raw(formula)
        )

        pack_formset = ProductFormulaPackMaterialFormSet(
            prefix="pack",
            initial=get_formula_initial_pack(formula),
            form_kwargs={
                "product": selected_product,
                "packing_material_queryset": packing_materials,
            }
        )

    return render(request, "core/product_formula_form.html", {
        "form": form,
        "raw_formset": raw_formset,
        "pack_formset": pack_formset,
        "title": f"Edit Formula - {formula.product.name}",
        "formula": formula,
        "selected_product": selected_product,
        "packing_materials": packing_materials,
    })


def save_formula_rows(formula, raw_formset, pack_formset):
    raw_rows = []
    pack_rows = []

    for row_form in raw_formset:
        data = row_form.cleaned_data

        if data.get("DELETE"):
            continue

        inventory_item = data.get("inventory_item")
        qty = data.get("quantity_required")

        if not inventory_item and not qty:
            continue

        if not inventory_item:
            raise ValidationError("Raw material item is required.")

        if not qty or qty <= 0:
            raise ValidationError(
                f"{inventory_item.name}: raw material quantity is required."
            )

        raw_rows.append(data)

    for row_form in pack_formset:
        data = row_form.cleaned_data

        if data.get("DELETE"):
            continue

        product_pack = data.get("product_pack")
        inventory_item = data.get("inventory_item")
        qty = data.get("quantity_required")

        if not product_pack and not inventory_item and not qty:
            continue

        if not product_pack:
            raise ValidationError("Product pack is required in packing material row.")

        if not inventory_item:
            raise ValidationError("Packing material item is required.")

        if not qty or qty <= 0:
            raise ValidationError(
                f"{inventory_item.name}: packing material quantity is required."
            )

        pack_rows.append(data)

    if not raw_rows:
        raise ValidationError("Please add at least one raw material.")

    formula.raw_materials.all().delete()
    formula.pack_materials.all().delete()

    for row in raw_rows:
        ProductFormulaRawMaterial.objects.create(
            formula=formula,
            inventory_item=row["inventory_item"],
            quantity_required=q3(row["quantity_required"]),
            unit=row.get("unit") or getattr(row["inventory_item"], "stock_unit", "KG"),
        )

    for row in pack_rows:
        ProductFormulaPackMaterial.objects.create(
            formula=formula,
            product_pack=row["product_pack"],
            inventory_item=row["inventory_item"],
            usage_basis=row.get("usage_basis") or "PER_UNIT",
            quantity_required=q3(row["quantity_required"]),
            unit=row.get("unit") or getattr(row["inventory_item"], "stock_unit", "PIECE"),
        )


def clean_pack_number(value):
    try:
        value = Decimal(str(value))
        if value == value.to_integral_value():
            return str(int(value))
        return str(value)
    except Exception:
        return str(value)


def plural_pack_type(pack):
    try:
        name = pack.get_packing_type_display()
    except Exception:
        name = getattr(pack, "packing_type", "")

    name = str(name or "").strip().title()

    if name.lower() == "box":
        return "Boxes"

    if name.lower() == "piece":
        return "Pieces"

    if name.lower() == "carton":
        return "Cartons"

    if name.lower() == "bag":
        return "Bags"

    if name.lower() == "packet":
        return "Packets"

    if name.lower() == "bottle":
        return "Bottles"

    if name.endswith("s"):
        return name

    return name + "s"


def smart_pack_label(pack):
    pack_size = clean_pack_number(pack.pack_size)
    pack_unit = pack.unit
    units_per_box = pack.units_per_box or 0
    packing_name = plural_pack_type(pack)

    return f"{pack_size} {pack_unit} | {units_per_box} {packing_name} in 1 Box"


def formula_json_data():
    formulas = ProductProductionFormula.objects.select_related(
        "product"
    ).prefetch_related(
        "raw_materials__inventory_item",
        "pack_materials__inventory_item",
        "pack_materials__product_pack",
    ).filter(
        is_active=True
    )

    data = {}

    for formula in formulas:
        product_packs = ProductPackSize.objects.filter(
            product=formula.product,
            is_active=True
        ).order_by("pack_size")

        raw_rows = []
        pack_rows = []
        pack_options = []

        for raw in formula.raw_materials.all():
            item = raw.inventory_item
            stock = q3(getattr(item, "stock_quantity", 0) or 0)
            rate = inventory_rate(item)

            raw_rows.append({
                "item_id": item.id,
                "item_name": item.name,
                "base_qty": str(q3(raw.quantity_required)),
                "unit": raw.unit,
                "stock": str(stock),
                "rate": str(rate),
            })

        for pack in product_packs:
            pack_options.append({
                "pack_id": pack.id,
                "label": smart_pack_label(pack),
                "pack_size": str(pack.pack_size),
                "unit": pack.unit,
                "units_per_box": pack.units_per_box,
                "current_stock_boxes": pack.stock_boxes,
            })

        for mat in formula.pack_materials.all():
            item = mat.inventory_item
            pack = mat.product_pack
            stock = q3(getattr(item, "stock_quantity", 0) or 0)
            rate = inventory_rate(item)

            pack_rows.append({
                "pack_id": pack.id,
                "pack_label": f"{pack.display_pack} | {pack.units_per_box} in 1 {pack.get_packing_type_display()}",
                "item_id": item.id,
                "item_name": item.name,
                "usage_basis": mat.usage_basis,
                "qty_required": str(q3(mat.quantity_required)),
                "unit": mat.unit,
                "stock": str(stock),
                "rate": str(rate),
            })

        data[str(formula.id)] = {
            "formula_id": formula.id,
            "product_name": formula.product.name,
            "base_output_qty": str(q3(formula.base_output_qty)),
            "base_output_unit": formula.base_output_unit,
            "raw_materials": raw_rows,
            "pack_materials": pack_rows,
            "packs": pack_options,
        }

    return data


def smart_amount(value):
    try:
        return q2(Decimal(str(value or "0")))
    except Exception:
        return Decimal("0.00")


def normalize_pack_output_qty(pack, boxes, target_unit):
    """
    Converts packed output to formula base unit.
    Example:
    250 ML × 40 bottles × 1 box = 10000 ML = 10 LITRE
    """
    boxes = Decimal(str(boxes or 0))
    pack_size = Decimal(str(pack.pack_size or 0))
    units_per_box = Decimal(str(pack.units_per_box or 0))

    qty = pack_size * units_per_box * boxes

    source_unit = str(pack.unit or "").upper()
    target_unit = str(target_unit or "").upper()

    if source_unit == target_unit:
        return q3(qty)

    if source_unit == "ML" and target_unit == "LITRE":
        return q3(qty / Decimal("1000"))

    if source_unit == "LITRE" and target_unit == "ML":
        return q3(qty * Decimal("1000"))

    if source_unit == "GRAM" and target_unit == "KG":
        return q3(qty / Decimal("1000"))

    if source_unit == "KG" and target_unit == "GRAM":
        return q3(qty * Decimal("1000"))

    return q3(qty)

def calculate_smart_production(formula, output_qty, pack_box_map, labour_cost=0, other_cost=0):
    output_qty = q3(output_qty)
    labour_cost = smart_amount(labour_cost)
    other_cost = smart_amount(other_cost)

    if output_qty <= 0:
        raise ValidationError("Output quantity must be greater than zero.")

    factor = output_qty / q3(formula.base_output_qty)

    required_raw = []
    required_pack = []
    shortages = []
    output_pack_rows = []

    total_raw_cost = Decimal("0.00")
    total_pack_cost = Decimal("0.00")
    total_boxes = 0
    packed_output_qty = Decimal("0.000")

    # Raw materials based on bulk output
    for raw in formula.raw_materials.select_related("inventory_item").all():
        item = raw.inventory_item
        req_qty = q3(raw.quantity_required * factor)
        available = q3(item.stock_quantity or 0)
        rate = inventory_rate(item)
        cost = q2(req_qty * rate)

        if available < req_qty:
            shortages.append({
                "name": item.name,
                "required": req_qty,
                "available": available,
                "unit": raw.unit,
            })

        total_raw_cost += cost

        required_raw.append({
            "item": item,
            "type": "RAW",
            "required_qty": req_qty,
            "available": available,
            "unit": raw.unit,
            "rate": rate,
            "cost": cost,
        })

    # Packing materials based on selected output boxes
    for mat in formula.pack_materials.select_related("inventory_item", "product_pack").all():
        boxes = int(pack_box_map.get(str(mat.product_pack_id), 0) or 0)

        if boxes <= 0:
            continue

        if mat.usage_basis == "PER_BOX":
            req_qty = q3(Decimal(boxes) * mat.quantity_required)
        else:
            req_qty = q3(Decimal(boxes) * Decimal(mat.product_pack.units_per_box) * mat.quantity_required)

        item = mat.inventory_item
        available = q3(item.stock_quantity or 0)
        rate = inventory_rate(item)
        cost = q2(req_qty * rate)

        if available < req_qty:
            shortages.append({
                "name": item.name,
                "required": req_qty,
                "available": available,
                "unit": mat.unit,
            })

        total_pack_cost += cost

        required_pack.append({
            "item": item,
            "type": "PACKING",
            "required_qty": req_qty,
            "available": available,
            "unit": mat.unit,
            "rate": rate,
            "cost": cost,
            "product_pack": mat.product_pack,
        })

    # Output packing match calculation
    product_packs = ProductPackSize.objects.filter(
        product=formula.product,
        is_active=True
    )

    for pack in product_packs:
        boxes = int(pack_box_map.get(str(pack.id), 0) or 0)

        if boxes <= 0:
            continue

        box_output_qty = normalize_pack_output_qty(
            pack=pack,
            boxes=boxes,
            target_unit=formula.base_output_unit
        )

        packed_output_qty += box_output_qty
        total_boxes += boxes

        output_pack_rows.append({
            "pack": pack,
            "boxes": boxes,
            "output_qty": box_output_qty,
            "unit": formula.base_output_unit,
        })

    packed_output_qty = q3(packed_output_qty)
    output_difference = q3(packed_output_qty - output_qty)
    output_matched = output_difference == Decimal("0.000")

    total_cost = q2(total_raw_cost + total_pack_cost + labour_cost + other_cost)

    cost_per_output_unit = Decimal("0.00")
    if output_qty > 0:
        cost_per_output_unit = q2(total_cost / output_qty)

    cost_per_box = Decimal("0.00")
    if total_boxes > 0:
        cost_per_box = q2(total_cost / Decimal(total_boxes))

    # distribute cost by output quantity
    for row in output_pack_rows:
        row_total_cost = q2(row["output_qty"] * cost_per_output_unit)
        row["total_cost"] = row_total_cost

        if row["boxes"] > 0:
            row["cost_per_box"] = q2(row_total_cost / Decimal(row["boxes"]))
        else:
            row["cost_per_box"] = Decimal("0.00")

    return {
        "required_raw": required_raw,
        "required_pack": required_pack,
        "shortages": shortages,

        "output_qty": output_qty,
        "packed_output_qty": packed_output_qty,
        "output_difference": output_difference,
        "output_matched": output_matched,
        "output_unit": formula.base_output_unit,
        "output_pack_rows": output_pack_rows,

        "total_raw_cost": q2(total_raw_cost),
        "total_pack_cost": q2(total_pack_cost),
        "labour_cost": labour_cost,
        "other_cost": other_cost,
        "total_cost": total_cost,

        "total_boxes": total_boxes,
        "cost_per_output_unit": cost_per_output_unit,
        "cost_per_box": cost_per_box,
    }



def smart_production_create(request):
    if not smart_production_check(request):
        return redirect("dashboard")

    formulas_json = formula_json_data()

    preview_result = None

    if request.method == "POST":
        form = SmartProductionRunForm(request.POST)

        if form.is_valid():
            formula = form.cleaned_data["formula"]
            output_qty = form.cleaned_data["output_qty"]
            labour_cost = request.POST.get("labour_cost") or "0"
            other_cost = request.POST.get("other_cost") or "0"

            pack_box_map = {}

            for pack in ProductPackSize.objects.filter(product=formula.product, is_active=True):
                key = f"pack_{pack.id}_boxes"
                pack_box_map[str(pack.id)] = int(request.POST.get(key) or 0)

            try:
                calc = calculate_smart_production(
                    formula=formula,
                    output_qty=output_qty,
                    pack_box_map=pack_box_map,
                    labour_cost=labour_cost,
                    other_cost=other_cost,
                )
                preview_result = calc

                if calc["shortages"]:
                    messages.error(request, "Stock is not enough. Please reduce production quantity or purchase missing items.")
                    raise ValidationError("Insufficient stock.")

                if calc["total_boxes"] <= 0:
                    messages.error(request, "Please enter at least one output pack box quantity.")
                    raise ValidationError("Output pack required.")
                if not calc["output_matched"]:
                    messages.error(
                        request,
                        f"Bulk output and packed output not matching. "
                        f"Bulk: {calc['output_qty']} {calc['output_unit']}, "
                        f"Packed: {calc['packed_output_qty']} {calc['output_unit']}."
                    )
                    raise ValidationError("Packed output mismatch.")

                with transaction.atomic():
                    run = ProductProductionRun.objects.create(
                        product=formula.product,
                        formula=formula,
                        output_qty=q3(output_qty),
                        output_unit=formula.base_output_unit,
                        production_date=form.cleaned_data["production_date"],
                        total_raw_cost=calc["total_raw_cost"],
                        total_packing_cost=calc["total_pack_cost"],
                        labour_cost=calc["labour_cost"],
                        other_cost=calc["other_cost"],
                        total_cost=calc["total_cost"],
                        note=form.cleaned_data.get("note"),
                        created_by=request.user,
                    )

                    for row in calc["required_raw"] + calc["required_pack"]:
                        reduce_inventory_item_stock(row["item"], row["required_qty"])

                        ProductProductionRunMaterial.objects.create(
                            run=run,
                            inventory_item=row["item"],
                            material_type=row["type"],
                            required_qty=row["required_qty"],
                            unit=row["unit"],
                            available_before=row["available"],
                            rate=row["rate"],
                            cost=row["cost"],
                        )

                    for pack_id, boxes in pack_box_map.items():
                        boxes = int(boxes or 0)

                        if boxes <= 0:
                            continue

                        pack = ProductPackSize.objects.get(id=pack_id, product=formula.product)

                        add_product_pack_boxes(pack, boxes)

                        output_row = next(
                            (row for row in calc["output_pack_rows"] if row["pack"].id == pack.id),
                            None
                        )

                        ProductProductionRunOutputPack.objects.create(
                            run=run,
                            product_pack=pack,
                            output_boxes=boxes,
                            units_per_box=pack.units_per_box,
                            cost_per_box=output_row["cost_per_box"] if output_row else calc["cost_per_box"],
                            total_cost=output_row["total_cost"] if output_row else q2(calc["cost_per_box"] * Decimal(boxes)),
                        )

                messages.success(request, "Production created successfully. Raw/Packing stock reduced and final product stock increased.")
                return redirect("smart_production_run_detail", run_id=run.id)

            except ValidationError:
                pass

            except Exception as error:
                messages.error(request, f"Production could not be created: {error}")

    else:
        form = SmartProductionRunForm(initial={
            "production_date": timezone.localdate()
        })

    return render(request, "core/smart_production_create.html", {
        "form": form,
        "formulas_json": json.dumps(formulas_json),
        "preview_result": preview_result,
        "title": "Smart Production / Packaging",
    })




def smart_production_run_list(request):
    if not smart_production_check(request):
        return redirect("dashboard")

    runs = ProductProductionRun.objects.select_related(
        "product",
        "created_by",
    ).prefetch_related(
        "output_packs",
    ).annotate(
        total_boxes=Sum("output_packs__output_boxes")
    ).order_by("-production_date", "-id")

    return render(request, "core/smart_production_run_list.html", {
        "runs": runs,
    })



def detail_decimal(value, default="0.000"):
    try:
        return Decimal(str(value or default))
    except Exception:
        return Decimal(default)


def detail_q3(value):
    return detail_decimal(value).quantize(Decimal("0.001"))


def normalize_detail_unit(unit):
    unit = str(unit or "").strip().upper()

    if unit in ["LTR", "LT", "L", "LITER"]:
        return "LITRE"

    if unit in ["G", "GM"]:
        return "GRAM"

    return unit


def detail_clean_number(value):
    value = detail_decimal(value)

    if value == value.to_integral_value():
        return str(int(value))

    return str(value)


def detail_plural_packing_name(pack):
    try:
        name = pack.get_packing_type_display()
    except Exception:
        name = getattr(pack, "packing_type", "")

    name = str(name or "").strip().title()

    if name.lower() == "bottle":
        return "Bottles"

    if name.lower() == "bag":
        return "Bags"

    if name.lower() == "packet":
        return "Packets"

    if name.lower() == "piece":
        return "Pieces"

    if name.lower() == "carton":
        return "Cartons"

    if name.lower() == "box":
        return "Boxes"

    if name.endswith("s"):
        return name

    return name + "s"


def detail_pack_label(pack, units_per_box=None):
    pack_size = detail_clean_number(pack.pack_size)
    unit = str(pack.unit or "").upper()
    units = units_per_box if units_per_box is not None else pack.units_per_box
    units = detail_clean_number(units)
    packing_name = detail_plural_packing_name(pack)

    return f"{pack_size} {unit} | {units} {packing_name} in 1 Box"


def detail_pack_output_qty(pack, boxes, target_unit):
    boxes = detail_decimal(boxes)
    pack_size = detail_decimal(pack.pack_size)
    units_per_box = detail_decimal(pack.units_per_box)

    total_qty = pack_size * units_per_box * boxes

    source_unit = normalize_detail_unit(pack.unit)
    target_unit = normalize_detail_unit(target_unit)

    if source_unit == target_unit:
        return detail_q3(total_qty)

    if source_unit == "ML" and target_unit == "LITRE":
        return detail_q3(total_qty / Decimal("1000"))

    if source_unit == "LITRE" and target_unit == "ML":
        return detail_q3(total_qty * Decimal("1000"))

    if source_unit == "GRAM" and target_unit == "KG":
        return detail_q3(total_qty / Decimal("1000"))

    if source_unit == "KG" and target_unit == "GRAM":
        return detail_q3(total_qty * Decimal("1000"))

    return detail_q3(total_qty)


def smart_production_run_detail(request, run_id):
    if not smart_production_check(request):
        return redirect("dashboard")

    run = get_object_or_404(
        ProductProductionRun.objects.select_related(
            "product",
            "formula",
            "created_by",
        ).prefetch_related(
            "output_packs__product_pack",
            "materials__inventory_item",
        ),
        id=run_id
    )

    output_rows = []
    packed_output_qty = Decimal("0.000")

    for output_pack in run.output_packs.select_related("product_pack").all():
        pack = output_pack.product_pack

        packed_qty = detail_pack_output_qty(
            pack=pack,
            boxes=output_pack.output_boxes,
            target_unit=run.output_unit,
        )

        packed_output_qty += packed_qty

        output_rows.append({
            "row": output_pack,
            "pack_label": detail_pack_label(pack, output_pack.units_per_box),
            "packed_output_qty": packed_qty,
        })

    packed_output_qty = detail_q3(packed_output_qty)
    bulk_output_qty = detail_q3(run.output_qty)
    output_difference = detail_q3(packed_output_qty - bulk_output_qty)
    output_matched = output_difference == Decimal("0.000")

    return render(request, "core/smart_production_run_detail.html", {
        "run": run,
        "output_rows": output_rows,
        "bulk_output_qty": bulk_output_qty,
        "packed_output_qty": packed_output_qty,
        "output_difference": output_difference,
        "output_matched": output_matched,
    })


# ==========================================================
# FINAL FIX - PURCHASE NEW ITEM FULL SAVE
# Saves New Item Name, SKU, Item Type, Unit, Qty, Rate, GST
# ==========================================================

from decimal import Decimal
from django.core.exceptions import ValidationError


def clean_purchase_text(value):
    return str(value or "").strip()


def get_purchase_decimal(value, default="0.00"):
    if value is None or value == "":
        return Decimal(default)

    return value


def update_inventory_item_from_purchase(item, row_data, purchase_type):
    sku_code = clean_purchase_text(row_data.get("sku_code"))
    item_type = row_data.get("item_type") or getattr(item, "item_type", "RAW_MATERIAL")
    unit = row_data.get("unit") or getattr(item, "stock_unit", "PIECE")
    purchase_price = get_purchase_decimal(row_data.get("purchase_price"), "0.00")

    update_fields = []

    if purchase_type and getattr(item, "purchase_type_id", None) != purchase_type.id:
        item.purchase_type = purchase_type
        update_fields.append("purchase_type")

    if sku_code:
        item.sku_code = sku_code
        update_fields.append("sku_code")

    if item_type:
        item.item_type = item_type
        update_fields.append("item_type")

    if unit:
        item.stock_unit = unit
        update_fields.append("stock_unit")

    item.average_purchase_price = purchase_price
    update_fields.append("average_purchase_price")

    if update_fields:
        item.save(update_fields=list(set(update_fields)))

    return item


def get_or_create_simple_inventory_item(row_data, purchase_type):
    selected_item = row_data.get("item")

    item_name = clean_purchase_text(row_data.get("new_item_name"))
    sku_code = clean_purchase_text(row_data.get("sku_code"))
    item_type = row_data.get("item_type") or "RAW_MATERIAL"
    unit = row_data.get("unit") or "PIECE"
    purchase_price = get_purchase_decimal(row_data.get("purchase_price"), "0.00")

    if selected_item:
        return update_inventory_item_from_purchase(
            item=selected_item,
            row_data=row_data,
            purchase_type=purchase_type,
        )

    if not item_name:
        if is_asset_purchase_type(purchase_type):
            raise ValidationError("Asset Name is required.")
        raise ValidationError("New Item Name is required.")

    if is_asset_purchase_type(purchase_type):
        item_type = get_inventory_fallback_item_type()
        unit = "PIECE"

    existing = InventoryItem.objects.filter(
        name__iexact=item_name,
        purchase_type=purchase_type,
    ).first()

    if not existing:
        existing = InventoryItem.objects.filter(
            name__iexact=item_name
        ).first()

    if existing:
        return update_inventory_item_from_purchase(
            item=existing,
            row_data={
                "sku_code": sku_code,
                "item_type": item_type,
                "unit": unit,
                "purchase_price": purchase_price,
            },
            purchase_type=purchase_type,
        )

    return InventoryItem.objects.create(
        purchase_type=purchase_type,
        item_type=item_type,
        name=item_name,
        sku_code=sku_code,
        stock_unit=unit,
        stock_quantity=Decimal("0.000"),
        average_purchase_price=purchase_price,
        minimum_stock_alert=Decimal("0.000"),
        is_active=True,
    )


def get_valid_purchase_item_rows(formset):
    valid_rows = []

    for item_form in formset:
        data = item_form.cleaned_data

        if data.get("DELETE"):
            continue

        item = data.get("item")
        new_item_name = clean_purchase_text(data.get("new_item_name"))
        quantity = data.get("quantity")
        purchase_price = data.get("purchase_price")

        if not item and not new_item_name and not quantity and purchase_price in [None, ""]:
            continue

        if not item and not new_item_name:
            raise ValidationError("Select item or enter new item name.")

        if not quantity or quantity <= 0:
            raise ValidationError("Quantity is required.")

        if purchase_price is None or purchase_price < 0:
            raise ValidationError("Purchase rate is required.")

        valid_rows.append({
            "item": item,
            "new_item_name": new_item_name,
            "sku_code": clean_purchase_text(data.get("sku_code")),
            "item_type": data.get("item_type") or "RAW_MATERIAL",
            "unit": data.get("unit") or "PIECE",
            "quantity": quantity,
            "purchase_price": purchase_price,
            "gst_percent": data.get("gst_percent") or Decimal("0.00"),
        })

    return valid_rows

# ==========================================================
# CLEAN STOCK USAGE HISTORY
# Only:
# 1. Purchase Entry
# 2. Purchase Return
# 3. Smart Production Formula Usage
# ==========================================================

from decimal import Decimal
from django.db.models import Q, Sum
from django.shortcuts import render, redirect


def stock_dec(value):
    if value in [None, ""]:
        return Decimal("0.000")
    return Decimal(str(value))


def amount_dec(value):
    if value in [None, ""]:
        return Decimal("0.00")
    return Decimal(str(value))


def purchase_stock_history(request):
    if not purchase_permission_check(request):
        return redirect("dashboard")

    search = (request.GET.get("q") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()

    items = InventoryItem.objects.filter(
        is_active=True
    ).order_by("name")

    if search:
        items = items.filter(
            Q(name__icontains=search) |
            Q(sku_code__icontains=search) |
            Q(purchase_type__name__icontains=search)
        )

    item_ids = list(items.values_list("id", flat=True))

    summary_rows = []
    movements = []

    for item in items:
        purchased_qty = PurchaseBillItem.objects.filter(
            inventory_item=item
        ).aggregate(total=Sum("quantity"))["total"] or Decimal("0.000")

        returned_qty = PurchaseReturnItem.objects.filter(
            inventory_item=item
        ).aggregate(total=Sum("quantity"))["total"] or Decimal("0.000")

        used_qty = ProductProductionRunMaterial.objects.filter(
            inventory_item=item
        ).aggregate(total=Sum("required_qty"))["total"] or Decimal("0.000")

        summary_rows.append({
            "item": item,
            "purchased_qty": purchased_qty,
            "used_qty": used_qty,
            "returned_qty": returned_qty,
            "current_stock": item.stock_quantity or Decimal("0.000"),
        })

    purchase_items = PurchaseBillItem.objects.select_related(
        "bill",
        "bill__party",
        "inventory_item",
    ).filter(
        inventory_item_id__in=item_ids
    )

    if from_date:
        purchase_items = purchase_items.filter(bill__bill_date__gte=from_date)

    if to_date:
        purchase_items = purchase_items.filter(bill__bill_date__lte=to_date)

    for item in purchase_items:
        qty = stock_dec(item.quantity)
        rate = amount_dec(item.purchase_price)
        gst = amount_dec(item.gst_percent)
        taxable = qty * rate
        cost = taxable + ((taxable * gst) / Decimal("100.00"))

        movements.append({
            "date": item.bill.bill_date,
            "type": "PURCHASED",
            "item_name": item.inventory_item.name,
            "sku": item.inventory_item.sku_code or "-",
            "reference": item.bill.bill_number or f"Bill #{item.bill.id}",
            "party": item.bill.party.party_name if item.bill.party else "-",
            "in_qty": qty,
            "out_qty": Decimal("0.000"),
            "unit": item.unit,
            "rate": rate,
            "cost": cost,
            "note": "Purchased stock added",
        })

    return_items = PurchaseReturnItem.objects.select_related(
        "purchase_return",
        "purchase_return__party",
        "inventory_item",
    ).filter(
        inventory_item_id__in=item_ids
    )

    if from_date:
        return_items = return_items.filter(purchase_return__return_date__gte=from_date)

    if to_date:
        return_items = return_items.filter(purchase_return__return_date__lte=to_date)

    for item in return_items:
        qty = stock_dec(item.quantity)
        rate = amount_dec(item.return_price)
        gst = amount_dec(item.gst_percent)
        taxable = qty * rate
        cost = taxable + ((taxable * gst) / Decimal("100.00"))

        movements.append({
            "date": item.purchase_return.return_date,
            "type": "RETURNED",
            "item_name": item.inventory_item.name,
            "sku": item.inventory_item.sku_code or "-",
            "reference": item.purchase_return.debit_note_number or f"Return #{item.purchase_return.id}",
            "party": item.purchase_return.party.party_name if item.purchase_return.party else "-",
            "in_qty": Decimal("0.000"),
            "out_qty": qty,
            "unit": item.unit,
            "rate": rate,
            "cost": cost,
            "note": item.purchase_return.reason or "Purchase return",
        })

    production_items = ProductProductionRunMaterial.objects.select_related(
        "run",
        "run__product",
        "inventory_item",
    ).filter(
        inventory_item_id__in=item_ids
    )

    if from_date:
        production_items = production_items.filter(run__production_date__gte=from_date)

    if to_date:
        production_items = production_items.filter(run__production_date__lte=to_date)

    for item in production_items:
        movements.append({
            "date": item.run.production_date,
            "type": "USED",
            "item_name": item.inventory_item.name,
            "sku": item.inventory_item.sku_code or "-",
            "reference": f"Production #{item.run.id}",
            "party": item.run.product.name if item.run.product else "-",
            "in_qty": Decimal("0.000"),
            "out_qty": item.required_qty or Decimal("0.000"),
            "unit": item.unit,
            "rate": item.rate,
            "cost": item.cost,
            "note": f"{item.get_material_type_display()} used",
        })

    movements.sort(key=lambda row: row["date"], reverse=True)

    total_purchased = sum((row["purchased_qty"] for row in summary_rows), Decimal("0.000"))
    total_used = sum((row["used_qty"] for row in summary_rows), Decimal("0.000"))
    total_returned = sum((row["returned_qty"] for row in summary_rows), Decimal("0.000"))
    total_remaining = sum((row["current_stock"] for row in summary_rows), Decimal("0.000"))

    return render(request, "core/purchase_stock_history.html", {
        "summary_rows": summary_rows,
        "movements": movements,
        "search": search,
        "from_date": from_date,
        "to_date": to_date,
        "total_purchased": total_purchased,
        "total_used": total_used,
        "total_returned": total_returned,
        "total_remaining": total_remaining,
    })

# ================= STRICT ACTION-ONLY PENDING BADGES API =================
# Replace your old pending-action-badges code in views.py with this full code.
# This counts ONLY actions the logged-in role can perform now.
# It does NOT count past/completed orders.

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.urls import reverse
from django.apps import apps


def badge_reverse(name):
    try:
        return reverse(name)
    except Exception:
        return None


def badge_model(name):
    try:
        return apps.get_model("core", name)
    except Exception:
        return None


def badge_count(model_name, **filters):
    Model = badge_model(model_name)

    if not Model:
        return 0

    try:
        return Model.objects.filter(**filters).count()
    except Exception:
        return 0


def badge_add(items, key, title, route_name, count, total_key=None):
    count = int(count or 0)

    if count <= 0:
        return

    url = badge_reverse(route_name)

    if not url:
        return

    items.append({
        "key": key,
        "title": title,
        "url": url,
        "count": count,
        "total_key": total_key or key,
    })


def badge_subordinate_ids(user):
    UserProfile = badge_model("UserProfile")

    if not UserProfile:
        return []

    try:
        return list(
            UserProfile.objects.filter(manager=user).values_list("user_id", flat=True)
        )
    except Exception:
        return []


def get_strict_order_action_counts(user, role):
    """
    DealerOrder actionable statuses from your workflow:

    SALES_APPROVAL               -> concerned Sales Officer can approve/reject
    ASM_APPROVAL                 -> concerned ASM can approve/reject
    RSM_APPROVAL                 -> concerned Regional Manager can approve/reject
    ACCOUNTANT_ORDER_APPROVAL    -> Accountant can approve/reject
    WAREHOUSE_REVIEW             -> Warehouse Manager can review
    ACCOUNTANT_INVOICE_REVIEW    -> Accountant can invoice review
    DISPATCH_PENDING             -> Warehouse Manager can dispatch
    DISPATCHED                   -> Dealer can accept delivery
    """
    DealerOrder = badge_model("DealerOrder")

    if not DealerOrder:
        return {
            "orders_link_count": 0,
            "warehouse_link_count": 0,
            "invoice_link_count": 0,
            "dealer_delivery_count": 0,
        }

    orders_link_count = 0
    warehouse_link_count = 0
    invoice_link_count = 0
    dealer_delivery_count = 0

    try:
        if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
            orders_link_count = DealerOrder.objects.filter(
                status="SALES_APPROVAL",
                concerned_sales_officer=user,
            ).count()

        elif role == "ASM":
            orders_link_count = DealerOrder.objects.filter(
                status="ASM_APPROVAL",
                concerned_asm=user,
            ).count()

        elif role == "REGIONAL_MANAGER":
            orders_link_count = DealerOrder.objects.filter(
                status="RSM_APPROVAL",
                concerned_rsm=user,
            ).count()

        elif role == "ACCOUNTANT":
            # Accountant can perform only these two order actions.
            accountant_order_approval = DealerOrder.objects.filter(
                status="ACCOUNTANT_ORDER_APPROVAL"
            ).count()

            accountant_invoice_review = DealerOrder.objects.filter(
                status="ACCOUNTANT_INVOICE_REVIEW"
            ).count()

            orders_link_count = accountant_order_approval + accountant_invoice_review
            invoice_link_count = accountant_invoice_review

        elif role == "WAREHOUSE_MANAGER":
            warehouse_review = DealerOrder.objects.filter(
                status="WAREHOUSE_REVIEW",
                warehouse_manager=user,
            ).count()

            dispatch_pending = DealerOrder.objects.filter(
                status="DISPATCH_PENDING",
                warehouse_manager=user,
            ).count()

            warehouse_link_count = warehouse_review + dispatch_pending

        elif role == "DEALER":
            dealer_delivery_count = DealerOrder.objects.filter(
                status="DISPATCHED",
                dealer_user=user,
            ).count()

            orders_link_count = dealer_delivery_count

        elif role == "ADMIN" or user.is_superuser:
            admin_approval = DealerOrder.objects.filter(
                status__in=[
                    "SALES_APPROVAL",
                    "ASM_APPROVAL",
                    "RSM_APPROVAL",
                    "ACCOUNTANT_ORDER_APPROVAL",
                ]
            ).count()

            admin_warehouse = DealerOrder.objects.filter(
                status__in=[
                    "WAREHOUSE_REVIEW",
                    "DISPATCH_PENDING",
                ]
            ).count()

            admin_invoice = DealerOrder.objects.filter(
                status="ACCOUNTANT_INVOICE_REVIEW"
            ).count()

            orders_link_count = admin_approval + admin_warehouse + admin_invoice
            warehouse_link_count = admin_warehouse
            invoice_link_count = admin_invoice

    except Exception:
        pass

    return {
        "orders_link_count": orders_link_count,
        "warehouse_link_count": warehouse_link_count,
        "invoice_link_count": invoice_link_count,
        "dealer_delivery_count": dealer_delivery_count,
    }


def get_pending_action_badges_for_user(user):
    profile = getattr(user, "profile", None)
    role = getattr(profile, "role", "")

    items = []
    unique_total_keys = {}

    # ==========================
    # OWN ATTENDANCE ACTION
    # ==========================
    clock_out_count = badge_count(
        "EmployeeAttendance",
        employee=user,
        clock_out_at__isnull=True,
    )

    badge_add(
        items,
        "clock_out_pending",
        "Clock Out Pending",
        "employee_attendance_dashboard",
        clock_out_count,
        "clock_out_pending"
    )

    # ==========================
    # TA/DA ACTIONS
    # ==========================
    subordinate_ids = badge_subordinate_ids(user)

    if user.is_superuser or role == "ADMIN":
        manager_tada_count = badge_count(
            "EmployeeAttendance",
            status="PENDING_MANAGER",
        )
    elif subordinate_ids:
        manager_tada_count = badge_count(
            "EmployeeAttendance",
            employee_id__in=subordinate_ids,
            status="PENDING_MANAGER",
        )
    else:
        manager_tada_count = 0

    badge_add(
        items,
        "manager_tada_approvals",
        "TA/DA Manager Approvals",
        "manager_attendance_approvals",
        manager_tada_count,
        "manager_tada_approvals"
    )

    if user.is_superuser or role in ["ADMIN", "HR"]:
        hr_tada_count = badge_count(
            "EmployeeAttendance",
            status="MANAGER_APPROVED",
        )

        badge_add(
            items,
            "hr_tada_claims",
            "HR TA/DA Claims",
            "hr_attendance_claims",
            hr_tada_count,
            "hr_tada_claims"
        )

    if user.is_superuser or role in ["ADMIN", "ACCOUNTANT"]:
        accountant_tada_count = badge_count(
            "EmployeeAttendance",
            status="HR_APPROVED",
        )

        badge_add(
            items,
            "accountant_tada_release",
            "Release TA/DA Funds",
            "accountant_attendance_claims",
            accountant_tada_count,
            "accountant_tada_release"
        )

    # ==========================
    # LEAVE ACTIONS
    # ==========================
    if user.is_superuser or role == "ADMIN":
        leave_manager_count = badge_count(
            "EmployeeLeaveRequest",
            status="PENDING",
        )
    elif subordinate_ids:
        leave_manager_count = badge_count(
            "EmployeeLeaveRequest",
            employee_id__in=subordinate_ids,
            status="PENDING",
        )
    else:
        leave_manager_count = 0

    badge_add(
        items,
        "manager_leave_approvals",
        "Leave Approvals",
        "manager_leave_approvals",
        leave_manager_count,
        "manager_leave_approvals"
    )

    # ==========================
    # FARMER MEET ACTIONS
    # ==========================
    farmer_count = 0

    if role in ["SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        farmer_count = badge_count(
            "FarmerMeetRequest",
            sales_officer=user,
            approval_status="PENDING_SALES_OFFICER",
        )

    elif role == "ASM":
        farmer_count = badge_count(
            "FarmerMeetRequest",
            asm=user,
            approval_status="PENDING_ASM",
        )

    elif role == "REGIONAL_MANAGER":
        farmer_count = badge_count(
            "FarmerMeetRequest",
            regional_manager=user,
            approval_status="PENDING_REGIONAL_MANAGER",
        )

    elif role == "STATE_HEAD":
        farmer_count = badge_count(
            "FarmerMeetRequest",
            state_head=user,
            approval_status="PENDING_STATE_HEAD",
        )

    elif user.is_superuser or role == "ADMIN":
        FarmerMeetRequest = badge_model("FarmerMeetRequest")
        if FarmerMeetRequest:
            try:
                farmer_count = FarmerMeetRequest.objects.filter(
                    approval_status__in=[
                        "PENDING_SALES_OFFICER",
                        "PENDING_ASM",
                        "PENDING_REGIONAL_MANAGER",
                        "PENDING_STATE_HEAD",
                    ]
                ).count()
            except Exception:
                farmer_count = 0

    badge_add(
        items,
        "farmer_meet_approvals",
        "Farmer Meet Approvals",
        "farmer_meet_list",
        farmer_count,
        "farmer_meet_approvals"
    )

    # ==========================
    # DEALER APPROVAL ACTIONS
    # ==========================
    dealer_approval_count = 0

    if user.is_superuser or role == "ADMIN":
        Dealer = badge_model("Dealer")
        if Dealer:
            try:
                dealer_approval_count = Dealer.objects.exclude(
                    approval_status__in=["APPROVED", "REJECTED"]
                ).count()
            except Exception:
                dealer_approval_count = 0

    elif role == "ASM":
        dealer_approval_count = badge_count(
            "Dealer",
            concerned_asm=user,
            approval_status="PENDING_ASM",
        )

    elif role == "REGIONAL_MANAGER":
        dealer_approval_count = badge_count(
            "Dealer",
            forwarded_regional_manager=user,
            approval_status="FORWARDED_RM",
        )

    elif role == "STATE_HEAD":
        dealer_approval_count = badge_count(
            "Dealer",
            forwarded_state_head=user,
            approval_status="FORWARDED_STATE_HEAD",
        )

    badge_add(
        items,
        "dealer_approvals",
        "Dealer Approvals",
        "dealer_approval_list",
        dealer_approval_count,
        "dealer_approvals"
    )

    # ==========================
    # STRICT ORDER ACTIONS
    # ==========================
    order_counts = get_strict_order_action_counts(user, role)

    badge_add(
        items,
        "order_actions",
        "Order Actions Pending",
        "dealer_order_list",
        order_counts["orders_link_count"],
        "order_actions"
    )

    badge_add(
        items,
        "warehouse_order_actions",
        "Warehouse Orders Pending",
        "warehouse_my_orders",
        order_counts["warehouse_link_count"],
        "order_actions"
    )

    # Show invoice badge only for actual accountant invoice review pending.
    badge_add(
        items,
        "invoice_review_actions",
        "Invoice Review Pending",
        "dealer_invoice_list",
        order_counts["invoice_link_count"],
        "order_actions"
    )

    # ==========================
    # DEALER REDEMPTION ACTIONS
    # ==========================
    redemption_count = 0

    if user.is_superuser or role in ["ADMIN", "ACCOUNTANT", "ASM", "SALES_OFFICER_SENIOR", "SALES_OFFICER_JUNIOR"]:
        DealerRedemptionRequest = badge_model("DealerRedemptionRequest")

        if DealerRedemptionRequest:
            try:
                redemption_count = DealerRedemptionRequest.objects.filter(
                    status="PENDING"
                ).count()
            except Exception:
                redemption_count = 0

    badge_add(
        items,
        "dealer_redemption_approvals",
        "Dealer Redemption Approvals",
        "dealer_redemption_approvals",
        redemption_count,
        "dealer_redemption_approvals"
    )

    # ==========================
    # ASSET MANAGER ACTIONS
    # ==========================
    asset_count = 0

    if user.is_superuser or role in ["ADMIN", "ASSET_MANAGER"]:
        AssetRequest = badge_model("AssetRequest")
        AssetReturnRequest = badge_model("AssetReturnRequest")

        if AssetRequest:
            try:
                asset_count += AssetRequest.objects.filter(status="PENDING").count()
            except Exception:
                pass

        if AssetReturnRequest:
            try:
                asset_count += AssetReturnRequest.objects.filter(status="PENDING").count()
            except Exception:
                pass

    badge_add(
        items,
        "asset_actions",
        "Asset Actions Pending",
        "asset_list",
        asset_count,
        "asset_actions"
    )

    # ==========================
    # TOTAL WITHOUT DUPLICATING SAME ACTION ON MULTIPLE LINKS
    # ==========================
    for item in items:
        total_key = item.get("total_key") or item["key"]
        unique_total_keys[total_key] = max(
            unique_total_keys.get(total_key, 0),
            item["count"]
        )

    total = sum(unique_total_keys.values())

    by_url = {}
    by_key = {}

    for item in items:
        by_url[item["url"]] = by_url.get(item["url"], 0) + item["count"]
        by_key[item["key"]] = by_key.get(item["key"], 0) + item["count"]

    return {
        "total": total,
        "items": items,
        "by_url": by_url,
        "by_key": by_key,
    }


@login_required
def pending_action_badges_api(request):
    data = get_pending_action_badges_for_user(request.user)
    return JsonResponse(data)

